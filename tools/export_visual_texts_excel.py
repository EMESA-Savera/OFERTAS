from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TRANSLATIONS_DIR = ROOT / "static" / "translations"
DEFAULT_OUTPUT = ROOT / "excel" / "textos_visuales_usuario.xlsx"

TRANSLATION_FILES = {
    "es": TRANSLATIONS_DIR / "es.json",
    "en": TRANSLATIONS_DIR / "en.json",
    "cs": TRANSLATIONS_DIR / "cs.json",
}

REFERENCE_PATTERNS = [
    re.compile(r'data-translate-key(?:-placeholder|-title|-aria-label|-value)?="([^"]+)"'),
    re.compile(r"\b(?:t|tf|translate)\(\s*'([^']+)'"),
    re.compile(r'\b(?:t|tf|translate)\(\s*"([^"]+)"'),
    re.compile(r'data-title-key="([^"]+)"'),
]

SCAN_FILES = [
    ROOT / "templates" / "index.html",
    ROOT / "templates" / "base.html",
    ROOT / "static" / "js" / "i18n.js",
    ROOT / "static" / "js" / "loginModal.js",
    ROOT / "static" / "js" / "navigationManager.js",
    ROOT / "static" / "js" / "ofertas.js",
    ROOT / "static" / "js" / "globalHeader.js",
]


@dataclass(frozen=True)
class LiteralRow:
    identifier: str
    source: str
    es: str
    en: str
    cs: str


MANUAL_LITERAL_ROWS = [
    LiteralRow("literal.breadcrumb", "templates/index.html", "Breadcrumb", "Breadcrumb", "Drobecková navigace"),
    LiteralRow("literal.main_navigation", "templates/index.html", "Navegación principal", "Main navigation", "Hlavní navigace"),
    LiteralRow("literal.sections", "templates/index.html", "Secciones", "Sections", "Sekce"),
    LiteralRow("literal.valid_email_title", "templates/index.html", "Indica al menos un correo electrónico válido", "Enter at least one valid email address", "Zadejte alespoň jednu platnou e-mailovou adresu"),
    LiteralRow("literal.etc.acronym", "templates/index.html", "ETC", "ETC", "ETC"),
    LiteralRow("literal.etc.step_title", "templates/index.html", "Paso 2: completar ETC", "Step 2: complete ETC", "Krok 2: dokončete ETC"),
    LiteralRow("literal.etc.saved_header_help", "templates/index.html", "La cabecera de la oferta ya está guardada. Completa ahora los datos ETC para finalizar el proceso.", "The quote header is already saved. Complete the ETC data now to finish the process.", "Hlavička nabídky je již uložena. Nyní doplňte údaje ETC a dokončete proces."),
    LiteralRow("literal.etc.responsible", "templates/index.html", "Responsable", "Responsible", "Odpovědná osoba"),
    LiteralRow("literal.etc.select_responsible", "templates/index.html", "Selecciona un responsable", "Select a responsible person", "Vyberte odpovědnou osobu"),
    LiteralRow("literal.etc.department_manager_help", "templates/index.html", "Si eliges un departamento, se propondrá automáticamente su manager.", "If you choose a department, its manager will be suggested automatically.", "Pokud vyberete oddělení, jeho vedoucí bude navržen automaticky."),
    LiteralRow("literal.etc.target_department", "templates/index.html", "Departamento destino", "Target department", "Cílové oddělení"),
    LiteralRow("literal.etc.no_department", "templates/index.html", "Sin departamento", "No department", "Bez oddělení"),
    LiteralRow("literal.etc.external_code", "templates/index.html", "Código externo", "External code", "Externí kód"),
    LiteralRow("literal.etc.client_reference", "templates/index.html", "Referencia cliente", "Client reference", "Reference klienta"),
    LiteralRow("literal.etc.incoterm", "templates/index.html", "Incoterm", "Incoterm", "Incoterm"),
    LiteralRow("literal.etc.extended_form", "templates/index.html", "Ext. formulario", "Extended form", "Rozšířený formulář"),
    LiteralRow("literal.etc.hide_extended_form", "static/js/ofertas.js", "Ocultar ext. formulario", "Hide extended form", "Skrýt rozšířený formulář"),
    LiteralRow("literal.etc.low_usage_help", "templates/index.html", "Muestra campos de menor uso detectados en el histórico del Excel.", "Shows lower-usage fields detected in the Excel history.", "Zobrazuje méně používaná pole zjištěná z historie v Excelu."),
    LiteralRow("literal.etc.priority", "templates/index.html", "Prioridad", "Priority", "Priorita"),
    LiteralRow("literal.etc.mark_urgent", "templates/index.html", "Marcar como urgente", "Mark as urgent", "Označit jako urgentní"),
    LiteralRow("literal.etc.internal_code", "templates/index.html", "Código interno", "Internal code", "Interní kód"),
    LiteralRow("literal.etc.commission_number", "templates/index.html", "Número comisión", "Commission number", "Číslo provize"),
    LiteralRow("literal.etc.project", "templates/index.html", "Proyecto", "Project", "Projekt"),
    LiteralRow("literal.etc.no_project", "templates/index.html", "Sin proyecto", "No project", "Bez projektu"),
    LiteralRow("literal.etc.po_original", "templates/index.html", "PO original", "Original PO", "Původní PO"),
    LiteralRow("literal.etc.b2b_order", "templates/index.html", "Pedido B2B", "B2B order", "B2B objednávka"),
    LiteralRow("literal.etc.offer_sent_date", "templates/index.html", "Fecha envío oferta", "Quote sent date", "Datum odeslání nabídky"),
    LiteralRow("literal.etc.requester_name", "templates/index.html", "Nombre solicitante", "Requester name", "Jméno žadatele"),
    LiteralRow("literal.etc.requester_email", "templates/index.html", "Email solicitante", "Requester email", "E-mail žadatele"),
    LiteralRow("literal.etc.total_material_eur", "templates/index.html", "Total material EUR", "Total material EUR", "Celkový materiál EUR"),
    LiteralRow("literal.etc.total_fee_eur", "templates/index.html", "Total fee EUR", "Total fee EUR", "Celkový poplatek EUR"),
    LiteralRow("literal.etc.material_summary", "templates/index.html", "Resumen material solicitado", "Requested material summary", "Souhrn požadovaného materiálu"),
    LiteralRow("literal.etc.client_notes", "templates/index.html", "Observaciones cliente", "Client notes", "Poznámky klienta"),
    LiteralRow("literal.etc.save_and_finish", "templates/index.html", "Guardar ETC y finalizar", "Save ETC and finish", "Uložit ETC a dokončit"),
    LiteralRow("literal.etc.no_users_in_department", "static/js/ofertas.js", "No hay usuarios configurados en este departamento", "There are no users configured in this department", "V tomto oddělení nejsou nakonfigurováni žádní uživatelé"),
    LiteralRow("literal.outlook.title", "templates/index.html", "Outlook", "Outlook", "Outlook"),
    LiteralRow("literal.bom.acronym", "templates/index.html", "BOM", "BOM", "BOM"),
    LiteralRow("literal.bom.title", "templates/index.html", "Materiales y precios", "Materials and prices", "Materiály a ceny"),
    LiteralRow("literal.bom.history_help", "templates/index.html", "Consulta y actualiza el histórico de precios de materiales.", "Review and update the material price history.", "Zobrazte a aktualizujte historii cen materiálů."),
    LiteralRow("literal.bom.search_material", "templates/index.html", "Buscar material", "Search material", "Hledat materiál"),
    LiteralRow("literal.bom.search_placeholder", "templates/index.html", "Escribe para filtrar materiales", "Type to filter materials", "Pište pro filtrování materiálů"),
    LiteralRow("literal.bom.material", "templates/index.html", "Material", "Material", "Materiál"),
    LiteralRow("literal.bom.current_price", "templates/index.html", "Precio actual", "Current price", "Aktuální cena"),
    LiteralRow("literal.bom.updated", "templates/index.html", "Actualizado", "Updated", "Aktualizováno"),
    LiteralRow("literal.bom.difference", "templates/index.html", "Diferencia", "Difference", "Rozdíl"),
    LiteralRow("literal.bom.loading_materials", "templates/index.html", "Cargando materiales...", "Loading materials...", "Načítání materiálů..."),
    LiteralRow("literal.bom.previous_price", "templates/index.html", "Precio anterior", "Previous price", "Předchozí cena"),
    LiteralRow("literal.bom.new_price", "templates/index.html", "Nuevo precio", "New price", "Nová cena"),
    LiteralRow("literal.bom.insert_history_help", "templates/index.html", "Se insertará un nuevo registro para mantener el histórico del material.", "A new record will be inserted to preserve the material history.", "Bude vložen nový záznam pro zachování historie materiálu."),
    LiteralRow("literal.bom.save_new_price", "templates/index.html", "Guardar nuevo precio", "Save new price", "Uložit novou cenu"),
    LiteralRow("literal.bom.back_to_list", "templates/index.html", "Volver al listado", "Back to list", "Zpět na seznam"),
    LiteralRow("literal.bom.offer_context", "static/js/ofertas.js", "Oferta {offerReference}. El cambio de precio inserta una nueva versión y conserva el histórico.", "Quote {offerReference}. The price change inserts a new version and preserves the history.", "Nabídka {offerReference}. Změna ceny vloží novou verzi a zachová historii."),
    LiteralRow("literal.bom.to", "static/js/ofertas.js", "Para:", "To:", "Komu:"),
    LiteralRow("literal.bom.no_subject", "static/js/ofertas.js", "(sin asunto)", "(no subject)", "(bez předmětu)"),
    LiteralRow("literal.states.sidebar_emoji", "templates/index.html", "Emoji del sidebar", "Sidebar emoji", "Emoji postranního panelu"),
    LiteralRow("literal.states.select_sidebar_emoji", "templates/index.html", "Seleccionar emoji para el sidebar", "Select emoji for the sidebar", "Vyberte emoji pro postranní panel"),
    LiteralRow("literal.states.emoji_help", "templates/index.html", "Se sugiere automáticamente según el nombre del estado, pero puedes cambiarlo.", "It is suggested automatically based on the state name, but you can change it.", "Navrhne se automaticky podle názvu stavu, ale můžete ho změnit."),
    LiteralRow("literal.states.department", "templates/index.html", "Departamento", "Department", "Oddělení"),
    LiteralRow("literal.states.edit_state", "templates/index.html", "Editar estado", "Edit state", "Upravit stav"),
    LiteralRow("literal.states.active", "templates/index.html", "Activo", "Active", "Aktivní"),
    LiteralRow("literal.users.title", "templates/index.html", "Usuarios", "Users", "Uživatelé"),
    LiteralRow("literal.users.description", "templates/index.html", "Alta de usuarios de OFERTAS a partir de General.Usuarios, con su rol y departamento.", "Create OFFERS users from General.Usuarios with their role and department.", "Zakládání uživatelů OFERTAS z General.Usuarios s jejich rolí a oddělením."),
    LiteralRow("literal.users.management", "templates/index.html", "Gestión de usuarios", "User management", "Správa uživatelů"),
    LiteralRow("literal.users.add_user", "templates/index.html", "Añadir usuario", "Add user", "Přidat uživatele"),
    LiteralRow("literal.users.view_users", "templates/index.html", "Ver usuarios", "View users", "Zobrazit uživatele"),
    LiteralRow("literal.users.operator_number", "templates/index.html", "Nº operario", "Operator no.", "Číslo operátora"),
    LiteralRow("literal.users.name", "templates/index.html", "Nombre", "Name", "Jméno"),
    LiteralRow("literal.users.email", "templates/index.html", "Email", "Email", "E-mail"),
    LiteralRow("literal.users.role", "templates/index.html", "Rol", "Role", "Role"),
    LiteralRow("literal.users.select_role", "templates/index.html", "Selecciona un rol", "Select a role", "Vyberte roli"),
    LiteralRow("literal.users.departments", "templates/index.html", "Departamentos", "Departments", "Oddělení"),
    LiteralRow("literal.users.add_department", "templates/index.html", "+ Departamento", "+ Department", "+ Oddělení"),
    LiteralRow("literal.users.save_user", "templates/index.html", "Guardar usuario", "Save user", "Uložit uživatele"),
    LiteralRow("literal.users.empty", "templates/index.html", "No hay usuarios creados todavía.", "There are no users yet.", "Zatím nejsou vytvořeni žádní uživatelé."),
    LiteralRow("literal.projects.title", "templates/index.html", "Proyectos", "Projects", "Projekty"),
    LiteralRow("literal.projects.description", "templates/index.html", "Alta y mantenimiento de proyectos disponibles para el formulario ETC.", "Create and maintain projects available for the ETC form.", "Zakládání a správa projektů dostupných pro formulář ETC."),
    LiteralRow("literal.projects.management", "templates/index.html", "Gestión de proyectos", "Project management", "Správa projektů"),
    LiteralRow("literal.projects.create_project", "templates/index.html", "Crear proyecto", "Create project", "Vytvořit projekt"),
    LiteralRow("literal.projects.view_projects", "templates/index.html", "Ver proyectos", "View projects", "Zobrazit projekty"),
    LiteralRow("literal.projects.project_name", "templates/index.html", "Nombre del proyecto", "Project name", "Název projektu"),
    LiteralRow("literal.projects.project_name_placeholder", "templates/index.html", "Introduce el nombre del proyecto", "Enter the project name", "Zadejte název projektu"),
    LiteralRow("literal.projects.save_project", "templates/index.html", "Guardar proyecto", "Save project", "Uložit projekt"),
    LiteralRow("literal.projects.project_description", "templates/index.html", "Descripción proyecto", "Project description", "Popis projektu"),
    LiteralRow("literal.projects.empty", "templates/index.html", "No hay proyectos creados todavía.", "There are no projects yet.", "Zatím nejsou vytvořeny žádné projekty."),
    LiteralRow("literal.table.id", "templates/index.html", "ID", "ID", "ID"),
    LiteralRow("literal.table.offer_number_duplicate", "templates/index.html", "Nº oferta", "Quote no.", "Číslo nabídky"),
    LiteralRow("literal.filters.users.operator_number", "static/js/ofertas.js", "Filtrar nº operario", "Filter operator no.", "Filtrovat číslo operátora"),
    LiteralRow("literal.filters.users.name", "static/js/ofertas.js", "Filtrar nombre", "Filter name", "Filtrovat jméno"),
    LiteralRow("literal.filters.users.role", "static/js/ofertas.js", "Filtrar rol", "Filter role", "Filtrovat roli"),
    LiteralRow("literal.filters.users.departments", "static/js/ofertas.js", "Filtrar departamentos", "Filter departments", "Filtrovat oddělení"),
    LiteralRow("literal.states.inactive_prefix", "static/js/ofertas.js", "[Inactivo] ", "[Inactive] ", "[Neaktivní] "),
    LiteralRow("literal.feedback.review_required_fields", "static/js/ofertas.js", "Revisa los campos obligatorios marcados en rojo.", "Review the required fields marked in red.", "Zkontrolujte povinná pole označená červeně."),
    LiteralRow("literal.feedback.header_ready", "static/js/ofertas.js", "Cabecera preparada. Completa ahora el ETC para guardar todo el proceso.", "Header ready. Complete the ETC now to save the whole process.", "Hlavička je připravena. Nyní doplňte ETC, aby se uložil celý proces."),
    LiteralRow("literal.feedback.prepare_etc_error", "static/js/ofertas.js", "Se produjo un error al preparar el formulario ETC.", "An error occurred while preparing the ETC form.", "Při přípravě formuláře ETC došlo k chybě."),
    LiteralRow("literal.feedback.missing_etc_identifier", "static/js/ofertas.js", "Indica al menos un identificador ETC: código externo, código interno, referencia cliente, número comisión o proyecto.", "Provide at least one ETC identifier: external code, internal code, client reference, commission number, or project.", "Zadejte alespoň jeden identifikátor ETC: externí kód, interní kód, referenci klienta, číslo provize nebo projekt."),
    LiteralRow("literal.feedback.save_header_first", "static/js/ofertas.js", "Primero debes guardar la cabecera de la oferta.", "You must save the quote header first.", "Nejprve musíte uložit hlavičku nabídky."),
    LiteralRow("literal.feedback.material_required", "static/js/ofertas.js", "El material es obligatorio.", "Material is required.", "Materiál je povinný."),
    LiteralRow("literal.feedback.new_price_required", "static/js/ofertas.js", "El nuevo precio es obligatorio.", "The new price is required.", "Nová cena je povinná."),
    LiteralRow("literal.feedback.bom_saved", "static/js/ofertas.js", "Precio BOM guardado correctamente.", "BOM price saved successfully.", "Cena BOM byla úspěšně uložena."),
    LiteralRow("literal.feedback.bom_save_error", "static/js/ofertas.js", "No se pudo guardar el nuevo precio BOM.", "Could not save the new BOM price.", "Novou cenu BOM se nepodařilo uložit."),
    LiteralRow("literal.feedback.bom_load_error", "static/js/ofertas.js", "No se pudieron cargar los materiales BOM.", "Could not load BOM materials.", "Materiály BOM se nepodařilo načíst."),
    LiteralRow("literal.feedback.loading_state", "static/js/ofertas.js", "Guardando...", "Saving...", "Ukládání..."),
    LiteralRow("literal.feedback.continuing_state", "static/js/ofertas.js", "Continuando...", "Continuing...", "Pokračování..."),
    LiteralRow("literal.feedback.accepting_state", "static/js/ofertas.js", "Aceptando...", "Accepting...", "Potvrzování..."),
]


def load_translations() -> dict[str, dict[str, str]]:
    loaded: dict[str, dict[str, str]] = {}
    for language, file_path in TRANSLATION_FILES.items():
        loaded[language] = json.loads(file_path.read_text(encoding="utf-8"))
    return loaded


def collect_translation_references() -> set[str]:
    keys: set[str] = set()
    for file_path in SCAN_FILES:
        if not file_path.exists():
            continue
        content = file_path.read_text(encoding="utf-8")
        for pattern in REFERENCE_PATTERNS:
            keys.update(match.group(1) for match in pattern.finditer(content))
    return keys


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)


def build_workbook(translations: dict[str, dict[str, str]]) -> Workbook:
    translation_keys = set().union(*(language_map.keys() for language_map in translations.values()))
    referenced_keys = collect_translation_references()
    export_keys = sorted(translation_keys | referenced_keys)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "textos_visuales"
    worksheet.append(["identificador", "origen", "tipo", "es", "en", "cs"])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for key in export_keys:
        worksheet.append([
            key,
            "static/translations/*.json",
            "translation_key",
            translations["es"].get(key, ""),
            translations["en"].get(key, ""),
            translations["cs"].get(key, ""),
        ])

    for row in MANUAL_LITERAL_ROWS:
        worksheet.append([row.identifier, row.source, "hardcoded_literal", row.es, row.en, row.cs])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)

    summary = workbook.create_sheet("resumen")
    summary.append(["metrica", "valor"])
    summary["A1"].font = summary["B1"].font = Font(bold=True)
    summary.append(["claves_traduccion", len(export_keys)])
    summary.append(["literales_hardcodeados", len(MANUAL_LITERAL_ROWS)])
    summary.append(["total_filas_datos", len(export_keys) + len(MANUAL_LITERAL_ROWS)])
    summary.append(["fuentes_escaneadas", len([path for path in SCAN_FILES if path.exists()])])
    autosize_columns(summary)

    return workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Exporta textos visuales del proyecto a un Excel multilingue.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Ruta de salida del archivo .xlsx")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    translations = load_translations()
    workbook = build_workbook(translations)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    output_path = args.output
    try:
        workbook.save(output_path)
    except PermissionError:
        output_path = output_path.with_name(f"{output_path.stem}_actualizado{output_path.suffix}")
        workbook.save(output_path)
    print(f"Excel generado en: {output_path}")


if __name__ == "__main__":
    main()
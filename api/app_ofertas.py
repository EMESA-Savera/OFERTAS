import base64
import csv
import hashlib
import importlib
import io
import json
import ipaddress
import mimetypes
import os
import re
import secrets
import shutil
import sys
import tempfile
import unicodedata
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from email.header import decode_header, make_header
from email import policy
from email.parser import BytesParser, Parser
from email.utils import parseaddr, parsedate_to_datetime
from html import escape as html_escape, unescape
from html.parser import HTMLParser
from urllib.parse import urlparse

import pyodbc
from dotenv import dotenv_values, load_dotenv
from flask import Flask, Response, has_request_context, jsonify, redirect, render_template, render_template_string, request, send_from_directory, session, url_for
from flask_cors import CORS
from flask_session import Session
from outlook_service import OutlookGraphError, OutlookGraphService
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename


def get_runtime_root():
    if getattr(sys, "frozen", False):
        return getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def load_runtime_env(project_dir):
    candidate_paths = []
    is_frozen = getattr(sys, "frozen", False)

    if is_frozen:
        executable_dir = os.path.dirname(sys.executable)
        production_paths = [
            os.path.join(executable_dir, ".env.production"),
            os.path.join(project_dir, ".env.production"),
        ]
        found_production_env = any(os.path.exists(path) for path in production_paths)
        candidate_paths.extend(production_paths)
        if not found_production_env:
            candidate_paths.append(os.path.join(executable_dir, ".env"))
    else:
        default_env_path = os.path.join(project_dir, ".env")
        if os.path.exists(default_env_path):
            load_dotenv(default_env_path, override=True)

        effective_env = (
            os.getenv("APP_ENV")
            or os.getenv("FLASK_ENV")
            or os.getenv("ENV")
            or ""
        ).strip().lower()
        if effective_env == "production":
            candidate_paths.append(os.path.join(project_dir, ".env.production"))

    seen = set()
    for env_path in candidate_paths:
        normalized = os.path.normpath(env_path)
        if normalized in seen:
            continue
        seen.add(normalized)
        if os.path.exists(env_path):
            load_dotenv(env_path, override=True)


PROJECT_DIR = get_runtime_root()
load_runtime_env(PROJECT_DIR)


def get_env_value(*names, default=None):
    for name in names:
        value = os.getenv(name)
        if value is not None and str(value).strip() != "":
            return value
    return default


def get_env_flag(*names, default=False):
    value = get_env_value(*names)
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def resolve_storage_dir(env_name, default_path):
    configured_path = get_env_value(env_name, default=None)
    if configured_path is None:
        return default_path

    normalized_path = os.path.expandvars(os.path.expanduser(str(configured_path).strip()))
    if not normalized_path:
        return default_path

    if os.path.isabs(normalized_path):
        return normalized_path

    return os.path.normpath(os.path.join(PROJECT_DIR, normalized_path))


class Config:
    SECRET_KEY = get_env_value("SESSION_SECRET", "SECRET_KEY", default="ofertas-dev-secret-key")
    APP_ENV = get_env_value("APP_ENV", "FLASK_ENV", default="development").lower()
    DEBUG = APP_ENV == "development"
    APP_PORT = int(get_env_value("PORT", "APP_PORT", default="3010"))
    AUTO_GENERATE_SSL_CERT = get_env_flag("AUTO_GENERATE_SSL_CERT", default=True)
    SSL_CERT_FILE = get_env_value("SSL_CERT_FILE", default="cert.pem")
    SSL_KEY_FILE = get_env_value("SSL_KEY_FILE", default="key.pem")

    DB_SERVER = get_env_value("DB_SERVER", "SQL_SERVER", "DATABASE_HOST", "ODBC_SERVER", default="EMEBIDWH")
    DB_DATABASE = get_env_value("DB_DATABASE", "SQL_DATABASE", "DATABASE_NAME", default="Digitalizacion")
    DB_USER = get_env_value("DB_USER", "SQL_USER", "DATABASE_USER", default="sa")
    DB_PASSWORD = get_env_value("DB_PASSWORD", "SQL_PASSWORD", "DATABASE_PASSWORD", default="")
    DB_DRIVER = get_env_value("DB_DRIVER", "SQL_DRIVER", "ODBC_DRIVER", default="ODBC Driver 18 for SQL Server")

    SESSION_COOKIE_NAME = get_env_value("SESSION_COOKIE_NAME", default="ofertas_session")
    SESSION_COOKIE_HTTPONLY = True
    SESSION_COOKIE_SAMESITE = "Lax"
    SESSION_COOKIE_SECURE = get_env_value("SESSION_COOKIE_SECURE", default="false").lower() == "true"
    SESSION_TYPE = "filesystem"
    SESSION_FILE_DIR = os.path.join(PROJECT_DIR, ".flask_session")
    SESSION_FILE_THRESHOLD = 500
    SESSION_USE_SIGNER = True
    SESSION_PERMANENT = True
    PERMANENT_SESSION_LIFETIME = timedelta(hours=12)


app = Flask(
    __name__,
    template_folder=os.path.join(PROJECT_DIR, "templates"),
    static_folder=os.path.join(PROJECT_DIR, "static"),
)
app.config.from_object(Config)
app.secret_key = app.config["SECRET_KEY"]
CORS(app, supports_credentials=True)
os.makedirs(app.config["SESSION_FILE_DIR"], exist_ok=True)
Session(app)

DEFAULT_RUNTIME_DATA_DIR = os.path.join(os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else PROJECT_DIR, "data")
RUNTIME_DATA_DIR = resolve_storage_dir("RUNTIME_DATA_DIR", DEFAULT_RUNTIME_DATA_DIR)
OFFER_ATTACHMENTS_DIR = resolve_storage_dir("OFFER_ATTACHMENTS_DIR", os.path.join(RUNTIME_DATA_DIR, "offer_attachments"))
IMPORTED_EMAIL_ATTACHMENTS_DIR = resolve_storage_dir("IMPORTED_EMAIL_ATTACHMENTS_DIR", os.path.join(RUNTIME_DATA_DIR, "imported_email_attachments"))
INTERNAL_CHAT_ENABLED = get_env_flag("ENABLE_INTERNAL_CHAT", default=False)
OFFER_CHAT_DIR = os.path.join(RUNTIME_DATA_DIR, "offer_chat")
OFFER_CHAT_READ_STATE_DIR = os.path.join(RUNTIME_DATA_DIR, "offer_chat_read_state")
MAX_OFFER_ATTACHMENT_SIZE_BYTES = 25 * 1024 * 1024
ALLOWED_OFFER_ATTACHMENT_EXTENSIONS = {
    ".pdf", ".csv", ".xls", ".xlsx", ".xlsm", ".ods",
    ".doc", ".docx", ".txt", ".rtf",
    ".png", ".jpg", ".jpeg", ".webp", ".gif",
    ".zip", ".7z", ".rar",
    ".eml", ".msg",
    ".dwg", ".dxf",
}
os.makedirs(OFFER_ATTACHMENTS_DIR, exist_ok=True)
os.makedirs(IMPORTED_EMAIL_ATTACHMENTS_DIR, exist_ok=True)
os.makedirs(OFFER_CHAT_DIR, exist_ok=True)
os.makedirs(OFFER_CHAT_READ_STATE_DIR, exist_ok=True)

IMAGE_ATTACHMENT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
TEXT_ATTACHMENT_EXTENSIONS = {".txt", ".csv", ".eml"}
INLINE_ATTACHMENT_EXTENSIONS = {".pdf"}


class DuplicateOfertaError(ValueError):
    pass


def normalize_offer_attachment_directory_name(numero_oferta, oferta_id):
    raw_value = str(numero_oferta or "").strip() or str(int(oferta_id))
    normalized_value = re.sub(r"[^a-zA-Z0-9_-]", "_", raw_value).strip("_")
    return normalized_value or str(int(oferta_id))


def get_offer_numero_by_id(oferta_id):
    with db_connection(autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT numero_oferta FROM ofertas.listado_ofertas WHERE id_oferta = ?",
            (int(oferta_id),),
        )
        row = cursor.fetchone()
    return row[0] if row else None


def get_offer_attachments_dir(oferta_id, create=False, numero_oferta=None):
    normalized_dir_name = normalize_offer_attachment_directory_name(numero_oferta or get_offer_numero_by_id(oferta_id), oferta_id)
    directory = os.path.join(OFFER_ATTACHMENTS_DIR, normalized_dir_name)
    legacy_directory = os.path.join(OFFER_ATTACHMENTS_DIR, str(int(oferta_id)))

    if os.path.normcase(legacy_directory) != os.path.normcase(directory) and os.path.isdir(legacy_directory):
        os.makedirs(os.path.dirname(directory), exist_ok=True)
        if not os.path.exists(directory):
            shutil.move(legacy_directory, directory)

    if create:
        os.makedirs(directory, exist_ok=True)
    return directory


def delete_offer_attachment_storage(oferta_id, numero_oferta=None):
    current_directory = get_offer_attachments_dir(oferta_id, numero_oferta=numero_oferta)
    legacy_directory = os.path.join(OFFER_ATTACHMENTS_DIR, str(int(oferta_id)))
    imported_bucket_directory = os.path.join(
        IMPORTED_EMAIL_ATTACHMENTS_DIR,
        normalize_imported_email_attachment_bucket(numero_oferta),
    )

    directories_to_remove = [current_directory, imported_bucket_directory]
    if os.path.normcase(legacy_directory) != os.path.normcase(current_directory):
        directories_to_remove.append(legacy_directory)

    for directory in directories_to_remove:
        try:
            shutil.rmtree(directory, ignore_errors=True)
        except Exception:
            app.logger.warning("No se pudo eliminar el almacenamiento de la oferta: %s", directory, exc_info=True)


def get_offer_chat_file_path(oferta_id):
    return os.path.join(OFFER_CHAT_DIR, f"{int(oferta_id)}.json")


def get_offer_chat_read_state_path(oferta_id):
    return os.path.join(OFFER_CHAT_READ_STATE_DIR, f"{int(oferta_id)}.json")


def load_offer_chat_messages(oferta_id):
    chat_path = get_offer_chat_file_path(oferta_id)
    if not os.path.exists(chat_path):
        return []

    try:
        with open(chat_path, "r", encoding="utf-8") as chat_file:
            payload = json.load(chat_file) or []
            return payload if isinstance(payload, list) else []
    except Exception:
        app.logger.warning("No se pudo leer el chat de la oferta %s", oferta_id, exc_info=True)
        return []


def save_offer_chat_messages(oferta_id, messages):
    chat_path = get_offer_chat_file_path(oferta_id)
    with open(chat_path, "w", encoding="utf-8") as chat_file:
        json.dump(messages or [], chat_file, ensure_ascii=False)


def load_offer_chat_read_state(oferta_id):
    state_path = get_offer_chat_read_state_path(oferta_id)
    if not os.path.exists(state_path):
        return {}

    try:
        with open(state_path, "r", encoding="utf-8") as state_file:
            payload = json.load(state_file) or {}
            return payload if isinstance(payload, dict) else {}
    except Exception:
        app.logger.warning("No se pudo leer el estado de lectura del chat de la oferta %s", oferta_id, exc_info=True)
        return {}


def save_offer_chat_read_state(oferta_id, state):
    state_path = get_offer_chat_read_state_path(oferta_id)
    with open(state_path, "w", encoding="utf-8") as state_file:
        json.dump(state or {}, state_file, ensure_ascii=False)


def get_offer_chat_user_key(user_data):
    num_operario = normalize_optional_text((user_data or {}).get("num_operario"), 50)
    if num_operario:
        return f"operario:{num_operario}"

    email = normalize_optional_text((user_data or {}).get("email"), 255)
    if email:
        return f"email:{email.lower()}"

    return None


def parse_offer_chat_timestamp(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)

    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        return None


def is_offer_chat_message_from_user(message, user_data):
    if not isinstance(message, dict):
        return False

    user_operario = normalize_optional_text((user_data or {}).get("num_operario"), 50)
    message_operario = normalize_optional_text(message.get("author_operario"), 50)
    if user_operario and message_operario and user_operario == message_operario:
        return True

    user_email = normalize_optional_text((user_data or {}).get("email"), 255)
    message_email = normalize_optional_text(message.get("author_email"), 255)
    if user_email and message_email and user_email.lower() == message_email.lower():
        return True

    return False


def get_offer_chat_unread_count(oferta_id, user_data, messages=None):
    user_key = get_offer_chat_user_key(user_data)
    if not user_key:
        return 0

    chat_messages = messages if messages is not None else load_offer_chat_messages(oferta_id)
    read_state = load_offer_chat_read_state(oferta_id)
    persisted_last_read_at = read_state.get(user_key)
    last_read_at = parse_offer_chat_timestamp(persisted_last_read_at)

    if persisted_last_read_at is None:
        latest_message_timestamp = None
        for message in reversed(chat_messages):
            latest_message_timestamp = parse_offer_chat_timestamp(message.get("created_at"))
            if latest_message_timestamp is not None:
                break

        bootstrap_timestamp = latest_message_timestamp or datetime.now(timezone.utc)
        read_state[user_key] = bootstrap_timestamp.isoformat().replace("+00:00", "Z")
        save_offer_chat_read_state(oferta_id, read_state)
        return 0

    unread_count = 0

    for message in chat_messages:
        if is_offer_chat_message_from_user(message, user_data):
            continue

        message_created_at = parse_offer_chat_timestamp(message.get("created_at"))
        if last_read_at is None:
            unread_count += 1
            continue

        if message_created_at and message_created_at > last_read_at:
            unread_count += 1

    return unread_count


def build_offer_chat_summary(oferta_id, user_data, messages=None):
    if not INTERNAL_CHAT_ENABLED:
        return {
            "chat_unread_count": 0,
            "chat_has_unread": False,
        }

    unread_count = get_offer_chat_unread_count(oferta_id, user_data, messages=messages)
    return {
        "chat_unread_count": unread_count,
        "chat_has_unread": unread_count > 0,
    }


def mark_offer_chat_as_read(oferta_id, user_data, messages=None):
    if not INTERNAL_CHAT_ENABLED:
        return build_offer_chat_summary(oferta_id, user_data, messages=messages)

    user_key = get_offer_chat_user_key(user_data)
    chat_messages = messages if messages is not None else load_offer_chat_messages(oferta_id)
    if not user_key:
        return build_offer_chat_summary(oferta_id, user_data, messages=chat_messages)

    read_state = load_offer_chat_read_state(oferta_id)
    read_timestamp = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    read_state[user_key] = read_timestamp
    save_offer_chat_read_state(oferta_id, read_state)

    return {
        "last_read_at": read_timestamp,
        **build_offer_chat_summary(oferta_id, user_data, messages=chat_messages),
    }


def build_offer_chat_author(user_data):
    author_name = normalize_optional_text(user_data.get("nombre"), 255) or "Usuario"
    author_email = normalize_optional_text(user_data.get("email"), 255)
    author_operario = normalize_optional_text(user_data.get("num_operario"), 50)
    return {
        "author_name": author_name,
        "author_email": author_email,
        "author_operario": author_operario,
    }


def append_offer_chat_message(oferta_id, user_data, message_text):
    if not INTERNAL_CHAT_ENABLED:
        raise RuntimeError("El chat interno está desactivado")

    normalized_message = normalize_required_text(message_text, "Mensaje", 2000)
    existing_messages = load_offer_chat_messages(oferta_id)
    author = build_offer_chat_author(user_data or {})
    message_entry = {
        "id": secrets.token_urlsafe(9),
        "message": normalized_message,
        "created_at": datetime.utcnow().isoformat() + "Z",
        **author,
    }
    existing_messages.append(message_entry)
    save_offer_chat_messages(oferta_id, existing_messages)
    return message_entry


def normalize_imported_email_attachment_bucket(bucket_name):
    normalized_bucket = re.sub(r"[^a-zA-Z0-9_-]", "_", str(bucket_name or "").strip())
    normalized_bucket = normalized_bucket.strip("_")
    return normalized_bucket or "pendientes"


def get_imported_email_attachments_dir(token, create=False, bucket_name=None):
    normalized_token = re.sub(r"[^a-zA-Z0-9_-]", "", str(token or "").strip())
    if not normalized_token:
        raise ValueError("Token de adjuntos importados no válido")

    normalized_bucket = normalize_imported_email_attachment_bucket(bucket_name)
    directory = os.path.join(IMPORTED_EMAIL_ATTACHMENTS_DIR, normalized_bucket, normalized_token)
    if create:
        os.makedirs(directory, exist_ok=True)
    return directory


def cleanup_imported_email_attachment_bucket(directory):
    parent_dir = os.path.dirname(directory)
    if not parent_dir or os.path.normcase(parent_dir) == os.path.normcase(IMPORTED_EMAIL_ATTACHMENTS_DIR):
        return

    try:
        if os.path.isdir(parent_dir) and not os.listdir(parent_dir):
            os.rmdir(parent_dir)
    except OSError:
        pass


def find_imported_email_attachments_dir(token):
    pending_dir = get_imported_email_attachments_dir(token)
    if os.path.isdir(pending_dir):
        return pending_dir

    try:
        bucket_entries = os.listdir(IMPORTED_EMAIL_ATTACHMENTS_DIR)
    except OSError:
        return pending_dir

    for bucket_name in bucket_entries:
        candidate_dir = get_imported_email_attachments_dir(token, bucket_name=bucket_name)
        if os.path.isdir(candidate_dir):
            return candidate_dir

    return pending_dir


def get_offer_attachment_metadata_path(file_path):
    return f"{file_path}.meta.json"


def load_offer_attachment_metadata(file_path):
    metadata_path = get_offer_attachment_metadata_path(file_path)
    if not os.path.exists(metadata_path):
        return {}

    try:
        with open(metadata_path, "r", encoding="utf-8") as metadata_file:
            return json.load(metadata_file) or {}
    except Exception:
        app.logger.warning("No se pudieron leer los metadatos del adjunto: %s", metadata_path, exc_info=True)
        return {}


def save_offer_attachment_metadata(file_path, metadata):
    metadata_path = get_offer_attachment_metadata_path(file_path)
    with open(metadata_path, "w", encoding="utf-8") as metadata_file:
        json.dump(metadata or {}, metadata_file, ensure_ascii=False)


def save_binary_attachment_to_dir(target_dir, original_name, content_bytes):
    normalized_name = os.path.basename(str(original_name or "").strip())
    if not normalized_name:
        raise ValueError("Debes seleccionar al menos un archivo.")

    if not is_allowed_offer_attachment(normalized_name):
        raise ValueError("Formato de archivo no permitido. Usa PDF, Excel, CSV, Word, imágenes, ZIP o correo.")

    binary_content = bytes(content_bytes or b"")
    if not binary_content:
        raise ValueError(f"El archivo adjunto '{normalized_name}' está vacío.")

    if len(binary_content) > MAX_OFFER_ATTACHMENT_SIZE_BYTES:
        raise ValueError(f"El archivo '{normalized_name}' supera el máximo de 25 MB.")

    safe_name = secure_filename(normalized_name)
    if not safe_name:
        raise ValueError("El nombre del archivo no es válido.")

    os.makedirs(target_dir, exist_ok=True)
    stored_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{safe_name}"
    file_path = os.path.join(target_dir, stored_name)
    with open(file_path, "wb") as output_file:
        output_file.write(binary_content)
    save_offer_attachment_metadata(file_path, {"original_name": normalized_name})
    return stored_name, file_path


def get_uploaded_file_size(file_storage):
    content_length = getattr(file_storage, "content_length", None)
    if isinstance(content_length, int) and content_length >= 0:
        return content_length

    stream = getattr(file_storage, "stream", None)
    if stream is None:
        return None

    current_position = None
    try:
        current_position = stream.tell()
        stream.seek(0, os.SEEK_END)
        size = stream.tell()
        stream.seek(current_position)
        return size
    except Exception:
        try:
            if current_position is not None:
                stream.seek(current_position)
        except Exception:
            pass
        return None


def is_allowed_offer_attachment(filename):
    extension = os.path.splitext(str(filename or ""))[1].lower()
    return bool(extension) and extension in ALLOWED_OFFER_ATTACHMENT_EXTENSIONS


def get_offer_attachment_extension(filename):
    return os.path.splitext(str(filename or ""))[1].lower()


def get_offer_attachment_preview_kind(filename):
    extension = get_offer_attachment_extension(filename)
    if extension in IMAGE_ATTACHMENT_EXTENSIONS:
        return "image"
    if extension in TEXT_ATTACHMENT_EXTENSIONS:
        return "text"
    if extension in INLINE_ATTACHMENT_EXTENSIONS:
        return "inline"
    return "unsupported"


def get_offer_attachment_mimetype(filename):
    extension = get_offer_attachment_extension(filename)
    explicit_map = {
        ".csv": "text/csv; charset=utf-8",
        ".eml": "text/plain; charset=utf-8",
        ".txt": "text/plain; charset=utf-8",
        ".pdf": "application/pdf",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".gif": "image/gif",
        ".doc": "application/msword",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xls": "application/vnd.ms-excel",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
        ".ods": "application/vnd.oasis.opendocument.spreadsheet",
        ".rtf": "application/rtf",
        ".zip": "application/zip",
        ".7z": "application/x-7z-compressed",
        ".rar": "application/vnd.rar",
        ".msg": "application/vnd.ms-outlook",
    }
    if extension in explicit_map:
        return explicit_map[extension]

    guessed_type, _ = mimetypes.guess_type(str(filename or ""))
    return guessed_type or "application/octet-stream"


def build_offer_attachment_payload(oferta_id, stored_name, numero_oferta=None):
    attachments_dir = get_offer_attachments_dir(oferta_id, numero_oferta=numero_oferta)
    file_path = os.path.join(attachments_dir, stored_name)
    if not os.path.isfile(file_path):
        return None

    metadata = load_offer_attachment_metadata(file_path)
    stats = os.stat(file_path)
    mime_type = get_offer_attachment_mimetype(metadata.get("original_name") or stored_name)
    preview_kind = get_offer_attachment_preview_kind(metadata.get("original_name") or stored_name)
    return {
        "stored_name": stored_name,
        "original_name": metadata.get("original_name") or stored_name,
        "size_bytes": stats.st_size,
        "updated_at": datetime.utcfromtimestamp(stats.st_mtime).isoformat() + "Z",
        "mime_type": mime_type,
        "preview_kind": preview_kind,
        "can_preview": preview_kind != "unsupported",
        "download_url": url_for("download_offer_attachment", oferta_id=int(oferta_id), filename=stored_name),
        "preview_url": url_for("preview_offer_attachment", oferta_id=int(oferta_id), filename=stored_name),
    }


def list_offer_attachments(oferta_id, numero_oferta=None):
    attachments_dir = get_offer_attachments_dir(oferta_id, numero_oferta=numero_oferta)
    if not os.path.isdir(attachments_dir):
        return []

    attachments = []
    for entry in os.listdir(attachments_dir):
        if entry.endswith(".meta.json"):
            continue
        attachment = build_offer_attachment_payload(oferta_id, entry, numero_oferta=numero_oferta)
        if attachment:
            attachments.append(attachment)

    attachments.sort(key=lambda item: item.get("updated_at") or "", reverse=True)
    return attachments


def ensure_offer_exists(cursor, oferta_id):
    cursor.execute("SELECT 1 FROM ofertas.listado_ofertas WHERE id_oferta = ?", (oferta_id,))
    return cursor.fetchone() is not None


def save_offer_attachment(oferta_id, file_storage):
    original_name = os.path.basename(str(getattr(file_storage, "filename", "") or "").strip())
    if not original_name:
        raise ValueError("Debes seleccionar al menos un archivo.")

    if not is_allowed_offer_attachment(original_name):
        raise ValueError("Formato de archivo no permitido. Usa PDF, Excel, CSV, Word, imágenes, ZIP o correo.")

    file_size = get_uploaded_file_size(file_storage)
    if file_size is not None and file_size > MAX_OFFER_ATTACHMENT_SIZE_BYTES:
        raise ValueError("Cada archivo puede ocupar como máximo 25 MB.")

    attachments_dir = get_offer_attachments_dir(oferta_id, create=True)
    stored_name, file_path = save_binary_attachment_to_dir(attachments_dir, original_name, file_storage.read())

    stream = getattr(file_storage, "stream", None)
    if stream is not None:
        try:
            stream.seek(0)
        except Exception:
            pass

    return build_offer_attachment_payload(oferta_id, stored_name)


def stage_imported_email_attachments(parsed_email, bucket_name=None):
    attachments = list((parsed_email or {}).get("attachments") or [])
    if not attachments:
        return None

    token = secrets.token_urlsafe(18)
    target_dir = get_imported_email_attachments_dir(token, create=True, bucket_name=bucket_name)
    staged_attachments = []
    try:
        for attachment in attachments:
            stored_name, file_path = save_binary_attachment_to_dir(
                target_dir,
                attachment.get("filename"),
                attachment.get("content_bytes"),
            )
            file_stat = os.stat(file_path)
            staged_attachments.append(
                {
                    "stored_name": stored_name,
                    "original_name": attachment.get("filename") or stored_name,
                    "size_bytes": file_stat.st_size,
                }
            )
    except Exception:
        shutil.rmtree(target_dir, ignore_errors=True)
        raise

    return {"token": token, "attachments": staged_attachments}


def move_staged_imported_email_attachments_to_offer(oferta_id, token, numero_oferta=None):
    source_dir = find_imported_email_attachments_dir(token)
    if not os.path.isdir(source_dir):
        return []

    target_dir = get_offer_attachments_dir(oferta_id, create=True, numero_oferta=numero_oferta)
    moved_names = []
    for entry in os.listdir(source_dir):
        if entry.endswith(".meta.json"):
            continue

        source_file = os.path.join(source_dir, entry)
        target_file = os.path.join(target_dir, entry)
        if not os.path.isfile(source_file):
            continue

        shutil.move(source_file, target_file)
        source_meta = get_offer_attachment_metadata_path(source_file)
        if os.path.exists(source_meta):
            shutil.move(source_meta, get_offer_attachment_metadata_path(target_file))
        moved_names.append(entry)

    shutil.rmtree(source_dir, ignore_errors=True)
    payloads = []
    for stored_name in moved_names:
        attachment_payload = build_offer_attachment_payload(oferta_id, stored_name, numero_oferta=numero_oferta)
        if attachment_payload:
            payloads.append(attachment_payload)
    cleanup_imported_email_attachment_bucket(source_dir)
    return payloads


def cleanup_offer_attachment_entries(oferta_id, stored_names):
    if not stored_names:
        return

    attachments_dir = get_offer_attachments_dir(oferta_id)
    for stored_name in stored_names:
        file_path = os.path.join(attachments_dir, stored_name)
        meta_path = get_offer_attachment_metadata_path(file_path)
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            app.logger.warning("No se pudo eliminar el adjunto temporal de la oferta: %s", file_path, exc_info=True)
        try:
            if os.path.exists(meta_path):
                os.remove(meta_path)
        except Exception:
            app.logger.warning("No se pudo eliminar el metadata temporal del adjunto: %s", meta_path, exc_info=True)


def get_configured_oauth_redirect_uri():
    apply_local_oauth_env_overrides()
    redirect_uri = get_env_value("OAUTH_REDIRECT_URI", "AZURE_REDIRECT_URI", "MICROSOFT_REDIRECT_URI", "OUTLOOK_REDIRECT_URI", default="")
    if not redirect_uri:
        return None
    parsed = urlparse(redirect_uri)
    if not parsed.scheme or not parsed.hostname:
        return None
    return parsed


def is_local_or_private_host(hostname):
    normalized = (hostname or "").strip().lower()
    if not normalized:
        return False
    if normalized in {"localhost", "127.0.0.1"}:
        return True

    try:
        parsed_ip = ipaddress.ip_address(normalized)
    except ValueError:
        return False

    return parsed_ip.is_loopback or parsed_ip.is_private


_LOCAL_ENV_CACHE = None


def load_local_env_values():
    global _LOCAL_ENV_CACHE
    if _LOCAL_ENV_CACHE is not None:
        return _LOCAL_ENV_CACHE

    local_env_path = os.path.join(PROJECT_DIR, ".env")
    if not os.path.exists(local_env_path):
        _LOCAL_ENV_CACHE = {}
        return _LOCAL_ENV_CACHE

    try:
        _LOCAL_ENV_CACHE = {
            str(key): str(value).strip()
            for key, value in dotenv_values(local_env_path).items()
            if key and value is not None and str(value).strip() != ""
        }
    except Exception:
        _LOCAL_ENV_CACHE = {}

    return _LOCAL_ENV_CACHE


def should_prefer_local_oauth_env():
    if getattr(sys, "frozen", False):
        return False
    if not has_request_context():
        return False

    # En producción no sobrescribir con valores de .env (desarrollo)
    if app.config.get("APP_ENV") == "production":
        return False

    current_host = (request.host.split(":", 1)[0] or "").strip().lower()
    return is_local_or_private_host(current_host)


def apply_local_oauth_env_overrides():
    if not should_prefer_local_oauth_env():
        return

    local_values = load_local_env_values()
    if not local_values:
        return

    for key, value in local_values.items():
        normalized_key = str(key).strip().upper()
        if (
            normalized_key.startswith(("OAUTH_", "AZURE_", "MICROSOFT_", "OUTLOOK_"))
            or normalized_key in {"CLIENT_ID", "CLIENT_SECRET", "TENANT_ID"}
        ):
            os.environ[normalized_key] = str(value)


def get_request_oauth_redirect_uri():
    apply_local_oauth_env_overrides()
    configured_redirect = get_configured_oauth_redirect_uri()
    if configured_redirect is None:
        return None

    if app.config.get("APP_ENV") == "production" and not should_prefer_local_oauth_env():
        return configured_redirect.geturl()

    current_host = (request.host.split(":", 1)[0] or "").strip().lower()
    configured_host = (configured_redirect.hostname or "").strip().lower()
    if not current_host:
        return configured_redirect.geturl()

    if not (is_local_or_private_host(current_host) and is_local_or_private_host(configured_host)):
        return configured_redirect.geturl()

    target_scheme = configured_redirect.scheme or request.scheme
    target_port = request.host.split(":", 1)[1].strip() if ":" in request.host else None
    if not target_port:
        configured_port = configured_redirect.port
        target_port = str(configured_port or (443 if target_scheme == "https" else 80))

    host_port = current_host
    if not ((target_scheme == "https" and target_port == "443") or (target_scheme == "http" and target_port == "80")):
        host_port = f"{host_port}:{target_port}"

    return f"{target_scheme}://{host_port}{configured_redirect.path}"


@app.before_request
def normalize_local_oauth_host():
    if request.method not in {"GET", "HEAD"}:
        return None

    if app.config.get("APP_ENV") == "production":
        return None

    configured_redirect = get_configured_oauth_redirect_uri()
    if not configured_redirect or not request.host:
        return None

    current_host = (request.host.split(":", 1)[0] or "").strip().lower()
    configured_host = (configured_redirect.hostname or "").strip().lower()
    local_aliases = {"127.0.0.1", "localhost"}

    if current_host == configured_host or {current_host, configured_host} - local_aliases:
        return None

    target_scheme = configured_redirect.scheme or request.scheme
    target_port = configured_redirect.port or (443 if target_scheme == "https" else 80)
    host_port = configured_host
    if not ((target_scheme == "https" and target_port == 443) or (target_scheme == "http" and target_port == 80)):
        host_port = f"{configured_host}:{target_port}"

    query_string = request.query_string.decode("utf-8") if request.query_string else ""
    target_url = f"{target_scheme}://{host_port}{request.path}"
    if query_string:
        target_url = f"{target_url}?{query_string}"
    return redirect(target_url, code=302)


READ_ONLY_OPERARIO_NUMBERS = {4}
GENERAL_USER_MICROSOFT_IDENTITY_COLUMNS = ("Email", "Correo", "Mail", "UPN", "Usuario", "Login")


@contextmanager
def db_connection(autocommit=False):
    conn = None
    last_error = None
    drivers_to_try = []

    if not app.config.get("DB_SERVER"):
        raise RuntimeError("Falta configurar DB_SERVER/SQL_SERVER en el archivo .env")

    if not app.config.get("DB_DATABASE"):
        raise RuntimeError("Falta configurar DB_DATABASE/SQL_DATABASE en el archivo .env")

    if not app.config.get("DB_USER"):
        raise RuntimeError("Falta configurar DB_USER/SQL_USER en el archivo .env")

    if not app.config.get("DB_PASSWORD"):
        raise RuntimeError("Falta configurar DB_PASSWORD/SQL_PASSWORD en el archivo .env")

    preferred_driver = app.config.get("DB_DRIVER")
    if preferred_driver:
        drivers_to_try.append(preferred_driver)

    for fallback_driver in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"):
        if fallback_driver not in drivers_to_try:
            drivers_to_try.append(fallback_driver)

    for driver in drivers_to_try:
        try:
            conn_str = (
                f"DRIVER={{{driver}}};"
                f"SERVER={app.config['DB_SERVER']};"
                f"DATABASE={app.config['DB_DATABASE']};"
                f"UID={app.config['DB_USER']};"
                f"PWD={app.config['DB_PASSWORD']};"
                "Encrypt=yes;"
                "TrustServerCertificate=yes;"
            )
            conn = pyodbc.connect(conn_str, timeout=8, autocommit=autocommit)
            break
        except Exception as exc:
            last_error = exc
            conn = None

    if conn is None:
        raise last_error or RuntimeError("No se pudo abrir la conexión ODBC")

    try:
        yield conn
    finally:
        conn.close()


def verify_password(password, hashed_password):
    if not hashed_password:
        return False

    try:
        if hashed_password.startswith("scrypt:") or hashed_password.startswith("pbkdf2:"):
            return check_password_hash(hashed_password, password)

        if hashed_password.startswith("sha256$"):
            parts = hashed_password.split("$")
            if len(parts) == 3:
                salt = parts[1]
                stored_hash = parts[2]
                calc = hashlib.sha256((password + salt).encode("utf-8")).hexdigest()
                return calc == stored_hash

        return password == hashed_password
    except Exception:
        return False


def get_logged_user_data():
    return session.get("user_data") or {}


def get_user_config(cursor, num_operario, fallback_role=None):
    cursor.execute(
        """
        SELECT TOP 1
            uc.email,
            uc.id_rol,
            r.nombre_rol,
            uc.id_departamento,
            d.nombre_departamento,
            d.descripcion,
            d.estado_activo
        FROM ofertas.usuarios_config uc
        LEFT JOIN ofertas.roles r
            ON r.id_rol = uc.id_rol
        LEFT JOIN ofertas.departamentos d
            ON d.id_departamento = uc.id_departamento
        WHERE uc.num_operario = ?
        """,
        (num_operario,),
    )
    row = cursor.fetchone()

    fallback_role_id = None
    fallback_role_name = normalize_optional_text(fallback_role, 100)
    if fallback_role is not None:
        try:
            fallback_role_id = int(fallback_role)
        except (TypeError, ValueError):
            fallback_role_id = None

        if fallback_role_id is not None:
            cursor.execute(
                "SELECT TOP 1 nombre_rol FROM ofertas.roles WHERE id_rol = ?",
                (fallback_role_id,),
            )
            fallback_role_row = cursor.fetchone()
            if fallback_role_row:
                fallback_role_name = fallback_role_row[0]
        elif fallback_role_name:
            cursor.execute(
                "SELECT TOP 1 id_rol, nombre_rol FROM ofertas.roles WHERE LOWER(LTRIM(RTRIM(nombre_rol))) = LOWER(LTRIM(RTRIM(?)))",
                (fallback_role_name,),
            )
            fallback_role_row = cursor.fetchone()
            if fallback_role_row:
                fallback_role_id = fallback_role_row[0]
                fallback_role_name = fallback_role_row[1]

    if row:
        department = None
        if row[3] is not None:
            department = {
                "id_departamento": row[3],
                "nombre_departamento": row[4],
                "descripcion": row[5],
                "estado_activo": bool(row[6]) if row[6] is not None else True,
            }
        return {
            "email": row[0],
            "id_rol": row[1] or fallback_role_id,
            "nombre_rol": row[2] or fallback_role_name or "Estandar",
            "departamento": department,
        }

    if fallback_role_name or fallback_role_id is not None:
        return {"email": None, "id_rol": fallback_role_id, "nombre_rol": fallback_role_name or "Estandar", "departamento": None}

    return {"email": None, "id_rol": None, "nombre_rol": "Estandar", "departamento": None}


def get_role_permission_level(role_name):
    normalized_role = str(role_name or "").strip().lower()
    if normalized_role == "manager":
        return 2
    return 1


def get_role_id_by_name(cursor, role_name):
    cursor.execute(
        """
        SELECT TOP 1 id_rol
        FROM ofertas.roles
        WHERE LOWER(LTRIM(RTRIM(nombre_rol))) = LOWER(LTRIM(RTRIM(?)))
        """,
        (role_name,),
    )
    row = cursor.fetchone()
    return row[0] if row else None


def get_outlook_automation_session_store():
    config = OutlookGraphService.get_config()
    if config.get("auth_mode") == "app-only":
        return None
    return session


def get_user_department_ids(user_data=None):
    current_user = user_data or get_logged_user_data()
    departments = current_user.get("departamentos") or []
    resolved_ids = set()

    for department in departments:
        if isinstance(department, dict):
            department_id = department.get("id_departamento")
        else:
            department_id = department

        try:
            if department_id is not None:
                resolved_ids.add(int(department_id))
        except (TypeError, ValueError):
            continue

    return resolved_ids


NOTIFICATION_STATE_TRANSLATIONS = {
    "pendiente": "Ceka",
    "pendiente tecnico": "Ceka - technicke oddeleni",
    "pendiente compras": "Ceka - nakup",
    "enviada": "Odeslana",
    "pedido": "Objednavka",
    "cancelada": "Zrusena",
}


NOTIFICATION_DEPARTMENT_TRANSLATIONS = {
    "administracion": "Administrativa",
    "tecnico": "Technicke oddeleni",
    "compras": "Nakup",
    "prueba": "Test",
}


NOTIFICATION_LOGO_PATH_CANDIDATES = (
    os.path.join(PROJECT_DIR, "static", "images", "Logo_EMESA.png"),
    os.path.join(PROJECT_DIR, "static", "images", ".Logo_EMESA.png"),
)


_notification_logo_data_uri = None


def normalize_notification_lookup_key(value):
    normalized = normalize_optional_text(value, 255) or ""
    normalized = normalized.strip().lower()
    for source, target in {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "ñ": "n",
    }.items():
        normalized = normalized.replace(source, target)
    return re.sub(r"\s+", " ", normalized)


def translate_notification_state_name(value):
    normalized = normalize_optional_text(value, 255)
    if not normalized:
        return normalized
    return NOTIFICATION_STATE_TRANSLATIONS.get(normalize_notification_lookup_key(normalized), normalized)


def translate_notification_department_name(value):
    normalized = normalize_optional_text(value, 255)
    if not normalized:
        return normalized
    return NOTIFICATION_DEPARTMENT_TRANSLATIONS.get(normalize_notification_lookup_key(normalized), normalized)


def get_notification_logo_url():
    public_candidates = [
        get_env_value("OAUTH_POST_LOGOUT_REDIRECT_URI", "PUBLIC_BASE_URL", default=""),
        get_env_value("OAUTH_REDIRECT_URI", default=""),
    ]

    for candidate in public_candidates:
        parsed = urlparse(str(candidate or "").strip())
        if parsed.scheme and parsed.netloc:
            return f"{parsed.scheme}://{parsed.netloc}/static/images/Logo_EMESA.png"

    return None


def get_notification_logo_data_uri():
    global _notification_logo_data_uri

    if _notification_logo_data_uri is not None:
        return _notification_logo_data_uri

    for candidate in NOTIFICATION_LOGO_PATH_CANDIDATES:
        if not os.path.exists(candidate):
            continue

        with open(candidate, "rb") as logo_file:
            encoded = base64.b64encode(logo_file.read()).decode("ascii")
        _notification_logo_data_uri = f"data:image/png;base64,{encoded}"
        return _notification_logo_data_uri

    _notification_logo_data_uri = ""
    return _notification_logo_data_uri


def get_notification_logo_src():
    return get_notification_logo_url() or get_notification_logo_data_uri() or ""


def build_offer_notification_message(
    oferta_label,
    estado_nombre,
    departamento_nombre,
    cliente,
    referencia,
    emisor,
    sender_department_name=None,
    assigned_by_name=None,
    detail_department_name=None,
    department_intro_message=None,
    source_department_name=None,
    target_department_name=None,
):
    estado_nombre = translate_notification_state_name(estado_nombre) or "Bez stavu"
    departamento_nombre = translate_notification_department_name(departamento_nombre)
    sender_department_name = translate_notification_department_name(sender_department_name)
    detail_department_name = translate_notification_department_name(
        detail_department_name if detail_department_name is not None else departamento_nombre
    )
    source_department_name = translate_notification_department_name(source_department_name)
    target_department_name = translate_notification_department_name(target_department_name)
    logo_src = get_notification_logo_src()

    # Referencia original en espanol:
    # subject = f"Oferta {oferta_label} en estado {estado_nombre}"
    # body = """<p>La oferta <strong>{oferta}</strong> ha llegado al estado <strong>{estado}</strong>.</p>
    # <p>Este estado esta asociado al departamento <strong>{departamento}</strong>.</p>
    # <ul>
    #   <li><strong>Oferta:</strong> {oferta}</li>
    #   <li><strong>Estado:</strong> {estado}</li>
    #   <li><strong>Departamento:</strong> {departamento}</li>
    #   <li><strong>Cliente:</strong> {cliente}</li>
    #   <li><strong>Referencia / asunto:</strong> {referencia}</li>
    #   <li><strong>Emisor:</strong> {emisor}</li>
    # </ul>"""
    subject = f"OFERTAS | Nabídka {oferta_label} byla převedena do stavu {estado_nombre}"
    assigned_by_html = ""
    if assigned_by_name:
        assigned_by_html = (
            '<tr>'
            '<td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Priradil</td>'
            '<td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{assigned_by}</td>'
            '</tr>'
        ).format(assigned_by=html_escape(assigned_by_name))

    logo_html = ""
    department_intro_html = ""
    if source_department_name or target_department_name:
        source_label = html_escape(source_department_name or "Bez oddeleni")
        target_label = html_escape(target_department_name or "Bez oddeleni")
        department_intro_html = """
            <p style="margin: 0 0 20px; font-size: 15px; color: #486581; line-height: 1.7;">
                Nabidka byla presunuta z oddeleni <strong style="color: #1a1a2e;">{source_department}</strong> do oddeleni <strong style="color: #1a1a2e;">{target_department}</strong>.
            </p>
        """.format(
            source_department=source_label,
            target_department=target_label,
        )
    elif detail_department_name:
        intro_message = department_intro_message or "Tento stav spada pod oddeleni"
        department_intro_html = """
            <p style="margin: 0 0 20px; font-size: 15px; color: #486581; line-height: 1.7;">
                {intro_message} <strong style="color: #1a1a2e;">{department_name}</strong>.
            </p>
        """.format(
            intro_message=html_escape(intro_message),
            department_name=html_escape(detail_department_name),
        )

    department_row_html = ""
    if source_department_name or target_department_name:
        department_row_html = """
            <tr>
                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Oddeleni puvodu</td>
                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{source_department}</td>
            </tr>
            <tr>
                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Oddeleni cile</td>
                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{target_department}</td>
            </tr>
        """.format(
            source_department=html_escape(source_department_name or "Bez oddeleni"),
            target_department=html_escape(target_department_name or "Bez oddeleni"),
        )
    elif detail_department_name:
        department_row_html = """
            <tr>
                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Oddeleni</td>
                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{department_name}</td>
            </tr>
        """.format(department_name=html_escape(detail_department_name))

    if logo_src:
        logo_html = (
            '<img src="{src}" alt="EMESA" '
            'style="display:block; max-width: 168px; width: 168px; height: auto; border: 0;" />'
        ).format(src=html_escape(logo_src))

    body = """<table role="presentation" cellpadding="0" cellspacing="0" border="0" width="100%" style="background-color: #f4f5f7; padding: 40px 20px;">
        <tr>
            <td align="center">
                <table role="presentation" cellpadding="0" cellspacing="0" border="0" width="100%" style="max-width: 620px; background-color: #ffffff; border-radius: 12px; overflow: hidden;">

                    <!-- Header -->
                    <tr>
                        <td style="padding: 32px 36px 28px 36px; background-color: #8b1e2d;">
                            <table role="presentation" cellpadding="0" cellspacing="0" border="0" width="100%">
                                <tr>
                                    <td style="vertical-align: middle;">
                                        <p style="margin: 0 0 4px 0; font-size: 11px; letter-spacing: 0.18em; text-transform: uppercase; color: rgba(255,255,255,0.72);">Automaticke oznameni</p>
                                        <p style="margin: 0; font-size: 26px; font-weight: 700; color: #ffffff; letter-spacing: -0.01em;">OFERTAS</p>
                                    </td>
                                    <td align="right" style="vertical-align: middle;">
                                        {logo_html}
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>

                    <!-- Body -->
                    <tr>
                        <td style="padding: 36px 36px 12px 36px;">

                            <!-- Greeting -->
                            <p style="margin: 0 0 20px 0; font-size: 16px; color: #1a1a2e;">
                                <strong>Dobry den,</strong>
                            </p>

                            <!-- Main message -->
                            <p style="margin: 0 0 8px 0; font-size: 15px; color: #334e68; line-height: 1.7;">
                                potvrzujeme, ze nabidka <strong style="color: #1a1a2e;">{oferta}</strong> byla prevedena do stavu
                                <strong style="color: #1a1a2e;">&bdquo;{estado}&rdquo;</strong>.
                            </p>

                            {department_intro_html}

                            <!-- Info card -->
                            <table role="presentation" cellpadding="0" cellspacing="0" border="0" width="100%" style="background-color: #f8fafc; border-radius: 10px; margin-top: 24px;">
                                <tr>
                                    <td style="padding: 24px 28px;">
                                        <p style="margin: 0 0 18px 0; font-size: 14px; font-weight: 600; color: #627d98; letter-spacing: 0.04em; text-transform: uppercase;">Prehled nabidky</p>

                                        <table role="presentation" cellpadding="0" cellspacing="0" border="0" width="100%">
                                            <tr>
                                                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; width: 170px; vertical-align: top;">Nabidka</td>
                                                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{oferta}</td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Stav</td>
                                                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{estado}</td>
                                            </tr>
                                            {department_row_html}
                                            <tr>
                                                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Zakaznik</td>
                                                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{cliente}</td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Reference / predmet e-mailu</td>
                                                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e; word-break: break-word;">{referencia}</td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 7px 0; font-size: 13px; font-weight: 600; color: #627d98; vertical-align: top;">Odesilatel</td>
                                                <td style="padding: 7px 0; font-size: 14px; color: #1a1a2e;">{emisor}</td>
                                            </tr>
                                            {assigned_by_html}
                                        </table>
                                    </td>
                                </tr>
                            </table>

                            <!-- Footer -->
                            <p style="margin: 28px 0 0 0; font-size: 12px; color: #9fb3c8; line-height: 1.7;">
                                Toto je automaticke upozorneni ze systemu OFERTAS. V pripade dotazu kontaktujte prosim prislusne oddeleni.
                            </p>
                        </td>
                    </tr>

                    <!-- Bottom spacer -->
                    <tr>
                        <td style="padding: 0 36px 36px 36px;">
                            <hr style="border: none; border-top: 1px solid #e4e7eb; margin: 24px 0 0 0;">
                            <p style="margin: 16px 0 0 0; font-size: 11px; color: #bcccdc; text-align: center;">
                                &copy; EMESA &middot; OFERTAS
                            </p>
                        </td>
                    </tr>

                </table>
            </td>
        </tr>
    </table>""".format(
        oferta=html_escape(oferta_label),
        estado=html_escape(estado_nombre),
        department_intro_html=department_intro_html,
        department_row_html=department_row_html,
        cliente=html_escape(cliente or "Bez zakaznika"),
        referencia=html_escape(referencia or "Bez reference"),
        emisor=html_escape(emisor or "Bez odesilatele"),
        assigned_by_html=assigned_by_html,
        logo_html=logo_html,
    )
    return {"subject": subject, "body": body}


def normalize_notification_matching_label(value):
    normalized = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    normalized = "".join(character for character in normalized if not unicodedata.combining(character))
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def get_manager_notification_recipients(cursor, department_id):
    if department_id is None:
        return []

    cursor.execute(
        """
        SELECT DISTINCT LOWER(LTRIM(RTRIM(uc.email)))
        FROM ofertas.usuarios_config uc
        LEFT JOIN ofertas.roles r
            ON r.id_rol = uc.id_rol
        WHERE uc.id_departamento = ?
          AND uc.email IS NOT NULL
          AND LTRIM(RTRIM(uc.email)) <> ''
          AND (
              uc.id_rol = 1
              OR LOWER(LTRIM(RTRIM(ISNULL(r.nombre_rol, '')))) = 'manager'
          )
        """,
        (department_id,),
    )
    return [row[0] for row in cursor.fetchall() if row and row[0]]


def resolve_notification_department_fallback(cursor, estado_nombre, current_department_id=None):
    normalized_state_name = normalize_notification_matching_label(estado_nombre)
    if not normalized_state_name:
        return None

    cursor.execute(
        """
        SELECT id_departamento, nombre_departamento
        FROM ofertas.departamentos
        WHERE ISNULL(estado_activo, 1) = 1
        ORDER BY id_departamento ASC
        """
    )
    candidates = []
    for row in cursor.fetchall():
        department_id = row[0]
        department_name = normalize_optional_text(row[1], 255)
        normalized_department_name = normalize_notification_matching_label(department_name)
        if not normalized_department_name or department_id == current_department_id:
            continue
        if normalized_department_name in normalized_state_name or normalized_state_name in normalized_department_name:
            candidates.append((len(normalized_department_name), department_id, department_name))

    if not candidates:
        return None

    candidates.sort(key=lambda item: (-item[0], item[1]))
    _, department_id, department_name = candidates[0]
    return {
        "id_departamento": department_id,
        "nombre_departamento": department_name,
    }


def build_estado_manager_notification(
    cursor,
    oferta_id,
    estado_id,
    sender_department_name=None,
    source_department_name=None,
    target_department_name=None,
):
    cursor.execute(
        """
        SELECT
            e.descripcion_estado,
            e.id_departamento,
            d.nombre_departamento
        FROM ofertas.estados e
        LEFT JOIN ofertas.departamentos d
            ON d.id_departamento = e.id_departamento
        WHERE e.id_estado = ?
        """,
        (estado_id,),
    )
    estado_row = cursor.fetchone()

    if not estado_row:
        return {"skipped": True, "message": "Estado no encontrado"}

    estado_nombre = normalize_optional_text(estado_row[0], 255) or f"#{estado_id}"
    id_departamento = estado_row[1]
    departamento_nombre = normalize_optional_text(estado_row[2], 255)
    display_target_department_name = normalize_optional_text(target_department_name, 255) or departamento_nombre
    display_source_department_name = normalize_optional_text(source_department_name, 255)

    if id_departamento is None:
        return {
            "skipped": True,
            "message": "El estado no tiene un departamento asociado",
            "estado": estado_nombre,
        }

    recipients = get_manager_notification_recipients(cursor, id_departamento)

    if not recipients:
        fallback_department = resolve_notification_department_fallback(
            cursor,
            estado_nombre,
            current_department_id=id_departamento,
        )
        if fallback_department is not None:
            fallback_recipients = get_manager_notification_recipients(cursor, fallback_department["id_departamento"])
            if fallback_recipients:
                recipients = fallback_recipients

    if not recipients:
        return {
            "skipped": True,
            "message": "No hay managers con email configurado en el departamento del estado",
            "estado": estado_nombre,
            "departamento": display_target_department_name,
        }

    cursor.execute(
        """
        SELECT
            lo.numero_oferta,
            lo.ref_cliente_asunto_email,
            c.descripcion_cliente,
            lo.nombre_emisor,
            lo.email_emisor
        FROM ofertas.listado_ofertas lo
        LEFT JOIN ofertas.clientes c
            ON c.id_cliente = lo.id_cliente
        WHERE lo.id_oferta = ?
        """,
        (oferta_id,),
    )
    oferta_row = cursor.fetchone()

    numero_oferta = normalize_optional_text(oferta_row[0], 100) if oferta_row else None
    referencia = normalize_optional_text(oferta_row[1], 500) if oferta_row else None
    cliente = normalize_optional_text(oferta_row[2], 255) if oferta_row else None
    emisor = format_sender_display(oferta_row[3], oferta_row[4]) if oferta_row else None

    oferta_label = numero_oferta or f"ID {oferta_id}"
    display_department_name = normalize_optional_text(sender_department_name, 255)
    message_payload = build_offer_notification_message(
        oferta_label=oferta_label,
        estado_nombre=estado_nombre,
        departamento_nombre=display_target_department_name,
        cliente=cliente,
        referencia=referencia,
        emisor=emisor,
        sender_department_name=sender_department_name,
        detail_department_name=display_department_name,
        department_intro_message="Změnu provedlo oddělení",
        source_department_name=display_source_department_name,
        target_department_name=display_target_department_name,
    )

    return {
        "skipped": False,
        "oferta_id": oferta_id,
        "to_recipients": recipients,
        "subject": message_payload["subject"],
        "body": message_payload["body"],
        "estado": estado_nombre,
        "departamento": display_target_department_name,
    }


def build_new_offer_notification(cursor, oferta_id, estado_id):
    """Construye la notificación para una oferta recién creada (no un cambio de estado)."""
    cursor.execute(
        """
        SELECT
            e.descripcion_estado,
            e.id_departamento,
            d.nombre_departamento
        FROM ofertas.estados e
        LEFT JOIN ofertas.departamentos d
            ON d.id_departamento = e.id_departamento
        WHERE e.id_estado = ?
        """,
        (estado_id,),
    )
    estado_row = cursor.fetchone()

    if not estado_row:
        return {"skipped": True, "message": "Estado no encontrado"}

    estado_nombre = normalize_optional_text(estado_row[0], 255) or f"#{estado_id}"
    id_departamento = estado_row[1]
    departamento_nombre = normalize_optional_text(estado_row[2], 255)

    if id_departamento is None:
        return {
            "skipped": True,
            "message": "El estado no tiene un departamento asociado",
            "estado": estado_nombre,
        }

    recipients = get_manager_notification_recipients(cursor, id_departamento)

    if not recipients:
        fallback_department = resolve_notification_department_fallback(
            cursor,
            estado_nombre,
            current_department_id=id_departamento,
        )
        if fallback_department is not None:
            fallback_recipients = get_manager_notification_recipients(cursor, fallback_department["id_departamento"])
            if fallback_recipients:
                recipients = fallback_recipients

    if not recipients:
        return {
            "skipped": True,
            "message": "No hay managers con email configurado en el departamento del estado",
            "estado": estado_nombre,
            "departamento": departamento_nombre,
        }

    cursor.execute(
        """
        SELECT
            lo.numero_oferta,
            lo.ref_cliente_asunto_email,
            c.descripcion_cliente,
            lo.nombre_emisor,
            lo.email_emisor
        FROM ofertas.listado_ofertas lo
        LEFT JOIN ofertas.clientes c
            ON c.id_cliente = lo.id_cliente
        WHERE lo.id_oferta = ?
        """,
        (oferta_id,),
    )
    oferta_row = cursor.fetchone()

    numero_oferta = normalize_optional_text(oferta_row[0], 100) if oferta_row else None
    referencia = normalize_optional_text(oferta_row[1], 500) if oferta_row else None
    cliente = normalize_optional_text(oferta_row[2], 255) if oferta_row else None
    emisor = format_sender_display(oferta_row[3], oferta_row[4]) if oferta_row else None

    oferta_label = numero_oferta or f"ID {oferta_id}"

    # Asunto específico para oferta nueva (distinto al de cambio de estado)
    estado_nombre_i18n = translate_notification_state_name(estado_nombre) or "Bez stavu"
    departamento_nombre_i18n = translate_notification_department_name(departamento_nombre)
    subject = f"OFERTAS | Nova nabidka {oferta_label} ve stavu {estado_nombre_i18n}"

    message_payload = build_offer_notification_message(
        oferta_label=oferta_label,
        estado_nombre=estado_nombre,
        departamento_nombre=departamento_nombre,
        cliente=cliente,
        referencia=referencia,
        emisor=emisor,
        detail_department_name=departamento_nombre_i18n,
        department_intro_message="Tato nabidka spada pod oddeleni",
    )
    message_payload["subject"] = subject

    return {
        "skipped": False,
        "oferta_id": oferta_id,
        "to_recipients": recipients,
        "subject": message_payload["subject"],
        "body": message_payload["body"],
        "estado": estado_nombre,
        "departamento": departamento_nombre,
    }


def build_reassignment_notification(cursor, oferta_id, assigned_by_name=None):
    cursor.execute(
        """
        SELECT
            lo.numero_oferta,
            e.descripcion_estado,
            d.nombre_departamento,
            c.descripcion_cliente,
            lo.ref_cliente_asunto_email,
            lo.nombre_emisor,
            lo.email_emisor,
            uc.email,
            gu.nombre
        FROM ofertas.listado_ofertas lo
        LEFT JOIN ofertas.estados e
            ON e.id_estado = lo.id_estado
        LEFT JOIN ofertas.departamentos d
            ON d.id_departamento = e.id_departamento
        LEFT JOIN ofertas.clientes c
            ON c.id_cliente = lo.id_cliente
        LEFT JOIN ofertas.oferta_etc oetc
            ON oetc.id_oferta_etc = lo.id_oferta
        LEFT JOIN ofertas.usuarios_config uc
            ON uc.num_operario = oetc.num_operario_responsable
        LEFT JOIN general.usuarios gu
            ON gu.num_operario = oetc.num_operario_responsable
        WHERE lo.id_oferta = ?
        """,
        (oferta_id,),
    )
    row = cursor.fetchone()
    if not row:
        return {"skipped": True, "message": "Oferta no encontrada"}

    recipient_email = normalize_optional_text(row[7], 255)
    if recipient_email:
        recipient_email = recipient_email.lower()
    if not recipient_email:
        return {"skipped": True, "message": "El usuario reasignado no tiene email configurado"}

    oferta_label = normalize_optional_text(row[0], 100) or f"ID {oferta_id}"
    estado_nombre = normalize_optional_text(row[1], 255) or "Bez stavu"
    departamento_nombre = normalize_optional_text(row[2], 255)
    cliente = normalize_optional_text(row[3], 255)
    referencia = normalize_optional_text(row[4], 500)
    emisor = format_sender_display(row[5], row[6])
    responsable_nombre = normalize_optional_text(row[8], 255)
    message_payload = build_offer_notification_message(
        oferta_label=oferta_label,
        estado_nombre=estado_nombre,
        departamento_nombre=departamento_nombre,
        cliente=cliente,
        referencia=referencia,
        emisor=emisor,
        assigned_by_name=assigned_by_name,
    )

    return {
        "skipped": False,
        "oferta_id": oferta_id,
        "to_recipients": [recipient_email],
        "subject": message_payload["subject"],
        "body": message_payload["body"],
        "estado": estado_nombre,
        "departamento": departamento_nombre,
        "responsable": responsable_nombre,
    }


def send_reassignment_notification(notification_payload):
    if not notification_payload:
        return {"success": False, "sent": False, "message": "No se ha generado la notificacion"}

    if notification_payload.get("skipped"):
        return {
            "success": True,
            "sent": False,
            "skipped": True,
            "message": notification_payload.get("message"),
            "recipients": [],
        }

    try:
        result = OutlookGraphService.send_mail(
            get_outlook_automation_session_store(),
            subject=notification_payload["subject"],
            body=notification_payload["body"],
            to_recipients=notification_payload["to_recipients"],
            is_html=True,
            save_to_sent_items=True,
        )
        return {
            "success": True,
            "sent": True,
            "message": "Aviso enviado al usuario reasignado",
            "recipients": notification_payload.get("to_recipients") or [],
            "account": result.get("account"),
        }
    except Exception as exc:
        app.logger.exception(
            "No se pudo enviar la notificacion de reasignacion para la oferta %s",
            notification_payload.get("oferta_id"),
        )
        return {
            "success": False,
            "sent": False,
            "message": str(exc),
            "recipients": notification_payload.get("to_recipients") or [],
        }


def send_estado_manager_notification(notification_payload):
    if not notification_payload:
        return {"success": False, "sent": False, "message": "No se ha generado la notificación"}

    if notification_payload.get("skipped"):
        return {
            "success": True,
            "sent": False,
            "skipped": True,
            "message": notification_payload.get("message"),
            "estado": notification_payload.get("estado"),
            "departamento": notification_payload.get("departamento"),
            "recipients": [],
        }

    try:
        result = OutlookGraphService.send_mail(
            get_outlook_automation_session_store(),
            subject=notification_payload["subject"],
            body=notification_payload["body"],
            to_recipients=notification_payload["to_recipients"],
            is_html=True,
            save_to_sent_items=True,
        )
        return {
            "success": True,
            "sent": True,
            "message": "Aviso enviado a los managers del departamento",
            "estado": notification_payload.get("estado"),
            "departamento": notification_payload.get("departamento"),
            "recipients": notification_payload.get("to_recipients") or [],
            "account": result.get("account"),
        }
    except Exception as exc:
        app.logger.exception(
            "No se pudo enviar la notificación automática para la oferta %s al estado %s",
            notification_payload.get("oferta_id"),
            notification_payload.get("estado"),
        )
        return {
            "success": False,
            "sent": False,
            "message": str(exc),
            "estado": notification_payload.get("estado"),
            "departamento": notification_payload.get("departamento"),
            "recipients": notification_payload.get("to_recipients") or [],
        }


def usuario_exists(cursor, num_operario):
    cursor.execute(
        """
        SELECT TOP 1 1
        FROM general.usuarios
        WHERE num_operario = ?
        """,
        (num_operario,),
    )
    return cursor.fetchone() is not None


def get_general_user(cursor, num_operario):
    cursor.execute(
        """
        SELECT TOP 1 id_usuario, num_operario, nombre, nivel_permisos, roles
        FROM general.usuarios
        WHERE num_operario = ?
        """,
        (num_operario,),
    )
    row = cursor.fetchone()
    if row is None:
        return None

    return {
        "id_usuario": row[0],
        "num_operario": row[1],
        "nombre": row[2],
        "nivel_permisos": row[3],
        "rol": row[4],
    }


def get_general_user_column_names(cursor):
    column_names = set()
    try:
        for row in cursor.columns(table="usuarios", schema="general"):
            raw_name = getattr(row, "column_name", None)
            if raw_name is None and len(row) > 3:
                raw_name = row[3]
            normalized = normalize_optional_text(raw_name, 128)
            if normalized:
                column_names.add(normalized.lower())
    except Exception:
        return set()
    return column_names


def get_general_user_by_column_value(cursor, column_name, value):
    normalized_value = normalize_optional_text(value, 255)
    if not normalized_value:
        return None

    cursor.execute(
        f"""
        SELECT TOP 1 id_usuario, num_operario, nombre, nivel_permisos, roles
        FROM general.usuarios
        WHERE LOWER(LTRIM(RTRIM([{column_name}]))) = LOWER(LTRIM(RTRIM(?)))
        """,
        (normalized_value,),
    )
    row = cursor.fetchone()
    if row is None:
        return None

    return {
        "id_usuario": row[0],
        "num_operario": row[1],
        "nombre": row[2],
        "nivel_permisos": row[3],
        "rol": row[4],
    }


def build_authenticated_user_data(cursor, general_user):
    num_operario = general_user["num_operario"]
    user_config = get_user_config(cursor, num_operario, fallback_role=general_user.get("rol"))
    departamentos = [user_config["departamento"]] if user_config["departamento"] else []
    return {
        "id": general_user["id_usuario"],
        "num_operario": num_operario,
        "usuario": num_operario,
        "nombre": general_user["nombre"],
        "nivel": general_user["nivel_permisos"],
        "nivel_permisos": general_user["nivel_permisos"],
        "id_rol": user_config["id_rol"],
        "rol": user_config["nombre_rol"],
        "read_only": int(num_operario) in READ_ONLY_OPERARIO_NUMBERS,
        "departamentos": departamentos,
        "success": True,
    }


def persist_authenticated_user(user_data):
    session["user_id"] = user_data["id"]
    session["user_data"] = user_data
    session.permanent = True


def refresh_authenticated_session_user():
    current_user = get_logged_user_data()
    if not current_user:
        return None

    num_operario = current_user.get("num_operario")
    if num_operario is None:
        return current_user

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            general_user = get_general_user(cursor, num_operario)
            if general_user is None:
                session.pop("user_data", None)
                session.pop("user_id", None)
                return None

            refreshed_user = build_authenticated_user_data(cursor, general_user)
            refreshed_user["read_only"] = is_read_only_user(refreshed_user)
            persist_authenticated_user(refreshed_user)
            return refreshed_user
    except Exception:
        current_user["read_only"] = is_read_only_user(current_user)
        session["user_data"] = current_user
        return current_user


def resolve_general_user_from_microsoft_profile(cursor, profile):
    employee_id = normalize_optional_text(profile.get("employee_id"), 50)
    if employee_id:
        try:
            user_by_employee_id = get_general_user(cursor, int(employee_id))
            if user_by_employee_id:
                return user_by_employee_id
        except (TypeError, ValueError):
            pass

    identity_values = []
    seen_values = set()
    for raw_value in (
        profile.get("mail"),
        profile.get("user_principal_name"),
        (profile.get("account") or {}).get("username"),
    ):
        normalized = normalize_optional_text(raw_value, 255)
        if normalized and normalized.lower() not in seen_values:
            identity_values.append(normalized)
            seen_values.add(normalized.lower())
            if "@" in normalized:
                alias = normalize_optional_text(normalized.split("@", 1)[0], 120)
                if alias and alias.lower() not in seen_values:
                    identity_values.append(alias)
                    seen_values.add(alias.lower())

    available_columns = get_general_user_column_names(cursor)
    for column_name in GENERAL_USER_MICROSOFT_IDENTITY_COLUMNS:
        if column_name.lower() not in available_columns:
            continue
        for value in identity_values:
            match = get_general_user_by_column_value(cursor, column_name, value)
            if match:
                return match

    return None


def is_read_only_user(user_data=None):
    current_user = user_data or get_logged_user_data()
    if not current_user:
        return False

    if current_user.get("read_only") is True:
        return True

    try:
        return int(current_user.get("num_operario")) in READ_ONLY_OPERARIO_NUMBERS
    except (TypeError, ValueError):
        return False


def read_only_response():
    return jsonify(
        {
            "success": False,
            "read_only": True,
            "message": "Usuario en modo solo lectura. No puede crear, editar, eliminar ni guardar cambios.",
        }
    ), 403


def is_manager_user(user_data=None):
    current_user = user_data or get_logged_user_data()
    if not current_user:
        return False

    try:
        if int(current_user.get("id_rol")) == 1:
            return True
    except (TypeError, ValueError):
        pass

    return str(current_user.get("rol") or current_user.get("nombre_rol") or "").strip().lower() == "manager"


def manager_only_response(message="Solo los usuarios con rol Manager pueden añadir o editar usuarios."):
    return jsonify(
        {
            "success": False,
            "manager_only": True,
            "message": message,
        }
    ), 403


def normalize_optional_text(value, max_length=None):
    text = str(value or "").strip()
    if not text:
        return None
    if max_length is not None:
        return text[:max_length]
    return text


def normalize_optional_email(value, field_name="Email"):
    email = normalize_optional_text(value, 255)
    if email is None:
        return None
    if "@" not in email:
        raise ValueError(f"El campo {field_name} debe contener un email válido")
    return email.lower()


def normalize_required_text(value, field_name, max_length=None):
    text = normalize_optional_text(value, max_length=max_length)
    if text is None:
        raise ValueError(f"El campo {field_name} es obligatorio")
    return text


def normalize_required_int(value, field_name):
    if value in (None, ""):
        raise ValueError(f"El campo {field_name} es obligatorio")

    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"El campo {field_name} debe ser numérico") from exc


def normalize_optional_int(value, field_name):
    if value in (None, ""):
        return None

    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"El campo {field_name} debe ser numérico") from exc


def normalize_optional_date(value, field_name):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None

    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"El campo {field_name} debe tener formato YYYY-MM-DD") from exc


def normalize_required_date(value, field_name):
    parsed_date = normalize_optional_date(value, field_name)
    if parsed_date is None:
        raise ValueError(f"El campo {field_name} es obligatorio")
    return parsed_date


def normalize_optional_decimal(value, field_name):
    if value in (None, ""):
        return None

    raw_value = str(value).strip().replace(",", ".")
    if not raw_value:
        return None

    try:
        return Decimal(raw_value)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"El campo {field_name} debe ser decimal") from exc


def normalize_optional_bool(value, field_name):
    if value in (None, ""):
        return None

    if isinstance(value, bool):
        return value

    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "t", "si", "sí", "s", "yes", "y"}:
        return True
    if normalized in {"0", "false", "f", "no", "n"}:
        return False

    raise ValueError(f"El campo {field_name} debe ser booleano")


def normalize_etc_priority(value):
    normalized = normalize_optional_text(value, 20)
    if normalized is None:
        return "NORMAL"

    priority = normalized.upper()
    valid_priorities = {"BAJA", "NORMAL", "ALTA", "CRITICA"}
    if priority not in valid_priorities:
        raise ValueError("El campo prioridad debe ser BAJA, NORMAL, ALTA o CRITICA")
    return priority


class HtmlToTextParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    @staticmethod
    def _is_quote_container(tag, attrs):
        if tag == "blockquote":
            return True

        attributes = {str(key or "").lower(): str(value or "") for key, value in attrs}
        style = attributes.get("style", "").lower()
        class_name = attributes.get("class", "").lower()

        if "border-left" in style and any(token in style for token in ("padding-left", "margin-left")):
            return True
        if any(token in class_name for token in ("quote", "quoted", "gmail_quote")):
            return True

        return False

    def handle_starttag(self, tag, attrs):
        if self._is_quote_container(tag, attrs):
            self.parts.append("\n----- quoted message -----\n")
            return

        if tag in {"br", "p", "div", "li", "tr", "table"}:
            self.parts.append("\n")

    def handle_endtag(self, tag):
        if tag in {"p", "div", "li", "tr", "table"}:
            self.parts.append("\n")

    def handle_data(self, data):
        if data:
            self.parts.append(data)

    def get_text(self):
        return "".join(self.parts)


def html_to_text(value):
    if not value:
        return ""

    parser = HtmlToTextParser()
    parser.feed(value)
    return unescape(parser.get_text())


def normalize_email_datetime(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    try:
        return parsedate_to_datetime(str(value))
    except (TypeError, ValueError, IndexError, OverflowError):
        return None


def decode_email_header(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return None

    try:
        decoded_value = str(make_header(decode_header(raw_value))).strip()
    except Exception:
        decoded_value = raw_value

    return decoded_value or None


def format_sender_display(sender_name=None, sender_email=None):
    normalized_name = normalize_optional_text(sender_name, 255)
    normalized_email = normalize_optional_text(sender_email, 255)
    if normalized_email:
        normalized_email = normalized_email.lower()

    if normalized_name and normalized_email:
        return f"{normalized_name} <{normalized_email}>"
    return normalized_email or normalized_name


def extract_sender_identity(value, fallback_email=None):
    raw_value = str(value or "").strip()
    parsed_name, parsed_email = parseaddr(raw_value)

    sender_name = decode_email_header(parsed_name)
    sender_email = normalize_optional_text(parsed_email or fallback_email, 255)
    if sender_email:
        sender_email = sender_email.lower()

    if sender_name is None and raw_value and raw_value != sender_email and "@" not in raw_value:
        sender_name = decode_email_header(raw_value)

    if sender_name and sender_email and sender_name.strip().lower() == sender_email.strip().lower():
        sender_name = None

    return {
        "sender_name": sender_name,
        "sender_email": sender_email,
        "sender_display": format_sender_display(sender_name, sender_email),
    }


def extract_sender_email(value):
    return extract_sender_identity(value).get("sender_email")


def extract_sender_domain(value):
    sender_email = extract_sender_email(value)
    if not sender_email or "@" not in sender_email:
        return sender_email, None

    raw_domain = sender_email.split("@", 1)[1].strip().lower().strip(".")
    if raw_domain.startswith("www."):
        raw_domain = raw_domain[4:]

    return sender_email, raw_domain or None


def get_domain_root(domain):
    normalized = str(domain or "").strip().lower()
    if not normalized:
        return None

    labels = [label for label in normalized.split(".") if label]
    if not labels:
        return None

    compound_suffixes = {
        ("co", "uk"),
        ("com", "es"),
        ("com", "mx"),
        ("com", "ar"),
        ("com", "br"),
        ("com", "co"),
        ("co", "jp"),
    }

    if len(labels) >= 3 and tuple(labels[-2:]) in compound_suffixes:
        return labels[-3]
    if len(labels) >= 2:
        return labels[-2]
    return labels[0]


def normalize_identifier(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def normalize_cliente_domain(value):
    normalized = normalize_optional_text(value, 255)
    if normalized is None:
        return None

    normalized = normalized.strip().lower()
    if "@" in normalized:
        normalized = normalized.split("@", 1)[1]

    normalized = normalized.strip(".")
    if normalized.startswith("www."):
        normalized = normalized[4:]

    if not normalized:
        return None

    if not re.fullmatch(r"[a-z0-9.-]+", normalized) or "." not in normalized:
        raise ValueError("Dominio cliente no válido")

    return normalized


def resolve_cliente_from_email(cursor, sender_email):
    sender_email, raw_domain = extract_sender_domain(sender_email)
    if not sender_email:
        return None

    domain_root = normalize_identifier(get_domain_root(raw_domain)) if raw_domain else None
    if not raw_domain:
        return {"sender_email": sender_email, "domain": None, "domain_root": domain_root, "cliente": None}

    try:
        domain = normalize_cliente_domain(raw_domain)
    except ValueError:
        return {
            "sender_email": sender_email,
            "domain": raw_domain,
            "domain_root": domain_root,
            "cliente": None,
        }

    domain_labels = [label for label in domain.split(".") if label]
    candidate_domains = [".".join(domain_labels[index:]) for index in range(len(domain_labels) - 1)] or [domain]
    placeholders = ", ".join("?" for _ in candidate_domains)

    cursor.execute(
        f"""
        SELECT id_cliente, descripcion_cliente, dominio
        FROM ofertas.clientes
        WHERE LOWER(LTRIM(RTRIM(ISNULL(dominio, '')))) IN ({placeholders})
        ORDER BY CASE WHEN LOWER(LTRIM(RTRIM(ISNULL(dominio, '')))) = ? THEN 0 ELSE 1 END,
                 LEN(LTRIM(RTRIM(ISNULL(dominio, '')))) DESC,
                 descripcion_cliente ASC
        """,
        tuple(candidate_domains + [domain]),
    )
    row = cursor.fetchone()

    best_match = None
    if row is not None:
        best_match = {
            "id_cliente": row[0],
            "descripcion_cliente": row[1],
            "dominio": row[2],
        }

    return {
        "sender_email": sender_email,
        "domain": domain,
        "domain_root": domain_root,
        "cliente": best_match,
    }


def clean_email_body(body_text):
    normalized = str(body_text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = normalized.split("\n")

    header_regex = re.compile(
        r"^(from|de|sent|enviado|date|datum|to|para|cc|subject|asunto|od|odeslano|odesláno|komu|predmet|předmět)\s*:",
        re.IGNORECASE,
    )
    original_message_regex = re.compile(
        r"^(?:-{2,}\s*(?:original|quoted)\s*message\s*-*|-{2,}\s*p[ůu]vodn[ií]\s*e[-‑–—]?mail\s*-*|_{5,})$",
        re.IGNORECASE,
    )
    forwarded_message_regex = re.compile(
        r"^(?:-{2,}\s*)?(?:forwarded message|mensaje reenviado|přeposlaná zpráva|preposlana zprava)(?:\s*-{2,})?$|^begin forwarded message:\s*$",
        re.IGNORECASE,
    )
    forwarding_intro_regex = re.compile(
        r"^(?:te\s+lo\s+reenv[ií]o|reenv[ií]o|reenviado|forwarding|forwarded|fyi|přepos[ií]l[aá]m|preposilam)\b",
        re.IGNORECASE,
    )
    mobile_signatures = {
        "sent from my iphone",
        "sent from my ipad",
        "enviado desde mi iphone",
        "enviado desde mi ipad",
        "obtener outlook para android",
    }
    signature_markers = (
        "--",
        "__",
        "saludos",
        "un saludo",
        "atentamente",
        "best regards",
        "kind regards",
        "muchas gracias",
        "gracias",
    )

    cleaned_lines = []
    signature_name = None
    seen_message_content = False

    def looks_like_forwarded_header_block(start_index):
        matched_headers = 0
        for candidate_line in lines[start_index:start_index + 5]:
            candidate = candidate_line.strip()
            if not candidate:
                continue
            if header_regex.match(candidate):
                matched_headers += 1
                continue
            break
        return matched_headers >= 2

    def append_quoted_divider():
        if not cleaned_lines:
            return
        while cleaned_lines and cleaned_lines[-1] == "":
            cleaned_lines.pop()
        if cleaned_lines and cleaned_lines[-1] != "----- quoted message -----":
            cleaned_lines.append("")
            cleaned_lines.append("----- quoted message -----")
            cleaned_lines.append("")

    for index, line in enumerate(lines):
        stripped = line.strip()
        lowered = stripped.lower()

        if not stripped:
            if cleaned_lines and cleaned_lines[-1] != "":
                cleaned_lines.append("")
            continue

        if header_regex.match(stripped) and looks_like_forwarded_header_block(index):
            append_quoted_divider()
            continue

        if stripped.startswith(">") or header_regex.match(stripped):
            continue

        if forwarded_message_regex.match(stripped):
            append_quoted_divider()
            continue

        if not seen_message_content and forwarding_intro_regex.match(stripped):
            continue

        if original_message_regex.match(stripped):
            append_quoted_divider()
            continue

        if lowered in mobile_signatures:
            break

        if any(lowered == marker or lowered.startswith(f"{marker},") for marker in signature_markers):
            for candidate in lines[index + 1:index + 4]:
                candidate = candidate.strip()
                if not candidate:
                    continue
                if len(candidate) <= 60 and "@" not in candidate and not re.search(r"\d{3,}", candidate):
                    signature_name = candidate
                break
            break

        cleaned_lines.append(stripped)
        seen_message_content = True

    while cleaned_lines and cleaned_lines[-1] == "":
        cleaned_lines.pop()

    cleaned_text = "\n".join(cleaned_lines).strip()
    cleaned_text = re.sub(r"\n{3,}", "\n\n", cleaned_text)

    if signature_name and signature_name.lower() not in cleaned_text.lower():
        cleaned_text = f"{cleaned_text}\n\n{signature_name}".strip()

    return cleaned_text


def extract_eml_attachments(message):
    attachments = []
    for part in message.walk():
        if part.is_multipart():
            continue

        filename = part.get_filename()
        disposition = (part.get_content_disposition() or "").lower()
        if not filename and disposition not in {"attachment", "inline"}:
            continue

        content_bytes = part.get_payload(decode=True)
        if not filename or not content_bytes:
            continue

        attachments.append(
            {
                "filename": decode_email_header(filename),
                "content_type": part.get_content_type(),
                "content_bytes": bytes(content_bytes),
            }
        )

    return attachments


def parse_eml_bytes(file_bytes):
    message = BytesParser(policy=policy.default).parsebytes(file_bytes)
    sender_identity = extract_sender_identity(message.get("from"))
    received_at = normalize_email_datetime(message.get("date"))
    subject = str(message.get("subject") or "").strip()

    body_part = None
    if message.is_multipart():
        body_part = message.get_body(preferencelist=("plain", "html"))

    if body_part is not None:
        body_text = body_part.get_content()
        if body_part.get_content_type() == "text/html":
            body_text = html_to_text(body_text)
    else:
        body_text = message.get_content()
        if message.get_content_type() == "text/html":
            body_text = html_to_text(body_text)

    reference_headers = []
    for header_name in ("references", "References"):
        reference_headers.extend(message.get_all(header_name, []))

    return {
        "sender_name": sender_identity["sender_name"],
        "sender_email": sender_identity["sender_email"],
        "sender_display": sender_identity["sender_display"],
        "received_at": received_at,
        "subject": subject,
        "body": clean_email_body(body_text),
        "attachments": extract_eml_attachments(message),
        "source_type": "email_file",
        "conversation_id": None,
        "internet_message_id": normalize_message_identifier(message.get("message-id") or message.get("Message-ID")),
        "in_reply_to_message_id": normalize_message_identifier(message.get("in-reply-to") or message.get("In-Reply-To")),
        "reference_message_ids": normalize_message_identifier_list(reference_headers),
    }


def parse_msg_bytes(file_bytes):
    try:
        extract_msg = importlib.import_module("extract_msg")
    except ImportError as exc:
        raise RuntimeError("Falta la dependencia 'extract-msg' para importar correos .msg") from exc

    temp_path = None
    message = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".msg") as temp_file:
            temp_file.write(file_bytes)
            temp_path = temp_file.name

        message = extract_msg.Message(temp_path)
        body_text = getattr(message, "body", None) or html_to_text(getattr(message, "htmlBody", None) or "")
        raw_headers = str(getattr(message, "header", "") or "")
        parsed_headers = Parser(policy=policy.default).parsestr(raw_headers) if raw_headers else None
        sender_identity = extract_sender_identity(
            getattr(message, "sender", None)
            or raw_headers,
            fallback_email=getattr(message, "sender_email", None),
        )

        attachments = []
        for attachment in list(getattr(message, "attachments", None) or []):
            filename = (
                getattr(attachment, "longFilename", None)
                or getattr(attachment, "filename", None)
                or getattr(attachment, "shortFilename", None)
                or getattr(attachment, "displayName", None)
            )
            content_bytes = getattr(attachment, "data", None)
            if callable(content_bytes):
                content_bytes = content_bytes()
            if isinstance(content_bytes, memoryview):
                content_bytes = content_bytes.tobytes()
            if isinstance(content_bytes, bytearray):
                content_bytes = bytes(content_bytes)

            if not filename or not isinstance(content_bytes, bytes) or not content_bytes:
                continue

            attachments.append(
                {
                    "filename": str(filename).strip(),
                    "content_type": getattr(attachment, "mimeType", None),
                    "content_bytes": content_bytes,
                }
            )

        return {
            "sender_name": sender_identity["sender_name"],
            "sender_email": sender_identity["sender_email"],
            "sender_display": sender_identity["sender_display"],
            "received_at": normalize_email_datetime(getattr(message, "date", None)),
            "subject": str(getattr(message, "subject", "") or "").strip(),
            "body": clean_email_body(body_text),
            "attachments": attachments,
            "source_type": "email_file",
            "conversation_id": None,
            "internet_message_id": normalize_message_identifier(parsed_headers.get("Message-ID") if parsed_headers else None),
            "in_reply_to_message_id": normalize_message_identifier(parsed_headers.get("In-Reply-To") if parsed_headers else None),
            "reference_message_ids": normalize_message_identifier_list(parsed_headers.get_all("References", []) if parsed_headers else []),
        }
    finally:
        close_message = getattr(message, "close", None)
        if callable(close_message):
            try:
                close_message()
            except Exception:
                app.logger.warning("No se pudo cerrar el mensaje .msg temporal", exc_info=True)

        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except PermissionError:
                app.logger.warning("No se pudo borrar el archivo temporal del correo .msg: %s", temp_path, exc_info=True)


def parse_uploaded_email(file_storage):
    filename = str(getattr(file_storage, "filename", "") or "")
    extension = os.path.splitext(filename)[1].lower()
    file_bytes = file_storage.read()

    if not file_bytes:
        raise ValueError("El archivo de correo está vacío")

    msg_signature = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

    def looks_like_eml_bytes(raw_bytes):
        if not raw_bytes:
            return False
        if raw_bytes.startswith(msg_signature):
            return False

        sample = raw_bytes[:8192]
        lowered = sample.lower()
        eml_markers = (b"\nfrom:", b"\nsubject:", b"\ndate:", b"\nmessage-id:", b"\nmime-version:")
        marker_hits = sum(1 for marker in eml_markers if marker in lowered)
        if marker_hits >= 2:
            return True

        # Some files include folded headers or start directly with a mail header line.
        return lowered.startswith((b"from:", b"subject:", b"date:", b"message-id:", b"mime-version:"))

    def detect_uploaded_email_format(raw_bytes, hinted_extension):
        hinted = str(hinted_extension or "").lower()
        if raw_bytes.startswith(msg_signature):
            return "msg"
        if looks_like_eml_bytes(raw_bytes):
            return "eml"
        if hinted == ".msg":
            return "msg"
        if hinted == ".eml":
            return "eml"
        return None

    detected_format = detect_uploaded_email_format(file_bytes, extension)
    if detected_format is None:
        raise ValueError("Formato de correo no soportado. Usa un archivo .eml o .msg")

    if detected_format == "eml":
        try:
            return parse_eml_bytes(file_bytes)
        except Exception:
            if file_bytes.startswith(msg_signature):
                return parse_msg_bytes(file_bytes)
            raise

    try:
        return parse_msg_bytes(file_bytes)
    except Exception:
        if looks_like_eml_bytes(file_bytes):
            return parse_eml_bytes(file_bytes)
        raise


def build_imported_email_response_data(parsed_email, cliente_resolution=None):
    matched_cliente = (cliente_resolution or {}).get("cliente")
    fecha_email = parsed_email.get("received_at")
    fecha_alta = datetime.now().date()

    return {
        "id_cliente": matched_cliente["id_cliente"] if matched_cliente else None,
        "cliente": matched_cliente["descripcion_cliente"] if matched_cliente else None,
        "nombre_emisor": parsed_email.get("sender_name"),
        "sender_email": parsed_email.get("sender_email"),
        "email_emisor": parsed_email.get("sender_email"),
        "emisor": parsed_email.get("sender_display"),
        "domain": (cliente_resolution or {}).get("domain"),
        "fecha_email": fecha_email.date().isoformat() if isinstance(fecha_email, datetime) else (fecha_email.isoformat() if fecha_email else None),
        "fecha_alta_oferta": fecha_alta.isoformat(),
        "ref_cliente_asunto_email": parsed_email.get("subject"),
        "observaciones": parsed_email.get("body"),
        "adjuntos_importados": [
            {
                "original_name": attachment.get("filename"),
                "size_bytes": len(attachment.get("content_bytes") or b""),
            }
            for attachment in (parsed_email.get("attachments") or [])
            if attachment.get("filename")
        ],
        "imported_email_metadata": serialize_imported_email_metadata(parsed_email),
    }


def resolve_cliente_for_sender_email(sender_email):
    with db_connection(autocommit=True) as conn:
        cursor = conn.cursor()
        try:
            return resolve_cliente_from_email(cursor, sender_email)
        except Exception as exc:
            app.logger.exception("Error al resolver cliente desde correo: %s", exc)
            return {
                "sender_email": sender_email,
                "domain": None,
                "domain_root": None,
                "cliente": None,
            }


def normalize_outlook_message_for_offer(message):
    body_content = message.get("body_content") or message.get("body_preview") or ""
    body_type = str(message.get("body_content_type") or "text").strip().lower()
    body_text = html_to_text(body_content) if body_type == "html" else str(body_content)
    body_text = clean_email_body(body_text)

    sender_name = normalize_optional_text(message.get("sender_name"), 255)
    sender_email = normalize_optional_text(message.get("sender_email"), 255)
    if sender_email:
        sender_email = sender_email.lower()

    return {
        "sender_name": sender_name,
        "sender_email": sender_email,
        "sender_display": format_sender_display(sender_name, sender_email),
        "subject": normalize_optional_text(message.get("subject"), 500),
        "received_at": normalize_email_datetime(message.get("received_at")),
        "body": body_text,
        "source_type": "outlook",
        "conversation_id": normalize_optional_text(message.get("conversation_id"), 255),
        "internet_message_id": normalize_message_identifier(message.get("internet_message_id")),
        "in_reply_to_message_id": None,
        "reference_message_ids": [],
    }


def normalize_message_identifier(value):
    if value is None:
        return None

    normalized = str(value).strip()
    if not normalized:
        return None

    normalized = normalized.strip("<>").strip().lower()
    return normalized or None


def normalize_message_identifier_list(values):
    if values is None:
        return []

    raw_values = values if isinstance(values, (list, tuple, set)) else [values]
    identifiers = []
    seen = set()

    for raw_value in raw_values:
        if raw_value is None:
            continue

        for candidate in re.findall(r"<([^>]+)>", str(raw_value)) or [raw_value]:
            normalized = normalize_message_identifier(candidate)
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            identifiers.append(normalized)

    return identifiers


def serialize_imported_email_metadata(parsed_email):
    if not isinstance(parsed_email, dict):
        return None

    metadata = {
        "source_type": normalize_optional_text(parsed_email.get("source_type"), 32),
        "conversation_id": normalize_optional_text(parsed_email.get("conversation_id"), 255),
        "internet_message_id": normalize_message_identifier(parsed_email.get("internet_message_id")),
        "in_reply_to_message_id": normalize_message_identifier(parsed_email.get("in_reply_to_message_id")),
        "reference_message_ids": normalize_message_identifier_list(parsed_email.get("reference_message_ids") or []),
        "subject": normalize_optional_text(parsed_email.get("subject"), 500),
        "sender_name": normalize_optional_text(parsed_email.get("sender_name"), 255),
        "sender_email": normalize_optional_text(parsed_email.get("sender_email"), 255),
        "received_at": serialize_date(parsed_email.get("received_at")),
        "body_sha256": build_email_body_sha256(parsed_email.get("body")),
    }
    if metadata["sender_email"]:
        metadata["sender_email"] = metadata["sender_email"].lower()

    if not any(
        metadata.get(key)
        for key in ("conversation_id", "internet_message_id", "in_reply_to_message_id", "reference_message_ids", "body_sha256")
    ):
        return None

    return metadata


def build_email_body_sha256(body_text):
    normalized_body = str(body_text or "").strip()
    if not normalized_body:
        return None
    return hashlib.sha256(normalized_body.encode("utf-8")).hexdigest()


def normalize_email_subject_for_matching(subject):
    normalized = normalize_optional_text(subject, 500)
    if not normalized:
        return None

    normalized = normalized.strip()
    prefix_regex = re.compile(r"^(?:(?:re|fw|fwd|rv|sv|wg|aw|tr|odp)\s*:\s*)+", re.IGNORECASE)
    previous = None
    while previous != normalized:
        previous = normalized
        normalized = prefix_regex.sub("", normalized).strip()

    return normalized.lower() or None


def oferta_email_tracking_table_exists(cursor):
    cursor.execute("SELECT 1 WHERE OBJECT_ID('ofertas.oferta_correos_importados', 'U') IS NOT NULL")
    return cursor.fetchone() is not None


def get_imported_email_metadata_for_match(parsed_email):
    if not isinstance(parsed_email, dict):
        return None

    metadata = serialize_imported_email_metadata(parsed_email)
    if metadata is None:
        return None
    return metadata


def get_offer_thread_match(cursor, parsed_email):
    if not oferta_email_tracking_table_exists(cursor):
        return None

    metadata = get_imported_email_metadata_for_match(parsed_email)
    if metadata is None:
        return None

    conversation_id = metadata.get("conversation_id")
    if conversation_id:
        cursor.execute(
            """
            SELECT TOP 1 lo.id_oferta, lo.numero_oferta, lo.ref_cliente_asunto_email
            FROM ofertas.oferta_correos_importados oci
            INNER JOIN ofertas.listado_ofertas lo
                ON lo.id_oferta = oci.id_oferta
            WHERE oci.conversation_id = ?
            ORDER BY oci.id_correo_importado DESC
            """,
            (conversation_id,),
        )
        row = cursor.fetchone()
        if row is not None:
            return {
                "id_oferta": row[0],
                "numero_oferta": row[1],
                "ref_cliente_asunto_email": row[2],
                "match_type": "conversation_id",
            }

    related_message_ids = []
    for key in ("in_reply_to_message_id", "reference_message_ids", "internet_message_id"):
        value = metadata.get(key)
        if isinstance(value, list):
            related_message_ids.extend(value)
        elif value:
            related_message_ids.append(value)

    related_message_ids = [item for index, item in enumerate(related_message_ids) if item and item not in related_message_ids[:index]]
    if not related_message_ids:
        return None

    placeholders = ", ".join("?" for _ in related_message_ids)
    cursor.execute(
        f"""
        SELECT TOP 1 lo.id_oferta, lo.numero_oferta, lo.ref_cliente_asunto_email
        FROM ofertas.oferta_correos_importados oci
        INNER JOIN ofertas.listado_ofertas lo
            ON lo.id_oferta = oci.id_oferta
        WHERE oci.internet_message_id IN ({placeholders})
        ORDER BY oci.id_correo_importado DESC
        """,
        tuple(related_message_ids),
    )
    row = cursor.fetchone()
    if row is None:
        return None

    return {
        "id_oferta": row[0],
        "numero_oferta": row[1],
        "ref_cliente_asunto_email": row[2],
        "match_type": "message_reference",
    }


def imported_email_already_registered(cursor, oferta_id, metadata):
    if not oferta_email_tracking_table_exists(cursor) or not metadata:
        return False

    internet_message_id = metadata.get("internet_message_id")
    if internet_message_id:
        cursor.execute(
            """
            SELECT TOP 1 1
            FROM ofertas.oferta_correos_importados
            WHERE id_oferta = ?
              AND internet_message_id = ?
            """,
            (oferta_id, internet_message_id),
        )
        if cursor.fetchone() is not None:
            return True

    body_sha256 = metadata.get("body_sha256")
    if body_sha256:
        cursor.execute(
            """
            SELECT TOP 1 1
            FROM ofertas.oferta_correos_importados
            WHERE id_oferta = ?
              AND body_sha256 = ?
            """,
            (oferta_id, body_sha256),
        )
        if cursor.fetchone() is not None:
            return True

    return False


def register_imported_email_metadata(cursor, oferta_id, parsed_email):
    if not oferta_email_tracking_table_exists(cursor):
        return False

    metadata = get_imported_email_metadata_for_match(parsed_email)
    if metadata is None or imported_email_already_registered(cursor, oferta_id, metadata):
        return False

    cursor.execute(
        """
        INSERT INTO ofertas.oferta_correos_importados (
            id_oferta,
            source_type,
            conversation_id,
            internet_message_id,
            in_reply_to_message_id,
            reference_message_ids_json,
            subject,
            sender_name,
            sender_email,
            received_at,
            body_sha256,
            fecha_registro
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, SYSUTCDATETIME())
        """,
        (
            oferta_id,
            metadata.get("source_type"),
            metadata.get("conversation_id"),
            metadata.get("internet_message_id"),
            metadata.get("in_reply_to_message_id"),
            json.dumps(metadata.get("reference_message_ids") or []),
            metadata.get("subject"),
            metadata.get("sender_name"),
            metadata.get("sender_email"),
            normalize_email_datetime(metadata.get("received_at")),
            metadata.get("body_sha256"),
        ),
    )
    return True


def discard_staged_imported_email_attachments(token):
    if not token:
        return
    target_dir = find_imported_email_attachments_dir(token)
    shutil.rmtree(target_dir, ignore_errors=True)
    cleanup_imported_email_attachment_bucket(target_dir)


def split_offer_conversation_segments(body_text):
    normalized = str(body_text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized:
        return []

    lines = normalized.split("\n")
    segments = []
    current_lines = []
    hard_divider_regex = re.compile(
        r"^(?:-{2,}\s*(?:original message|mensaje original|quoted message|p[ůu]vodn[ií]\s*e[-‑–—]?mail)\s*-*|_{5,})$",
        re.IGNORECASE,
    )
    wrote_regex = re.compile(r"^on\s.+wrote:$", re.IGNORECASE)
    header_line_regex = re.compile(
        r"^(from|de|sent|enviado|date|datum|to|para|cc|subject|asunto|od|odeslano|odesláno|komu|predmet|předmět)\s*:",
        re.IGNORECASE,
    )

    def push_segment():
        text = "\n".join(current_lines).strip()
        if text:
            segments.append(text)
        current_lines.clear()

    for index, line in enumerate(lines):
        trimmed = line.strip()

        if not trimmed:
            if current_lines and current_lines[-1] != "":
                current_lines.append("")
            continue

        next_lines = [item.strip() for item in lines[index:index + 4] if item.strip()]
        header_matches = sum(1 for item in next_lines if header_line_regex.match(item))
        starts_quoted_block = hard_divider_regex.match(trimmed) or wrote_regex.match(trimmed) or header_matches >= 2

        if starts_quoted_block:
            push_segment()
            continue

        current_lines.append(trimmed)

    push_segment()
    return segments or [normalized]


def normalize_offer_conversation_segment(segment):
    header_line_regex = re.compile(
        r"^(from|de|sent|enviado(?:\s+el)?|date|datum|to|para|cc|subject|asunto|od|odeslano|odesláno|komu|predmet|předmět)\s*:",
        re.IGNORECASE,
    )
    original_marker_regex = re.compile(
        r"^(?:-{2,}\s*(?:original message|mensaje original|quoted message|p[ůu]vodn[ií]\s*e[-‑–—]?mail)\s*-*|_{5,})$",
        re.IGNORECASE,
    )

    normalized_lines = []
    raw_segment = unicodedata.normalize("NFKC", str(segment or ""))
    raw_segment = raw_segment.replace("\u00a0", " ").replace("\u200b", "")
    raw_segment = raw_segment.replace("‑", "-").replace("–", "-").replace("—", "-")

    for raw_line in raw_segment.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"<mailto:([^>]+)>", r"\1", line, flags=re.IGNORECASE)
        line = re.sub(r"\s+", " ", line).strip()
        if header_line_regex.match(line) or original_marker_regex.match(line):
            continue
        normalized_lines.append(line)

    return "\n".join(normalized_lines).strip().lower()


def build_offer_conversation_segment_fingerprints(segment):
    normalized_segment = normalize_offer_conversation_segment(segment)
    if not normalized_segment:
        return set()

    paragraphs = [
        re.sub(r"\s+", " ", paragraph).strip()
        for paragraph in re.split(r"\n{2,}", normalized_segment)
    ]
    substantial_paragraphs = [paragraph for paragraph in paragraphs if len(paragraph) >= 80]

    if not substantial_paragraphs and len(normalized_segment) >= 160:
        substantial_paragraphs = [re.sub(r"\s+", " ", normalized_segment).strip()]

    return {
        hashlib.sha256(paragraph.encode("utf-8")).hexdigest()
        for paragraph in substantial_paragraphs
        if paragraph
    }


def build_offer_conversation_match_context(body_text):
    segments = split_offer_conversation_segments(body_text)
    normalized_segments = []
    normalized_segment_keys = set()
    paragraph_fingerprints = set()

    for segment in segments:
        normalized_segment = normalize_offer_conversation_segment(segment)
        if not normalized_segment:
            continue
        normalized_segments.append(normalized_segment)
        normalized_segment_keys.add(normalized_segment)
        paragraph_fingerprints.update(build_offer_conversation_segment_fingerprints(segment))

    return {
        "normalized_segments": normalized_segments,
        "normalized_segment_keys": normalized_segment_keys,
        "paragraph_fingerprints": paragraph_fingerprints,
    }


def offer_conversation_segment_already_exists(segment, existing_context):
    normalized_segment = normalize_offer_conversation_segment(segment)
    if not normalized_segment:
        return True

    existing_segment_keys = existing_context.get("normalized_segment_keys") or set()
    if normalized_segment in existing_segment_keys:
        return True

    existing_normalized_segments = existing_context.get("normalized_segments") or []
    if any(
        len(existing_segment) >= 120 and len(normalized_segment) >= 120
        and (normalized_segment in existing_segment or existing_segment in normalized_segment)
        for existing_segment in existing_normalized_segments
    ):
        return True

    candidate_fingerprints = build_offer_conversation_segment_fingerprints(segment)
    existing_fingerprints = existing_context.get("paragraph_fingerprints") or set()
    if candidate_fingerprints and candidate_fingerprints.issubset(existing_fingerprints):
        return True

    return False


def prepend_imported_message_to_observaciones(existing_observaciones, new_body):
    new_segments = split_offer_conversation_segments(new_body)
    if not new_segments:
        return str(existing_observaciones or "").strip()

    existing_text = str(existing_observaciones or "").strip()
    existing_context = build_offer_conversation_match_context(existing_text)

    merged_new_segments = []
    for segment in new_segments:
        if offer_conversation_segment_already_exists(segment, existing_context):
            continue

        normalized_segment = normalize_offer_conversation_segment(segment)
        if normalized_segment:
            existing_context["normalized_segments"].append(normalized_segment)
            existing_context["normalized_segment_keys"].add(normalized_segment)
            existing_context["paragraph_fingerprints"].update(build_offer_conversation_segment_fingerprints(segment))
        merged_new_segments.append(segment.strip())

    if not merged_new_segments:
        return existing_text

    merged_new_text = "\n\n----- quoted message -----\n\n".join(merged_new_segments)
    if not existing_text:
        return merged_new_text

    return f"{merged_new_text}\n\n----- quoted message -----\n\n{existing_text}"


def imported_email_adds_new_offer_content(existing_observaciones, new_body):
    new_segments = split_offer_conversation_segments(new_body)
    if not new_segments:
        return False

    existing_context = build_offer_conversation_match_context(existing_observaciones)

    for segment in new_segments:
        if not offer_conversation_segment_already_exists(segment, existing_context):
            return True

    return False


def sync_imported_emails_into_offer(cursor, oferta_id, parsed_emails, imported_email_attachment_token=None):
    if not ensure_offer_exists(cursor, oferta_id):
        raise ValueError("Oferta no encontrada")

    cursor.execute(
        """
        SELECT numero_oferta, observaciones, fecha_email, ref_cliente_asunto_email, nombre_emisor, email_emisor
        FROM ofertas.listado_ofertas
        WHERE id_oferta = ?
        """,
        (oferta_id,),
    )
    offer_row = cursor.fetchone()
    if offer_row is None:
        raise ValueError("Oferta no encontrada")

    numero_oferta, observaciones_actuales, _, _, _, _ = offer_row
    ordered_messages = sorted(
        [item for item in (parsed_emails or []) if isinstance(item, dict)],
        key=lambda item: normalize_email_datetime(item.get("received_at")) or datetime.min,
    )

    imported_count = 0
    skipped_count = 0
    latest_message = None
    current_observaciones = observaciones_actuales

    for parsed_email in ordered_messages:
        metadata = get_imported_email_metadata_for_match(parsed_email)
        if metadata is not None and imported_email_already_registered(cursor, oferta_id, metadata):
            skipped_count += 1
            continue

        if not imported_email_adds_new_offer_content(current_observaciones, parsed_email.get("body")):
            skipped_count += 1
            continue

        current_observaciones = prepend_imported_message_to_observaciones(current_observaciones, parsed_email.get("body"))
        register_imported_email_metadata(cursor, oferta_id, parsed_email)
        latest_message = parsed_email
        imported_count += 1

    attachments = []
    if imported_count > 0:
        latest_received_at = normalize_email_datetime((latest_message or {}).get("received_at"))
        cursor.execute(
            """
            UPDATE ofertas.listado_ofertas
            SET fecha_email = ?,
                ref_cliente_asunto_email = ?,
                observaciones = ?,
                nombre_emisor = ?,
                email_emisor = ?
            WHERE id_oferta = ?
            """,
            (
                latest_received_at.date() if latest_received_at else None,
                normalize_optional_text((latest_message or {}).get("subject"), 500),
                current_observaciones,
                normalize_optional_text((latest_message or {}).get("sender_name"), 255),
                normalize_optional_text((latest_message or {}).get("sender_email"), 255),
                oferta_id,
            ),
        )
        if imported_email_attachment_token:
            attachments = move_staged_imported_email_attachments_to_offer(
                oferta_id,
                imported_email_attachment_token,
                numero_oferta=numero_oferta,
            )
    elif imported_email_attachment_token:
        discard_staged_imported_email_attachments(imported_email_attachment_token)

    return {
        "id_oferta": oferta_id,
        "numero_oferta": numero_oferta,
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "adjuntos": attachments,
    }


def build_import_result_payload(response_data=None, sync_result=None, message=None, message_key=None, message_params=None):
    if sync_result is not None:
        return {
            "success": True,
            "mode": "synced_existing_offer",
            "message": message,
            "message_key": message_key,
            "message_params": message_params or {},
            **sync_result,
        }

    return {
        "success": True,
        "mode": "form_prefill",
        "message": message,
        "message_key": message_key,
        "message_params": message_params or {},
        "data": response_data or {},
    }


def build_oferta_payload(data):
    sender_identity = extract_sender_identity(
        data.get("emisor") or format_sender_display(data.get("nombre_emisor"), data.get("email_emisor")),
        fallback_email=data.get("email_emisor"),
    )

    if sender_identity["sender_name"] is None:
        sender_identity["sender_name"] = normalize_optional_text(data.get("nombre_emisor"), 255)

    if sender_identity["sender_email"] is None:
        raise ValueError("El campo Quién lo envía debe incluir un correo electrónico")

    sender_display = format_sender_display(sender_identity["sender_name"], sender_identity["sender_email"])

    return {
        "fecha_email": normalize_required_date(data.get("fecha_email"), "fecha_email"),
        "fecha_alta_oferta": normalize_optional_date(data.get("fecha_alta_oferta"), "fecha_alta_oferta"),
        "ref_cliente_asunto_email": normalize_optional_text(data.get("ref_cliente_asunto_email"), 500),
        "id_cliente": normalize_required_int(data.get("id_cliente") if data.get("id_cliente") is not None else data.get("cliente"), "Cliente"),
        "id_bom": normalize_optional_int(data.get("id_bom"), "BOM"),
        "observaciones": normalize_optional_text(data.get("observaciones")),
        "nombre_emisor": sender_identity["sender_name"],
        "email_emisor": sender_identity["sender_email"],
        "emisor": sender_display,
    }


def build_oferta_update_payload(data):
    payload = build_oferta_payload(data)
    payload["id_estado"] = normalize_required_int(data.get("id_estado"), "Estado")
    return payload


def build_oferta_etc_payload(data):
    payload = {
        "fecha_recepcion": normalize_optional_date(data.get("fecha_recepcion"), "fecha_recepcion"),
        "fecha_envio_oferta": normalize_optional_date(data.get("fecha_envio_oferta"), "fecha_envio_oferta"),
        "fecha_limite_respuesta": normalize_optional_date(data.get("fecha_limite_respuesta"), "fecha_limite_respuesta"),
        "id_estado": normalize_optional_int(data.get("id_estado"), "Estado") or 1,
        "id_cliente": normalize_optional_int(data.get("id_cliente") if data.get("id_cliente") is not None else data.get("cliente"), "Cliente"),
        "num_operario_responsable": normalize_required_int(data.get("num_operario_responsable"), "Num_operario_responsable"),
        "id_departamento_destino": normalize_required_int(data.get("id_departamento_destino"), "id_departamento_destino"),
        "codigo_externo_oferta": normalize_required_text(data.get("codigo_externo_oferta"), "Código externo", 100),
        "codigo_interno_oferta": normalize_optional_text(data.get("codigo_interno_oferta"), 100),
        "referencia_cliente": normalize_required_text(data.get("referencia_cliente"), "Referencia cliente", 255),
        "numero_comision": normalize_optional_text(data.get("numero_comision"), 100),
        "po_original": normalize_optional_text(data.get("po_original"), 100),
        "pedido_b2b": normalize_optional_text(data.get("pedido_b2b"), 100),
        "proyecto": normalize_optional_text(data.get("proyecto"), 255),
        "sales_orders": normalize_optional_text(data.get("sales_orders"), 100),
        "request_delivery_date": normalize_optional_date(data.get("request_delivery_date"), "request_delivery_date"),
        "nombre_solicitante": normalize_optional_text(data.get("nombre_solicitante"), 255),
        "email_solicitante": normalize_optional_text(data.get("email_solicitante"), 255),
        "empresa_solicitante": normalize_optional_text(data.get("empresa_solicitante"), 255),
        "incoterm": normalize_optional_text(data.get("incoterm"), 50),
        "moneda": (normalize_optional_text(data.get("moneda"), 3) or "EUR").upper(),
        "prioridad": normalize_etc_priority(data.get("prioridad")),
        "es_urgente": normalize_optional_bool(data.get("es_urgente"), "es_urgente"),
        "resumen_material_solicitado": normalize_optional_text(data.get("resumen_material_solicitado")),
        "resumen_material_ofertado": normalize_optional_text(data.get("resumen_material_ofertado")),
        "total_material_eur": normalize_optional_decimal(data.get("total_material_eur"), "total_material_eur"),
        "total_fee_eur": normalize_optional_decimal(data.get("total_fee_eur"), "total_fee_eur"),
        "observaciones_cliente": normalize_optional_text(data.get("observaciones_cliente")),
        "observaciones_tecnicas": normalize_optional_text(data.get("observaciones_tecnicas")),
        "observaciones_internas": normalize_optional_text(data.get("observaciones_internas")),
        "origen_registro": (normalize_optional_text(data.get("origen_registro"), 20) or "MANUAL").upper(),
        "activo": normalize_optional_bool(data.get("activo"), "activo"),
    }

    if payload["es_urgente"] is None:
        payload["es_urgente"] = False
    if payload["activo"] is None:
        payload["activo"] = True

    identifying_values = (
        payload["codigo_externo_oferta"],
        payload["codigo_interno_oferta"],
        payload["referencia_cliente"],
        payload["numero_comision"],
        payload["proyecto"],
    )
    if not any(identifying_values):
        raise ValueError(
            "Debes indicar al menos un identificador para la oferta ETC: código externo, código interno, referencia de cliente, número de comisión o proyecto"
        )

    return payload


def build_oferta_estado_payload(data):
    return {
        "id_estado": normalize_required_int(data.get("id_estado"), "Estado siguiente"),
        "fecha_limite": normalize_optional_date(data.get("fecha_limite"), "Fecha límite"),
        "comentario": normalize_optional_text(data.get("comentario")),
    }


def build_cliente_payload(data):
    return {
        "descripcion_cliente": normalize_required_text(data.get("descripcion_cliente"), "Descripción cliente", 255),
        "dominio": normalize_cliente_domain(data.get("dominio")),
    }


def build_proyecto_payload(data):
    return {
        "descripcion_proyecto": normalize_required_text(data.get("descripcion_proyecto"), "Descripción proyecto", 255),
    }


def build_estado_payload(data):
    payload = {
        "descripcion_estado": normalize_required_text(data.get("descripcion_estado"), "Descripción estado", 255),
        "orden": normalize_optional_int(data.get("orden"), "orden"),
        "emoji_sidebar": normalize_optional_text(data.get("emoji_sidebar"), 8),
        "activo": normalize_optional_bool(data.get("activo"), "activo"),
    }

    if payload["activo"] is None:
        payload["activo"] = True

    return payload


def build_usuario_payload(data):
    role_id = normalize_optional_int(data.get("id_rol"), "id_rol")
    role_name = normalize_optional_text(data.get("rol") or data.get("nombre_rol"), 100)
    if role_id is None and role_name is None:
        raise ValueError("El campo Rol es obligatorio")

    return {
        "num_operario": normalize_required_int(data.get("num_operario"), "Num_operario"),
        "nombre": normalize_required_text(data.get("nombre"), "Nombre", 255),
        "email": normalize_optional_email(data.get("email"), "Email"),
        "id_rol": role_id,
        "rol": role_name,
        "id_departamento": normalize_optional_int(data.get("id_departamento"), "id_departamento"),
    }


def get_available_offer_columns():
    return [
        {"value": "id_oferta", "label": "ID oferta", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "estado", "label": "Estado", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "fecha_email", "label": "Fecha e-mail", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "fecha_alta_oferta", "label": "Fecha alta oferta", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "fecha_limite", "label": "Fecha límite", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "fecha_envio_oferta", "label": "Fecha envío oferta", "source": "ofertas.oferta_etc"},
        {"value": "numero_oferta", "label": "Nº oferta", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "ref_cliente_asunto_email", "label": "Ref. cliente / asunto e-mail", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "cliente", "label": "Cliente", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "emisor", "label": "Emisor", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "observaciones_oferta", "label": "Observaciones oferta", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "tipo_interaccion", "label": "Tipos interacción", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "fecha_interaccion", "label": "Fechas interacción", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "observaciones_interaccion", "label": "Observaciones interacción", "source": "ofertas.vw_listado_ofertas_interacciones"},
        {"value": "nombre_responsable", "label": "Responsable", "source": "ofertas.oferta_etc"},
        {"value": "nombre_departamento_destino", "label": "Departamento destino", "source": "ofertas.oferta_etc"},
        {"value": "codigo_externo_oferta", "label": "Código externo", "source": "ofertas.oferta_etc"},
        {"value": "codigo_interno_oferta", "label": "Material number", "source": "ofertas.oferta_etc"},
        {"value": "referencia_cliente", "label": "Referencia cliente", "source": "ofertas.oferta_etc"},
        {"value": "numero_comision", "label": "Commission number", "source": "ofertas.oferta_etc"},
        {"value": "proyecto", "label": "Proyecto", "source": "ofertas.oferta_etc"},
        {"value": "nombre_solicitante", "label": "Nombre solicitante", "source": "ofertas.oferta_etc"},
        {"value": "email_solicitante", "label": "Email solicitante", "source": "ofertas.oferta_etc"},
        {"value": "incoterm", "label": "Incoterm", "source": "ofertas.oferta_etc"},
        {"value": "prioridad", "label": "Prioridad", "source": "ofertas.oferta_etc"},
        {"value": "total_material_eur", "label": "Total material EUR", "source": "ofertas.oferta_etc"},
        {"value": "total_fee_eur", "label": "Total fee EUR", "source": "ofertas.oferta_etc"},
        {"value": "observaciones_cliente", "label": "Observaciones cliente", "source": "ofertas.oferta_etc"},
        {"value": "pedido_b2b", "label": "Pedido B2B", "source": "ofertas.oferta_etc"},
        {"value": "po_original", "label": "PO original", "source": "ofertas.oferta_etc"},
        {"value": "sales_orders", "label": "Sales Orders", "source": "ofertas.oferta_etc"},
        {"value": "request_delivery_date", "label": "Request Delivery Date", "source": "ofertas.oferta_etc"},
    ]


def get_available_offer_column_map():
    return {column["value"]: column for column in get_available_offer_columns()}


def normalize_offer_column_name(column_name):
    normalized = normalize_optional_text(column_name, 100)
    if normalized is None:
        return None

    legacy_map = {
        "id_oferta": "id_oferta",
        "id_estado": "estado",
        "Estado": "estado",
        "fecha_email": "fecha_email",
        "fecha_alta_oferta": "fecha_alta_oferta",
        "fecha_limite": "fecha_limite",
        "numero_oferta": "numero_oferta",
        "ref_cliente_asunto_email": "ref_cliente_asunto_email",
        "Cliente": "cliente",
        "Emisor": "emisor",
        "Observaciones": "observaciones_oferta",
        "Observaciones_oferta": "observaciones_oferta",
        "Oferta_Interacciones.Tipo_interaccion": "tipo_interaccion",
        "Tipo_interaccion": "tipo_interaccion",
        "Oferta_Interacciones.fecha_interaccion": "fecha_interaccion",
        "fecha_interaccion": "fecha_interaccion",
        "Oferta_Interacciones.Observaciones": "observaciones_interaccion",
        "Observaciones_interaccion": "observaciones_interaccion",
        "Concepto": None,
        "columna1": None,
        "id_cliente": None,
        "Oferta_Interacciones.id_interaccion": None,
    }

    return legacy_map.get(normalized, normalized.lower())


def build_configuracion_columna_payload(data):
    return {
        "columna": normalize_required_text(data.get("columna"), "columna", 100),
        "descripcion_columna": normalize_optional_text(data.get("descripcion_columna"), 255),
        "orden_columna": normalize_optional_int(data.get("orden_columna"), "orden columna"),
    }


GLOBAL_CONFIG_SCOPE_ID = -1
GLOBAL_CONFIG_SCOPE_LABEL = "__GLOBAL_COLUMN_SCOPE__"
DEFAULT_GLOBAL_CONFIG_COLUMNS = [
    "numero_oferta",
    "fecha_alta_oferta",
    "cliente",
    "emisor",
    "observaciones_oferta",
    "estado",
    "codigo_externo_oferta",
    "sales_orders",
    "request_delivery_date",
]

def is_global_config_scope(estado_id):
    return estado_id == GLOBAL_CONFIG_SCOPE_ID

def parse_config_scope_id(raw_value):
    try:
        estado_id = int(str(raw_value).strip())
    except (TypeError, ValueError):
        raise ValueError("Estado no válido")

    if estado_id == GLOBAL_CONFIG_SCOPE_ID:
        return estado_id
    if estado_id <= 0:
        raise ValueError("Estado no válido")
    return estado_id

def ensure_config_scope_exists(cursor, estado_id):
    if is_global_config_scope(estado_id):
        ensure_global_config_scope_row(cursor)
        return True

    cursor.execute(
        "SELECT 1 FROM ofertas.estados WHERE id_estado = ?",
        (estado_id,),
    )
    return cursor.fetchone() is not None


def ensure_global_config_scope_row(cursor):
    cursor.execute(
        "SELECT 1 FROM ofertas.estados WHERE id_estado = ?",
        (GLOBAL_CONFIG_SCOPE_ID,),
    )
    if cursor.fetchone() is not None:
        return

    cursor.execute("SET IDENTITY_INSERT ofertas.estados ON")
    try:
        cursor.execute(
            """
            INSERT INTO ofertas.estados (id_estado, descripcion_estado, orden, id_departamento, emoji_sidebar, activo)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (GLOBAL_CONFIG_SCOPE_ID, GLOBAL_CONFIG_SCOPE_LABEL, None, None, '', 0),
        )
    finally:
        cursor.execute("SET IDENTITY_INSERT ofertas.estados OFF")


def ensure_global_config_default_columns(cursor):
    ensure_global_config_scope_row(cursor)
    cursor.execute(
        "SELECT COUNT(*) FROM ofertas.configuracioncolumnas WHERE id_estado = ?",
        (GLOBAL_CONFIG_SCOPE_ID,),
    )
    existing_count = cursor.fetchone()[0] or 0
    if existing_count:
        return

    available_column_map = get_available_offer_column_map()
    for index, column_name in enumerate(DEFAULT_GLOBAL_CONFIG_COLUMNS, start=1):
        column_info = available_column_map.get(column_name)
        if not column_info:
            continue

        cursor.execute(
            """
            INSERT INTO ofertas.configuracioncolumnas (id_estado, columna, descripcion_columna, orden_columna)
            VALUES (?, ?, ?, ?)
            """,
            (GLOBAL_CONFIG_SCOPE_ID, column_name, column_info["label"], index),
        )



def get_next_numero_oferta(cursor):
    current_year = datetime.now().year
    cursor.execute(
        """
        SELECT ISNULL(MAX(TRY_CONVERT(INT, RIGHT(numero_oferta, 5))), 0)
        FROM ofertas.listado_ofertas
        WHERE LEFT(ISNULL(numero_oferta, ''), 4) = ?
        """,
        (str(current_year),),
    )
    current_max = cursor.fetchone()[0] or 0
    return f"{current_year}{current_max + 1:05d}"


def serialize_date(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def serialize_decimal(value):
    if value is None:
        return None
    return float(value)


def get_history_actor_label(user_data):
    if not isinstance(user_data, dict):
        return None

    display_name = normalize_optional_text(user_data.get("nombre") or user_data.get("display_name"), 255)
    email = normalize_optional_text(user_data.get("email"), 255)
    num_operario = normalize_optional_text(user_data.get("num_operario"), 50)

    if display_name and num_operario:
        return f"{display_name} ({num_operario})"
    if display_name:
        return display_name
    if email and num_operario:
        return f"{email} ({num_operario})"
    if email:
        return email
    if num_operario:
        return f"Operario {num_operario}"
    return None


def format_history_change_value(value):
    if value is None:
        return "Sin valor"
    if isinstance(value, bool):
        return "Si" if value else "No"
    if isinstance(value, Decimal):
        normalized = value.quantize(Decimal("0.01"))
        return format(normalized, "f").rstrip("0").rstrip(".") or "0"
    if isinstance(value, float):
        return format(value, ".2f").rstrip("0").rstrip(".") or "0"
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            pass

    normalized = normalize_optional_text(value)
    return normalized if normalized is not None else "Sin valor"


def build_history_change_entries(before_snapshot, after_snapshot, field_labels, actor_label=None):
    if not before_snapshot or not after_snapshot:
        return []

    actor_suffix = f"\nActualizado por: {actor_label}" if actor_label else ""
    entries = []
    for field_name, label in field_labels:
        before_value = format_history_change_value(before_snapshot.get(field_name))
        after_value = format_history_change_value(after_snapshot.get(field_name))
        if before_value == after_value:
            continue

        observation = f"{label}: {before_value} -> {after_value}{actor_suffix}"
        if len(observation) > 500:
            observation = f"{observation[:497].rstrip()}..."
        entries.append(observation)

    return entries


def insert_offer_interaction_entry(cursor, oferta_id, interaction_type, interaction_date, observaciones=None, fecha_limite=None):
    cursor.execute(
        """
        INSERT INTO ofertas.oferta_interacciones (
            id_oferta,
            tipo_interaccion,
            fecha_interaccion,
            fecha_limite,
            observaciones
        )
        VALUES (?, ?, ?, ?, ?)
        """,
        (oferta_id, interaction_type, interaction_date, fecha_limite, observaciones),
    )


def get_offer_audit_snapshot(cursor, oferta_id):
    cursor.execute(
        """
        SELECT
            e.descripcion_estado,
            lo.fecha_email,
            lo.fecha_alta_oferta,
            lo.ref_cliente_asunto_email,
            c.descripcion_cliente,
            b.material,
            lo.observaciones,
            lo.nombre_emisor,
            lo.email_emisor
        FROM ofertas.listado_ofertas lo
        LEFT JOIN ofertas.estados e
            ON e.id_estado = lo.id_estado
        LEFT JOIN ofertas.clientes c
            ON c.id_cliente = lo.id_cliente
        LEFT JOIN ofertas.materiales_precio b
            ON b.id_material_precio = lo.id_bom
        WHERE lo.id_oferta = ?
        """,
        (oferta_id,),
    )
    row = cursor.fetchone()
    if row is None:
        return None

    return {
        "estado": row[0],
        "fecha_email": row[1],
        "fecha_alta_oferta": row[2],
        "ref_cliente_asunto_email": row[3],
        "cliente": row[4],
        "bom": row[5],
        "observaciones": row[6],
        "emisor": format_sender_display(row[7], row[8]),
    }


def get_offer_etc_audit_snapshot(cursor, oferta_etc_id):
    cursor.execute(
        """
        SELECT
            o.fecha_recepcion,
            o.fecha_envio_oferta,
            o.fecha_limite_respuesta,
            e.descripcion_estado,
            c.descripcion_cliente,
            COALESCE(NULLIF(LTRIM(RTRIM(gu.nombre)), ''), CAST(o.num_operario_responsable AS NVARCHAR(50))),
            d.nombre_departamento,
            o.codigo_externo_oferta,
            o.codigo_interno_oferta,
            o.referencia_cliente,
            o.numero_comision,
            o.po_original,
            o.pedido_b2b,
            o.sales_orders,
            o.request_delivery_date,
            o.proyecto,
            o.nombre_solicitante,
            o.email_solicitante,
            o.empresa_solicitante,
            o.incoterm,
            o.moneda,
            o.prioridad,
            o.es_urgente,
            o.resumen_material_solicitado,
            o.resumen_material_ofertado,
            o.total_material_eur,
            o.total_fee_eur,
            o.observaciones_cliente,
            o.observaciones_tecnicas,
            o.observaciones_internas,
            o.origen_registro,
            o.activo
        FROM ofertas.oferta_etc o
        LEFT JOIN ofertas.estados e
            ON e.id_estado = o.id_estado
        LEFT JOIN ofertas.clientes c
            ON c.id_cliente = o.id_cliente
        LEFT JOIN general.usuarios gu
            ON gu.num_operario = o.num_operario_responsable
        LEFT JOIN ofertas.departamentos d
            ON d.id_departamento = o.id_departamento_destino
        WHERE o.id_oferta_etc = ?
        """,
        (oferta_etc_id,),
    )
    row = cursor.fetchone()
    if row is None:
        return None

    return {
        "fecha_recepcion": row[0],
        "fecha_envio_oferta": row[1],
        "fecha_limite_respuesta": row[2],
        "estado": row[3],
        "cliente": row[4],
        "responsable": row[5],
        "departamento_destino": row[6],
        "codigo_externo_oferta": row[7],
        "codigo_interno_oferta": row[8],
        "referencia_cliente": row[9],
        "numero_comision": row[10],
        "po_original": row[11],
        "pedido_b2b": row[12],
        "sales_orders": row[13],
        "request_delivery_date": row[14],
        "proyecto": row[15],
        "nombre_solicitante": row[16],
        "email_solicitante": row[17],
        "empresa_solicitante": row[18],
        "incoterm": row[19],
        "moneda": row[20],
        "prioridad": row[21],
        "es_urgente": bool(row[22]) if row[22] is not None else False,
        "resumen_material_solicitado": row[23],
        "resumen_material_ofertado": row[24],
        "total_material_eur": row[25],
        "total_fee_eur": row[26],
        "observaciones_cliente": row[27],
        "observaciones_tecnicas": row[28],
        "observaciones_internas": row[29],
        "origen_registro": row[30],
        "activo": bool(row[31]) if row[31] is not None else True,
    }


def materiales_precio_table_exists(cursor):
    cursor.execute("SELECT 1 WHERE OBJECT_ID('ofertas.materiales_precio', 'U') IS NOT NULL")
    return cursor.fetchone() is not None


def materiales_precio_has_fecha_creacion(cursor):
    cursor.execute("SELECT 1 WHERE COL_LENGTH('ofertas.materiales_precio', 'fecha_creacion') IS NOT NULL")
    return cursor.fetchone() is not None


def oferta_has_bom_column(cursor):
    cursor.execute("SELECT 1 WHERE COL_LENGTH('ofertas.listado_ofertas', 'id_bom') IS NOT NULL")
    return cursor.fetchone() is not None


BOM_CATALOG_COLUMN_DEFINITIONS = (
    {"key": "part_nr", "header": "Part Nr", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "mat_description", "header": "Mat Description", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "new_sales_price", "header": "New Sales Price", "sql_type": "DECIMAL(18, 2)", "max_length": None},
    {"key": "notas", "header": "NOTAS", "sql_type": "NVARCHAR(500)", "max_length": 500},
    {"key": "nuevos_cp357", "header": "NUEVOS CP357", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "nuevos_cp361", "header": "NUEVOS CP361", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "nuevos_cp365", "header": "NUEVOS CP365", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "anulados_cp365", "header": "ANULADOS CP365", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "nuevos_cp369", "header": "NUEVOS CP369", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "anulados_cp369", "header": "ANULADOS CP369", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "nuevos_cp373", "header": "NUEVOS CP373", "sql_type": "NVARCHAR(255)", "max_length": 255},
    {"key": "anulados_cp373", "header": "ANULADOS CP373", "sql_type": "NVARCHAR(255)", "max_length": 255},
)

BOM_CATALOG_COLUMNS_BY_KEY = {column["key"]: column for column in BOM_CATALOG_COLUMN_DEFINITIONS}
BOM_CATALOG_COLUMNS_BY_HEADER = {
    re.sub(r"\s+", " ", str(column["header"]).strip()).lower(): column
    for column in BOM_CATALOG_COLUMN_DEFINITIONS
}
BOM_REQUIRED_IMPORT_KEYS = ("part_nr", "mat_description", "new_sales_price")


def materiales_precio_has_column(cursor, column_name):
    cursor.execute(
        f"SELECT 1 WHERE COL_LENGTH('ofertas.materiales_precio', '{column_name}') IS NOT NULL"
    )
    return cursor.fetchone() is not None


def ensure_materiales_precio_catalog_columns(cursor):
    for column in BOM_CATALOG_COLUMN_DEFINITIONS:
        if not materiales_precio_has_column(cursor, column["key"]):
            cursor.execute(
                f"ALTER TABLE ofertas.materiales_precio ADD {column['key']} {column['sql_type']} NULL"
            )


def normalize_bom_identifier(value):
    normalized = normalize_optional_text(value, 255)
    return normalized.lower() if normalized else None


def build_bom_catalog_record(row):
    record = {
        "id_bom": row[0],
        "id_material_precio": row[0],
        "material": row[1],
        "precio": serialize_decimal(row[2]),
        "fecha_creacion": serialize_date(row[3]),
        "part_nr": row[4],
        "mat_description": row[5],
        "new_sales_price": serialize_decimal(row[6]),
        "notas": row[7],
        "nuevos_cp357": row[8],
        "nuevos_cp361": row[9],
        "nuevos_cp365": row[10],
        "anulados_cp365": row[11],
        "nuevos_cp369": row[12],
        "anulados_cp369": row[13],
        "nuevos_cp373": row[14],
        "anulados_cp373": row[15],
        "precio_anterior": None,
        "diferencia_precio": None,
    }
    return record


def build_bom_payload(data):
    part_nr = normalize_required_text(data.get("part_nr"), "Part Nr", 255)
    mat_description = normalize_required_text(data.get("mat_description"), "Mat Description", 255)
    new_sales_price = normalize_optional_decimal(data.get("new_sales_price"), "New Sales Price")
    if new_sales_price is None:
        raise ValueError("El campo New Sales Price es obligatorio")
    if new_sales_price < 0:
        raise ValueError("El campo New Sales Price no puede ser negativo")

    return {
        "part_nr": part_nr,
        "mat_description": mat_description,
        "new_sales_price": new_sales_price.quantize(Decimal("0.01")),
        "notas": normalize_optional_text(data.get("notas"), 500),
        "material": mat_description,
        "precio": new_sales_price.quantize(Decimal("0.01")),
        "nuevos_cp357": None,
        "nuevos_cp361": None,
        "nuevos_cp365": None,
        "anulados_cp365": None,
        "nuevos_cp369": None,
        "anulados_cp369": None,
        "nuevos_cp373": None,
        "anulados_cp373": None,
    }


def normalize_bom_csv_header(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def decode_uploaded_csv(uploaded_file):
    file_bytes = uploaded_file.read()
    if not file_bytes:
        raise ValueError("El archivo CSV está vacío")

    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise ValueError("No se pudo leer el archivo CSV con una codificación compatible")


def parse_bom_catalog_rows(rows, source_label="archivo"):
    if not rows:
        raise ValueError(f"El {source_label} no contiene datos")

    raw_headers = rows[0]
    header_index_map = {}
    for index, raw_header in enumerate(raw_headers):
        normalized_header = normalize_bom_csv_header(raw_header)
        column = BOM_CATALOG_COLUMNS_BY_HEADER.get(normalized_header)
        if column and column["key"] not in header_index_map:
            header_index_map[column["key"]] = index

    missing_keys = [key for key in BOM_REQUIRED_IMPORT_KEYS if key not in header_index_map]
    if missing_keys:
        missing_headers = ", ".join(BOM_CATALOG_COLUMNS_BY_KEY[key]["header"] for key in missing_keys)
        raise ValueError(f"El {source_label} no contiene las columnas obligatorias: {missing_headers}")

    imported_rows = []
    seen_identifiers = set()
    for row_number, row in enumerate(rows[1:], start=2):
        raw_part_nr = row[header_index_map["part_nr"]] if header_index_map["part_nr"] < len(row) else ""
        raw_mat_description = row[header_index_map["mat_description"]] if header_index_map["mat_description"] < len(row) else ""
        raw_new_sales_price = row[header_index_map["new_sales_price"]] if header_index_map["new_sales_price"] < len(row) else ""

        # Some Excel exports include section rows with only a title in the first column.
        # They are not BOM materials and should be ignored during import.
        if normalize_optional_text(raw_part_nr, 255) and not normalize_optional_text(raw_mat_description, 255) and not normalize_optional_text(raw_new_sales_price):
            continue

        row_values = {}
        for column in BOM_CATALOG_COLUMN_DEFINITIONS:
            raw_value = row[header_index_map[column["key"]]] if column["key"] in header_index_map and header_index_map[column["key"]] < len(row) else ""
            if column["key"] == "new_sales_price":
                decimal_value = normalize_optional_decimal(raw_value, column["header"])
                if decimal_value is None:
                    raise ValueError(f"La columna {column['header']} es obligatoria en la fila {row_number}")
                if decimal_value < 0:
                    raise ValueError(f"La columna {column['header']} no puede ser negativa en la fila {row_number}")
                row_values[column["key"]] = decimal_value.quantize(Decimal("0.01"))
            else:
                row_values[column["key"]] = normalize_optional_text(raw_value, column["max_length"])

        if not any(value not in (None, "") for value in row_values.values()):
            continue

        if row_values["part_nr"] is None:
            raise ValueError(f"La columna Part Nr es obligatoria en la fila {row_number}")
        if row_values["mat_description"] is None:
            raise ValueError(f"La columna Mat Description es obligatoria en la fila {row_number}")

        identifier = normalize_bom_identifier(row_values["part_nr"])
        if identifier in seen_identifiers:
            raise ValueError(f"El CSV contiene un Part Nr repetido: {row_values['part_nr']}")
        seen_identifiers.add(identifier)

        row_values["material"] = row_values["mat_description"]
        row_values["precio"] = row_values["new_sales_price"]
        imported_rows.append(row_values)

    if not imported_rows:
        raise ValueError(f"El {source_label} no contiene filas válidas para importar")

    return imported_rows


def parse_bom_catalog_csv(uploaded_file):
    csv_text = decode_uploaded_csv(uploaded_file)
    lines = csv_text.splitlines()
    if not lines:
        raise ValueError("El archivo CSV no contiene datos")

    header_line = lines[0]
    if ';' in header_line:
        delimiter = ';'
    elif '\t' in header_line:
        delimiter = '\t'
    else:
        delimiter = ','

    reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
    rows = list(reader)
    return parse_bom_catalog_rows(rows, source_label="archivo CSV")


def parse_bom_catalog_excel(uploaded_file):
    file_bytes = uploaded_file.read()
    if not file_bytes:
        raise ValueError("El archivo Excel está vacío")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("No está disponible el soporte para importar archivos Excel (.xlsx/.xlsm)") from exc

    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {str(exc)}") from exc

    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            normalized_row = []
            for value in row:
                if value is None:
                    normalized_row.append("")
                else:
                    normalized_row.append(str(value).strip())
            if any(cell != "" for cell in normalized_row):
                rows.append(normalized_row)

        if not rows:
            continue

        try:
            return parse_bom_catalog_rows(rows, source_label=f"hoja '{worksheet.title}'")
        except ValueError as exc:
            if "no contiene las columnas obligatorias" in str(exc):
                continue
            raise

    raise ValueError("El archivo Excel no contiene ninguna hoja con las columnas BOM esperadas")


def upsert_bom_catalog_row(cursor, payload, bom_id=None):
    values = [
        payload["material"],
        payload["precio"],
        payload["part_nr"],
        payload["mat_description"],
        payload["new_sales_price"],
        payload.get("notas"),
        payload.get("nuevos_cp357"),
        payload.get("nuevos_cp361"),
        payload.get("nuevos_cp365"),
        payload.get("anulados_cp365"),
        payload.get("nuevos_cp369"),
        payload.get("anulados_cp369"),
        payload.get("nuevos_cp373"),
        payload.get("anulados_cp373"),
    ]

    if bom_id is None:
        cursor.execute(
            """
            INSERT INTO ofertas.materiales_precio (
                material,
                precio,
                part_nr,
                mat_description,
                new_sales_price,
                notas,
                nuevos_cp357,
                nuevos_cp361,
                nuevos_cp365,
                anulados_cp365,
                nuevos_cp369,
                anulados_cp369,
                nuevos_cp373,
                anulados_cp373
            )
            OUTPUT INSERTED.id_material_precio
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            tuple(values),
        )
        inserted_row = cursor.fetchone()
        return inserted_row[0] if inserted_row else None

    cursor.execute(
        """
        UPDATE ofertas.materiales_precio
        SET material = ?,
            precio = ?,
            part_nr = ?,
            mat_description = ?,
            new_sales_price = ?,
            notas = ?,
            nuevos_cp357 = ?,
            nuevos_cp361 = ?,
            nuevos_cp365 = ?,
            anulados_cp365 = ?,
            nuevos_cp369 = ?,
            anulados_cp369 = ?,
            nuevos_cp373 = ?,
            anulados_cp373 = ?
        WHERE id_material_precio = ?
        """,
        tuple(values + [bom_id]),
    )
    return bom_id


def find_conflicting_bom(cursor, payload, exclude_bom_id=None):
    conditions = [
        "(part_nr IS NOT NULL AND LTRIM(RTRIM(part_nr)) <> '' AND LOWER(LTRIM(RTRIM(part_nr))) = LOWER(LTRIM(RTRIM(?))))",
        "(material IS NOT NULL AND LTRIM(RTRIM(material)) <> '' AND LOWER(LTRIM(RTRIM(material))) = LOWER(LTRIM(RTRIM(?))))",
    ]
    params = [payload["part_nr"], payload["material"]]
    where_clause = " OR ".join(conditions)
    query = f"SELECT TOP 1 id_material_precio, material, part_nr FROM ofertas.materiales_precio WHERE ({where_clause})"
    if exclude_bom_id is not None:
        query += " AND id_material_precio <> ?"
        params.append(exclude_bom_id)

    cursor.execute(query, tuple(params))
    return cursor.fetchone()


def replace_bom_catalog_from_rows(cursor, imported_rows):
    ensure_offer_bom_price_override_schema(cursor)

    cursor.execute(
        """
        SELECT
            id_material_precio,
            material,
            part_nr,
            mat_description
        FROM ofertas.materiales_precio
        """
    )
    existing_rows = cursor.fetchall()

    existing_by_part_nr = {}
    existing_by_material = {}
    for row in existing_rows:
        bom_id = int(row[0])
        part_nr_key = normalize_bom_identifier(row[2])
        material_key = normalize_bom_identifier(row[1] or row[3])
        if part_nr_key:
            existing_by_part_nr[part_nr_key] = bom_id
        if material_key:
            existing_by_material[material_key] = bom_id

    matched_ids = set()
    inserted_count = 0
    updated_count = 0
    deleted_count = 0
    detached_links = 0
    deleted_override_count = 0
    affected_offer_count = 0

    for payload in imported_rows:
        part_nr_key = normalize_bom_identifier(payload.get("part_nr"))
        material_key = normalize_bom_identifier(payload.get("material"))
        matched_bom_id = None
        if part_nr_key and part_nr_key in existing_by_part_nr:
            matched_bom_id = existing_by_part_nr[part_nr_key]
        elif material_key and material_key in existing_by_material:
            matched_bom_id = existing_by_material[material_key]

        if matched_bom_id is not None:
            upsert_bom_catalog_row(cursor, payload, bom_id=matched_bom_id)
            matched_ids.add(matched_bom_id)
            updated_count += 1
        else:
            inserted_bom_id = upsert_bom_catalog_row(cursor, payload, bom_id=None)
            if inserted_bom_id is not None:
                matched_ids.add(int(inserted_bom_id))
            inserted_count += 1

    obsolete_bom_ids = [int(row[0]) for row in existing_rows if int(row[0]) not in matched_ids]
    for obsolete_bom_id in obsolete_bom_ids:
        delete_result = delete_bom_catalog_material(cursor, obsolete_bom_id)
        deleted_count += 1
        detached_links += int(delete_result.get("detached_links", 0) or 0)
        deleted_override_count += int(delete_result.get("deleted_override_count", 0) or 0)
        affected_offer_count += int(delete_result.get("affected_offer_count", 0) or 0)

    return {
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "deleted_count": deleted_count,
        "detached_links": detached_links,
        "deleted_override_count": deleted_override_count,
        "affected_offer_count": affected_offer_count,
    }


def format_bom_csv_decimal(value):
    serialized = serialize_decimal(value)
    if serialized is None:
        return ""
    return str(serialized).replace(".", ",")


def build_bom_catalog_csv_content(rows):
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow([column["header"] for column in BOM_CATALOG_COLUMN_DEFINITIONS])
    for row in rows:
        writer.writerow([
            row.get("part_nr") or "",
            row.get("mat_description") or row.get("material") or "",
            format_bom_csv_decimal(row.get("new_sales_price") if row.get("new_sales_price") is not None else row.get("precio")),
            row.get("notas") or "",
            row.get("nuevos_cp357") or "",
            row.get("nuevos_cp361") or "",
            row.get("nuevos_cp365") or "",
            row.get("anulados_cp365") or "",
            row.get("nuevos_cp369") or "",
            row.get("anulados_cp369") or "",
            row.get("nuevos_cp373") or "",
            row.get("anulados_cp373") or "",
        ])
    return buffer.getvalue()


def ensure_bom_catalog_schema(cursor):
    if not materiales_precio_table_exists(cursor):
        raise ValueError("La tabla ofertas.materiales_precio no existe todavía")

    if not materiales_precio_has_fecha_creacion(cursor):
        raise ValueError(
            "La tabla ofertas.materiales_precio debe incluir la columna fecha_creacion antes de usar BOM en ofertas"
        )

    ensure_materiales_precio_catalog_columns(cursor)

    if not oferta_has_bom_column(cursor):
        cursor.execute("ALTER TABLE ofertas.listado_ofertas ADD id_bom INT NULL")

    cursor.execute(
        """
        SELECT OBJECT_NAME(referenced_object_id)
        FROM sys.foreign_keys
        WHERE name = 'fk_listado_ofertas_bom'
          AND parent_object_id = OBJECT_ID('ofertas.listado_ofertas')
        """
    )
    fk_row = cursor.fetchone()
    if fk_row and fk_row[0] != 'materiales_precio':
        cursor.execute(
            "ALTER TABLE ofertas.listado_ofertas DROP CONSTRAINT fk_listado_ofertas_bom"
        )
        fk_row = None

    if fk_row is None:
        cursor.execute(
            """
            ALTER TABLE ofertas.listado_ofertas WITH CHECK
            ADD CONSTRAINT fk_listado_ofertas_bom
            FOREIGN KEY (id_bom) REFERENCES ofertas.materiales_precio(id_material_precio)
            """
        )


def ensure_offer_bom_links_schema(cursor):
    ensure_bom_catalog_schema(cursor)
    cursor.execute(
        """
        IF OBJECT_ID('ofertas.oferta_bom_materiales', 'U') IS NULL
        BEGIN
            CREATE TABLE ofertas.oferta_bom_materiales (
                id_oferta INT NOT NULL,
                id_material_precio INT NOT NULL,
                fecha_asignacion DATETIME2(0) NOT NULL CONSTRAINT df_oferta_bom_materiales_fecha_asignacion DEFAULT (SYSDATETIME()),
                cantidad INT NOT NULL CONSTRAINT df_oferta_bom_materiales_cantidad DEFAULT (1),
                CONSTRAINT pk_oferta_bom_materiales PRIMARY KEY CLUSTERED (id_oferta ASC, id_material_precio ASC),
                CONSTRAINT fk_oferta_bom_materiales_oferta FOREIGN KEY (id_oferta) REFERENCES ofertas.listado_ofertas(id_oferta),
                CONSTRAINT fk_oferta_bom_materiales_material FOREIGN KEY (id_material_precio) REFERENCES ofertas.materiales_precio(id_material_precio)
            )
        END
        ELSE IF COL_LENGTH('ofertas.oferta_bom_materiales', 'cantidad') IS NULL
        BEGIN
            ALTER TABLE ofertas.oferta_bom_materiales
            ADD cantidad INT NOT NULL CONSTRAINT df_oferta_bom_materiales_cantidad DEFAULT (1) WITH VALUES
        END
        """
    )
    cursor.execute(
        """
        INSERT INTO ofertas.oferta_bom_materiales (id_oferta, id_material_precio)
        SELECT lo.id_oferta, lo.id_bom
        FROM ofertas.listado_ofertas lo
        LEFT JOIN ofertas.oferta_bom_materiales obm
            ON obm.id_oferta = lo.id_oferta
           AND obm.id_material_precio = lo.id_bom
        WHERE lo.id_bom IS NOT NULL
          AND obm.id_oferta IS NULL
        """
    )


def ensure_offer_bom_price_override_schema(cursor):
    ensure_offer_bom_links_schema(cursor)
    cursor.execute(
        """
        IF OBJECT_ID('ofertas.oferta_bom_precio_override', 'U') IS NULL
        BEGIN
            CREATE TABLE ofertas.oferta_bom_precio_override (
                id_override INT IDENTITY(1,1) NOT NULL,
                id_oferta INT NOT NULL,
                id_material_precio INT NOT NULL,
                precio_catalogo_snapshot DECIMAL(18, 2) NOT NULL,
                precio_oferta DECIMAL(18, 2) NOT NULL,
                num_operario INT NULL,
                comentario NVARCHAR(500) NULL,
                activo BIT NOT NULL CONSTRAINT df_oferta_bom_precio_override_activo DEFAULT ((1)),
                fecha_creacion DATETIME2(0) NOT NULL CONSTRAINT df_oferta_bom_precio_override_fecha_creacion DEFAULT (SYSDATETIME()),
                fecha_actualizacion DATETIME2(0) NOT NULL CONSTRAINT df_oferta_bom_precio_override_fecha_actualizacion DEFAULT (SYSDATETIME()),
                CONSTRAINT pk_oferta_bom_precio_override PRIMARY KEY CLUSTERED (id_override ASC),
                CONSTRAINT fk_oferta_bom_precio_override_oferta FOREIGN KEY (id_oferta) REFERENCES ofertas.listado_ofertas(id_oferta),
                CONSTRAINT fk_oferta_bom_precio_override_material FOREIGN KEY (id_material_precio) REFERENCES ofertas.materiales_precio(id_material_precio)
            )
        END
        """
    )
    cursor.execute(
        """
        IF NOT EXISTS (
            SELECT 1
            FROM sys.indexes
            WHERE name = 'ux_oferta_bom_precio_override_activo'
              AND object_id = OBJECT_ID('ofertas.oferta_bom_precio_override')
        )
        BEGIN
            CREATE UNIQUE INDEX ux_oferta_bom_precio_override_activo
            ON ofertas.oferta_bom_precio_override (id_oferta, id_material_precio)
            WHERE activo = 1
        END
        """
    )


def get_offer_bom_material_catalog_state(cursor, oferta_id, material_id):
    ensure_offer_bom_price_override_schema(cursor)
    cursor.execute(
        """
        SELECT TOP 1
            obm.id_oferta,
            obm.id_material_precio,
            mp.material,
            mp.precio,
            mp.fecha_creacion
        FROM ofertas.oferta_bom_materiales obm
        INNER JOIN ofertas.materiales_precio mp
            ON mp.id_material_precio = obm.id_material_precio
        WHERE obm.id_oferta = ?
          AND obm.id_material_precio = ?
        """,
        (oferta_id, material_id),
    )
    row = cursor.fetchone()
    if row is None:
        raise ValueError("El material BOM no esta asociado a la oferta")

    return {
        "id_oferta": row[0],
        "id_material_precio": row[1],
        "material": row[2],
        "precio_catalogo": row[3],
        "fecha_catalogo": row[4],
    }


def upsert_offer_bom_material_price_override(cursor, oferta_id, material_id, precio_oferta, user_data=None, comentario=None):
    catalog_state = get_offer_bom_material_catalog_state(cursor, oferta_id, material_id)

    precio_normalizado = precio_oferta.quantize(Decimal("0.01"))
    if precio_normalizado < 0:
        raise ValueError("El precio BOM de la oferta no puede ser negativo")

    num_operario = None
    try:
        num_operario = int((user_data or {}).get("num_operario"))
    except (TypeError, ValueError, AttributeError):
        num_operario = None

    comentario_normalizado = normalize_optional_text(comentario, 500)

    cursor.execute(
        """
        SELECT TOP 1 id_override
        FROM ofertas.oferta_bom_precio_override
        WHERE id_oferta = ?
          AND id_material_precio = ?
          AND activo = 1
        ORDER BY id_override DESC
        """,
        (oferta_id, material_id),
    )
    existing_row = cursor.fetchone()

    if existing_row is None:
        cursor.execute(
            """
            INSERT INTO ofertas.oferta_bom_precio_override (
                id_oferta,
                id_material_precio,
                precio_catalogo_snapshot,
                precio_oferta,
                num_operario,
                comentario,
                activo,
                fecha_actualizacion
            )
            VALUES (?, ?, ?, ?, ?, ?, 1, SYSDATETIME())
            """,
            (
                oferta_id,
                material_id,
                catalog_state["precio_catalogo"],
                precio_normalizado,
                num_operario,
                comentario_normalizado,
            ),
        )
        created = True
    else:
        cursor.execute(
            """
            UPDATE ofertas.oferta_bom_precio_override
            SET precio_catalogo_snapshot = ?,
                precio_oferta = ?,
                num_operario = ?,
                comentario = ?,
                activo = 1,
                fecha_actualizacion = SYSDATETIME()
            WHERE id_override = ?
            """,
            (
                catalog_state["precio_catalogo"],
                precio_normalizado,
                num_operario,
                comentario_normalizado,
                existing_row[0],
            ),
        )
        created = False

    return {
        "created": created,
        "material": catalog_state["material"],
        "precio_catalogo": catalog_state["precio_catalogo"],
        "precio_oferta": precio_normalizado,
    }


def remove_offer_bom_material_price_override(cursor, oferta_id, material_id):
    catalog_state = get_offer_bom_material_catalog_state(cursor, oferta_id, material_id)

    cursor.execute(
        """
        UPDATE ofertas.oferta_bom_precio_override
        SET activo = 0,
            fecha_actualizacion = SYSDATETIME()
        WHERE id_oferta = ?
          AND id_material_precio = ?
          AND activo = 1
        """,
        (oferta_id, material_id),
    )

    return {
        "removed": cursor.rowcount > 0,
        "material": catalog_state["material"],
        "precio_catalogo": catalog_state["precio_catalogo"],
    }


def load_offer_bom_materials_map(cursor, offer_ids):
    normalized_ids = []
    for offer_id in offer_ids or []:
        try:
            normalized_ids.append(int(offer_id))
        except (TypeError, ValueError):
            continue

    if not normalized_ids:
        return {}

    ensure_offer_bom_price_override_schema(cursor)

    placeholders = ", ".join(["?"] * len(normalized_ids))
    cursor.execute(
        f"""
        WITH
        active_overrides AS (
            SELECT
                o.id_oferta,
                o.id_material_precio,
                o.precio_catalogo_snapshot,
                o.precio_oferta,
                o.fecha_actualizacion
            FROM ofertas.oferta_bom_precio_override o
            WHERE o.activo = 1
        )
        SELECT
            obm.id_oferta,
            obm.id_material_precio,
            mp.material,
            mp.part_nr,
            COALESCE(ao.precio_oferta, mp.precio) AS precio,
            CASE
                WHEN ao.id_material_precio IS NOT NULL THEN ao.fecha_actualizacion
                ELSE mp.fecha_creacion
            END AS fecha_creacion,
            obm.fecha_asignacion,
            obm.cantidad,
            mp.precio AS precio_catalogo,
            mp.fecha_creacion AS fecha_catalogo,
            ao.precio_catalogo_snapshot,
            ao.precio_oferta,
            CAST(CASE WHEN ao.id_material_precio IS NULL THEN 0 ELSE 1 END AS BIT) AS tiene_precio_override,
            CASE
                WHEN ao.id_material_precio IS NULL THEN NULL
                ELSE ao.precio_oferta - mp.precio
            END AS diferencia_precio
        FROM ofertas.oferta_bom_materiales obm
        INNER JOIN ofertas.materiales_precio mp
            ON mp.id_material_precio = obm.id_material_precio
        LEFT JOIN active_overrides ao
            ON ao.id_oferta = obm.id_oferta
           AND ao.id_material_precio = obm.id_material_precio
        WHERE obm.id_oferta IN ({placeholders})
        ORDER BY obm.id_oferta ASC, obm.fecha_asignacion DESC, mp.material ASC
        """,
        tuple(normalized_ids),
    )

    materials_by_offer = {offer_id: [] for offer_id in normalized_ids}
    for row in cursor.fetchall():
        materials_by_offer.setdefault(row[0], []).append(
            {
                "id_material_precio": row[1],
                "material": row[2],
                "part_nr": row[3],
                "precio": serialize_decimal(row[4]),
                "fecha_creacion": serialize_date(row[5]),
                "fecha_asignacion": serialize_date(row[6]),
                "cantidad": row[7] if row[7] is not None else 1,
                "precio_catalogo": serialize_decimal(row[8]),
                "fecha_catalogo": serialize_date(row[9]),
                "precio_catalogo_snapshot": serialize_decimal(row[10]),
                "precio_oferta": serialize_decimal(row[11]),
                "tiene_precio_override": bool(row[12]),
                "precio_anterior": None,
                "diferencia_precio": serialize_decimal(row[13]),
            }
        )

    return materials_by_offer


def list_offer_bom_materials(cursor, oferta_id):
    return load_offer_bom_materials_map(cursor, [oferta_id]).get(int(oferta_id), [])


def sync_offer_primary_bom(cursor, oferta_id):
    cursor.execute(
        """
        SELECT TOP 1 obm.id_material_precio
        FROM ofertas.oferta_bom_materiales obm
        WHERE obm.id_oferta = ?
        ORDER BY obm.fecha_asignacion DESC, obm.id_material_precio DESC
        """,
        (oferta_id,),
    )
    row = cursor.fetchone()
    cursor.execute(
        "UPDATE ofertas.listado_ofertas SET id_bom = ? WHERE id_oferta = ?",
        (row[0] if row else None, oferta_id),
    )
    return row[0] if row else None


def add_offer_bom_material_link(cursor, oferta_id, material_id):
    ensure_offer_bom_links_schema(cursor)
    ensure_offer_bom_price_override_schema(cursor)

    if not ensure_offer_exists(cursor, oferta_id):
        raise ValueError("Oferta no encontrada")

    cursor.execute(
        "SELECT material, precio FROM ofertas.materiales_precio WHERE id_material_precio = ?",
        (material_id,),
    )
    material_row = cursor.fetchone()
    if material_row is None:
        raise ValueError("BOM no encontrado")

    cursor.execute(
        "SELECT 1 FROM ofertas.oferta_bom_materiales WHERE id_oferta = ? AND id_material_precio = ?",
        (oferta_id, material_id),
    )
    already_exists = cursor.fetchone() is not None

    if not already_exists:
        precio_catalogo = material_row[1]

        cursor.execute(
            "INSERT INTO ofertas.oferta_bom_materiales (id_oferta, id_material_precio, cantidad) VALUES (?, ?, 1)",
            (oferta_id, material_id),
        )

        # Crear override automático para congelar el precio en el momento de la vinculación.
        # Así, si el precio del catálogo cambia después, esta oferta conserva el precio original.
        cursor.execute(
            """
            INSERT INTO ofertas.oferta_bom_precio_override (
                id_oferta, id_material_precio,
                precio_catalogo_snapshot, precio_oferta,
                num_operario, comentario, activo,
                fecha_actualizacion
            )
            VALUES (?, ?, ?, ?, NULL, N'Precio inicial al vincular BOM', 1, SYSDATETIME())
            """,
            (oferta_id, material_id, precio_catalogo, precio_catalogo),
        )

    sync_offer_primary_bom(cursor, oferta_id)
    return {
        "added": not already_exists,
        "material": material_row[0],
        "cantidad": 1,
    }


def remove_offer_bom_material_link(cursor, oferta_id, material_id):
    ensure_offer_bom_links_schema(cursor)
    ensure_offer_bom_price_override_schema(cursor)

    if not ensure_offer_exists(cursor, oferta_id):
        raise ValueError("Oferta no encontrada")

    cursor.execute(
        "SELECT material FROM ofertas.materiales_precio WHERE id_material_precio = ?",
        (material_id,),
    )
    material_row = cursor.fetchone()
    if material_row is None:
        raise ValueError("BOM no encontrado")

    cursor.execute(
        "DELETE FROM ofertas.oferta_bom_materiales WHERE id_oferta = ? AND id_material_precio = ?",
        (oferta_id, material_id),
    )
    removed = cursor.rowcount > 0
    if removed:
        cursor.execute(
            """
            UPDATE ofertas.oferta_bom_precio_override
            SET activo = 0,
                fecha_actualizacion = SYSDATETIME()
            WHERE id_oferta = ?
              AND id_material_precio = ?
              AND activo = 1
            """,
            (oferta_id, material_id),
        )
    sync_offer_primary_bom(cursor, oferta_id)
    return {
        "removed": removed,
        "material": material_row[0],
    }


def delete_bom_catalog_material(cursor, bom_id):
    ensure_offer_bom_price_override_schema(cursor)

    cursor.execute(
        "SELECT material FROM ofertas.materiales_precio WHERE id_material_precio = ?",
        (bom_id,),
    )
    material_row = cursor.fetchone()
    if material_row is None:
        raise ValueError("BOM no encontrado")

    cursor.execute(
        "SELECT DISTINCT id_oferta FROM ofertas.oferta_bom_materiales WHERE id_material_precio = ?",
        (bom_id,),
    )
    affected_offer_ids = [int(row[0]) for row in cursor.fetchall()]

    cursor.execute(
        "SELECT COUNT(*) FROM ofertas.oferta_bom_precio_override WHERE id_material_precio = ?",
        (bom_id,),
    )
    override_count_row = cursor.fetchone()
    override_count = int(override_count_row[0]) if override_count_row else 0

    cursor.execute(
        "DELETE FROM ofertas.oferta_bom_precio_override WHERE id_material_precio = ?",
        (bom_id,),
    )

    cursor.execute(
        "DELETE FROM ofertas.oferta_bom_materiales WHERE id_material_precio = ?",
        (bom_id,),
    )
    detached_links = cursor.rowcount

    cursor.execute(
        "UPDATE ofertas.listado_ofertas SET id_bom = NULL WHERE id_bom = ?",
        (bom_id,),
    )

    for offer_id in affected_offer_ids:
        sync_offer_primary_bom(cursor, offer_id)

    cursor.execute(
        "DELETE FROM ofertas.materiales_precio WHERE id_material_precio = ?",
        (bom_id,),
    )
    if cursor.rowcount == 0:
        raise ValueError("BOM no encontrado")

    return {
        "material": material_row[0],
        "affected_offer_count": len(affected_offer_ids),
        "detached_links": detached_links,
        "deleted_override_count": override_count,
    }


def attach_offer_bom_materials(cursor, offers):
    if not offers:
        return offers

    materials_by_offer = load_offer_bom_materials_map(cursor, [offer.get("id_oferta") for offer in offers])
    for offer in offers:
        selected_materials = materials_by_offer.get(offer.get("id_oferta"), [])
        offer["bom_materiales"] = selected_materials
        offer["nombre_bom"] = ", ".join(
            material["material"]
            for material in selected_materials
            if material.get("material")
        ) or offer.get("nombre_bom")

    return offers


def build_bom_payload(data):
    return {
        "part_nr": normalize_required_text(data.get("part_nr"), "Part Nr", 255),
        "mat_description": normalize_required_text(data.get("mat_description"), "Mat Description", 255),
        "new_sales_price": normalize_optional_decimal(data.get("new_sales_price"), "New Sales Price"),
        "notas": normalize_optional_text(data.get("notas"), 500),
    }


def build_bom_mutation_payload(data):
    payload = build_bom_payload(data)
    price = payload.get("new_sales_price")
    if price is None:
        raise ValueError("El campo New Sales Price es obligatorio")
    if price < 0:
        raise ValueError("El campo New Sales Price no puede ser negativo")

    payload["new_sales_price"] = price.quantize(Decimal("0.01"))
    payload["material"] = payload["mat_description"]
    payload["precio"] = payload["new_sales_price"]
    payload["nuevos_cp357"] = None
    payload["nuevos_cp361"] = None
    payload["nuevos_cp365"] = None
    payload["anulados_cp365"] = None
    payload["nuevos_cp369"] = None
    payload["anulados_cp369"] = None
    payload["nuevos_cp373"] = None
    payload["anulados_cp373"] = None
    return payload


def list_boms_catalog(cursor, only_active=False):
    materiales = list_materiales_precio_snapshot(cursor)
    return materiales


def build_material_precio_payload(data):
    payload = build_bom_mutation_payload(
        {
            "part_nr": data.get("part_nr") or data.get("material"),
            "mat_description": data.get("mat_description") or data.get("material"),
            "new_sales_price": data.get("new_sales_price") or data.get("precio"),
            "notas": data.get("notas"),
        }
    )
    return payload


def list_materiales_precio_snapshot(cursor):
    cursor.execute(
        """
        SELECT
            mp.id_material_precio,
            mp.material,
            mp.precio,
            mp.fecha_creacion,
            mp.part_nr,
            mp.mat_description,
            mp.new_sales_price,
            mp.notas,
            mp.nuevos_cp357,
            mp.nuevos_cp361,
            mp.nuevos_cp365,
            mp.anulados_cp365,
            mp.nuevos_cp369,
            mp.anulados_cp369,
            mp.nuevos_cp373,
            mp.anulados_cp373
        FROM ofertas.materiales_precio mp
        WHERE mp.material IS NOT NULL
          AND LTRIM(RTRIM(mp.material)) <> ''
        ORDER BY COALESCE(mp.part_nr, mp.material) ASC, mp.id_material_precio ASC
        """
    )

    materiales = []
    for row in cursor.fetchall():
        materiales.append(build_bom_catalog_record(row))

    return materiales


def serialize_interaction_date_list(value):
    if value is None:
        return None
    return str(value)


def oferta_etc_table_exists(cursor):
    cursor.execute("SELECT 1 WHERE OBJECT_ID('ofertas.oferta_etc', 'U') IS NOT NULL")
    return cursor.fetchone() is not None


def oferta_email_subject_exists(cursor, fecha_email, ref_cliente_asunto_email, excluded_id=None):
    query = """
        SELECT id_oferta, ref_cliente_asunto_email
        FROM ofertas.listado_ofertas
        WHERE fecha_email = ?
    """
    params = [fecha_email]

    if excluded_id is not None:
        query += " AND id_oferta <> ?"
        params.append(excluded_id)

    cursor.execute(query, tuple(params))
    normalized_input_subject = normalize_email_subject_for_matching(ref_cliente_asunto_email)

    for row in cursor.fetchall():
        if normalize_email_subject_for_matching(row[1]) == normalized_input_subject:
            return True

    return False


def cliente_exists(cursor, descripcion_cliente, excluded_id=None):
    query = """
        SELECT TOP 1 id_cliente
        FROM ofertas.clientes
        WHERE LTRIM(RTRIM(descripcion_cliente)) = LTRIM(RTRIM(?))
    """
    params = [descripcion_cliente]

    if excluded_id is not None:
        query += " AND id_cliente <> ?"
        params.append(excluded_id)

    cursor.execute(query, tuple(params))
    return cursor.fetchone() is not None


def cliente_domain_exists(cursor, dominio, excluded_id=None):
    if not dominio:
        return False

    query = """
        SELECT TOP 1 id_cliente
        FROM ofertas.clientes
        WHERE LOWER(LTRIM(RTRIM(ISNULL(dominio, '')))) = LOWER(LTRIM(RTRIM(?)))
    """
    params = [dominio]

    if excluded_id is not None:
        query += " AND id_cliente <> ?"
        params.append(excluded_id)

    cursor.execute(query, tuple(params))
    return cursor.fetchone() is not None


def proyecto_exists(cursor, descripcion_proyecto, excluded_id=None):
    query = """
        SELECT TOP 1 id_proyecto
        FROM ofertas.proyectos
        WHERE LTRIM(RTRIM(descripcion_proyecto)) = LTRIM(RTRIM(?))
    """
    params = [descripcion_proyecto]

    if excluded_id is not None:
        query += " AND id_proyecto <> ?"
        params.append(excluded_id)

    cursor.execute(query, tuple(params))
    return cursor.fetchone() is not None


def estado_exists(cursor, descripcion_estado, excluded_id=None):
    query = """
        SELECT TOP 1 id_estado
        FROM ofertas.estados
        WHERE LTRIM(RTRIM(descripcion_estado)) = LTRIM(RTRIM(?))
    """
    params = [descripcion_estado]

    if excluded_id is not None:
        query += " AND id_estado <> ?"
        params.append(excluded_id)

    cursor.execute(query, tuple(params))
    return cursor.fetchone() is not None


def get_estado_metadata(cursor, estado_id):
    cursor.execute(
        """
        SELECT
            e.id_estado,
            e.descripcion_estado,
            ISNULL(e.activo, 1) AS activo,
            e.emoji_sidebar,
            e.id_departamento,
            d.nombre_departamento
        FROM ofertas.estados e
        LEFT JOIN ofertas.departamentos d
            ON d.id_departamento = e.id_departamento
        WHERE id_estado = ?
        """,
        (estado_id,),
    )
    row = cursor.fetchone()
    if row is None:
        return None

    return {
        "id_estado": row[0],
        "descripcion_estado": row[1],
        "activo": bool(row[2]),
        "emoji_sidebar": row[3],
        "id_departamento": row[4],
        "nombre_departamento": row[5],
    }


def resolve_active_estado_for_department(cursor, department_id, fallback_estado_id=None):
    normalized_department_id = normalize_optional_int(department_id, "Departamento")
    if normalized_department_id is None:
        return fallback_estado_id

    cursor.execute(
        """
        SELECT TOP 1 id_estado
        FROM ofertas.estados
        WHERE id_departamento = ?
          AND ISNULL(activo, 1) = 1
        ORDER BY CASE WHEN orden IS NULL THEN 1 ELSE 0 END,
                 orden ASC,
                 id_estado ASC
        """,
        (normalized_department_id,),
    )
    row = cursor.fetchone()
    if row is not None:
        return row[0]

    return fallback_estado_id


def sync_offer_estado_with_etc(cursor, oferta_id, estado_id):
    normalized_oferta_id = normalize_optional_int(oferta_id, "Oferta")
    normalized_estado_id = normalize_optional_int(estado_id, "Estado")
    if normalized_oferta_id is None or normalized_estado_id is None:
        return False

    cursor.execute(
        "UPDATE ofertas.listado_ofertas SET id_estado = ? WHERE id_oferta = ?",
        (normalized_estado_id, normalized_oferta_id),
    )
    return cursor.rowcount > 0


def configuracion_columna_exists(cursor, id_estado, columna, excluded_id=None):
    normalized_target = normalize_offer_column_name(columna)
    cursor.execute(
        """
        SELECT id_config, columna
        FROM ofertas.configuracioncolumnas
        WHERE id_estado = ?
        """,
        (id_estado,),
    )

    for row in cursor.fetchall():
        config_id = row[0]
        existing_column = normalize_offer_column_name(row[1])
        if excluded_id is not None and config_id == excluded_id:
            continue
        if existing_column == normalized_target:
            return True
    return False


@app.route("/")
def index():
    template_path = os.path.join(PROJECT_DIR, "templates", "index.html")
    with open(template_path, encoding="utf-8") as template_file:
        return render_template_string(template_file.read())


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify(
        {
            "success": True,
            "service": "OFERTAS",
            "port": app.config["APP_PORT"],
            "database": app.config["DB_DATABASE"],
            "db_password_configured": bool(app.config.get("DB_PASSWORD")),
        }
    )


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(silent=True) or {}
    usuario = (data.get("usuario") or "").strip()
    password = (data.get("password") or "").strip()

    if not usuario or not password:
        return jsonify({"success": False, "message": "Usuario y contraseña requeridos"}), 400

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT id_usuario, num_operario, nombre, nivel_permisos, roles, contrasena
                FROM general.usuarios
                WHERE num_operario = ?
                """,
                (usuario,),
            )
            result = cursor.fetchone()

            if not result:
                return jsonify({"success": False, "message": "Usuario no encontrado"}), 401

            if not verify_password(password, result[5]):
                return jsonify({"success": False, "message": "Contraseña incorrecta"}), 401

            user_data = build_authenticated_user_data(
                cursor,
                {
                    "id_usuario": result[0],
                    "num_operario": result[1],
                    "nombre": result[2],
                    "nivel_permisos": result[3],
                    "rol": result[4],
                },
            )

            persist_authenticated_user(user_data)

            return jsonify(user_data)
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"Error del servidor: {str(exc)}"}), 500


@app.route("/api/logout", methods=["POST"])
def api_logout():
    try:
        session.clear()
        return jsonify({"success": True, "message": "Sesión cerrada exitosamente"})
    except Exception as exc:
        return jsonify({"success": False, "message": f"Error del servidor: {str(exc)}"}), 500


@app.route("/auth/logout", methods=["GET"])
def auth_logout():
    session.clear()
    return redirect("/")


def start_microsoft_login():
    apply_local_oauth_env_overrides()
    session["microsoft_auth_context"] = "app"
    session["microsoft_post_auth_redirect"] = url_for("index")
    try:
        effective_redirect_uri = get_request_oauth_redirect_uri()
        print(
            f"[oauth] login env={app.config.get('APP_ENV')} "
            f"client_id={OutlookGraphService.get_config().get('client_id')} "
            f"tenant_id={OutlookGraphService.get_config().get('tenant_id')} "
            f"redirect_uri={effective_redirect_uri}",
            flush=True,
        )
        return redirect(
            OutlookGraphService.start_auth_flow(
                session,
                scopes=OutlookGraphService.get_login_scopes(),
                redirect_uri=effective_redirect_uri,
            )
        )
    except OutlookGraphError as exc:
        return redirect(url_for("index", auth_error=str(exc)))
    except Exception as exc:
        app.logger.exception("No se pudo iniciar el login de Microsoft")
        return redirect(url_for("index", auth_error=f"No se pudo iniciar la autenticación con Microsoft: {exc}"))


@app.route("/auth/microsoft/login", methods=["GET"])
def auth_microsoft_login():
    return start_microsoft_login()


@app.route("/auth/login", methods=["GET"])
def auth_login():
    return start_microsoft_login()


@app.route("/auth/microsoft/callback", methods=["GET"])
def auth_microsoft_callback():
    return auth_outlook_callback()


@app.route("/auth/callback", methods=["GET"])
def auth_callback():
    return auth_outlook_callback()


@app.route("/api/me", methods=["GET"])
def api_me():
    user_data = refresh_authenticated_session_user()
    if not user_data:
        return jsonify({"success": False, "message": "No hay sesión activa"}), 401
    return jsonify({"success": True, "user": user_data})


@app.route("/api/session/check", methods=["GET"])
def api_session_check():
    user_data = refresh_authenticated_session_user()
    return jsonify({"success": bool(user_data), "authenticated": bool(user_data), "user": user_data or None})


@app.route("/api/outlook/status", methods=["GET"])
def api_outlook_status():
    apply_local_oauth_env_overrides()
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para usar Outlook"}), 401

    try:
        status = OutlookGraphService.get_status(session)
        account = status.get("account") or {}
        status["mailbox"] = account.get("username")
        status["login_url"] = url_for("auth_outlook_login")
        status["disconnect_url"] = url_for("api_outlook_disconnect")
        return jsonify({"success": True, **status})
    except OutlookGraphError as exc:
        return jsonify({"success": False, "message": str(exc)}), 409
    except Exception as exc:
        app.logger.exception("No se pudo resolver el estado de Outlook")
        return jsonify({"success": False, "message": f"No se pudo consultar Outlook: {str(exc)}"}), 500


@app.route("/auth/outlook/login", methods=["GET"])
def auth_outlook_login():
    apply_local_oauth_env_overrides()
    user_data = get_logged_user_data()
    if not user_data:
        return redirect(url_for("index"))

    try:
        session["microsoft_auth_context"] = "outlook"
        should_open_outlook = str(request.args.get("open_outlook", "1")).strip().lower() not in {"0", "false", "no"}
        session["outlook_post_auth_redirect"] = url_for("index", open_outlook=1) if should_open_outlook else url_for("index")
        return redirect(OutlookGraphService.start_auth_flow(session, redirect_uri=get_request_oauth_redirect_uri()))
    except OutlookGraphError as exc:
        return redirect(url_for("index", outlook_error=str(exc)))
    except Exception as exc:
        app.logger.exception("No se pudo iniciar el login de Outlook")
        return redirect(url_for("index", outlook_error=f"No se pudo iniciar la autenticación de Outlook: {exc}"))


@app.route("/auth/outlook/callback", methods=["GET"])
def auth_outlook_callback():
    apply_local_oauth_env_overrides()
    auth_context = session.pop("microsoft_auth_context", "outlook")
    user_data = get_logged_user_data()
    if auth_context != "app" and not user_data:
        return redirect(url_for("index"))

    redirect_target = (
        session.pop("microsoft_post_auth_redirect", None)
        if auth_context == "app"
        else session.pop("outlook_post_auth_redirect", None)
    ) or (url_for("index") if auth_context == "app" else url_for("index", open_outlook=1))

    try:
        auth_result = OutlookGraphService.complete_auth_flow(
            session,
            dict(request.args),
            require_access_token=(auth_context != "app"),
        )
        if auth_context == "app":
            with db_connection(autocommit=True) as conn:
                cursor = conn.cursor()
                profile = auth_result.get("profile") or {}
                general_user = resolve_general_user_from_microsoft_profile(cursor, profile)
                if not general_user:
                    OutlookGraphService.disconnect(session)
                    session.pop("user_id", None)
                    session.pop("user_data", None)
                    return redirect(url_for("index", auth_error="No se ha encontrado tu usuario en general.usuarios para esta cuenta de Microsoft."))

                persist_authenticated_user(build_authenticated_user_data(cursor, general_user))
                session["microsoft_profile"] = profile
        return redirect(redirect_target)
    except OutlookGraphError as exc:
        error_param = "auth_error" if auth_context == "app" else "outlook_error"
        return redirect(url_for("index", **{error_param: str(exc)}))
    except RuntimeError as exc:
        error_param = "auth_error" if auth_context == "app" else "outlook_error"
        return redirect(url_for("index", **{error_param: str(exc)}))


@app.route("/api/outlook/disconnect", methods=["POST"])
def api_outlook_disconnect():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para desconectar Outlook"}), 401

    OutlookGraphService.disconnect(session)
    return jsonify({"success": True, "message": "La cuenta de Outlook se ha desconectado correctamente."})


@app.route("/api/outlook/messages", methods=["GET"])
def api_outlook_messages():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar Outlook"}), 401

    folder = request.args.get("folder", "inbox")
    top = request.args.get("top", default=20, type=int)

    try:
        result = OutlookGraphService.list_messages(session, folder=folder, top=top)
        return jsonify({"success": True, **result})
    except OutlookGraphError as exc:
        return jsonify({"success": False, "message": str(exc)}), 409
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo consultar Outlook: {str(exc)}"}), 500


@app.route("/api/outlook/messages/<path:message_id>", methods=["GET"])
def api_outlook_message_detail(message_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar Outlook"}), 401

    try:
        message = OutlookGraphService.get_message(session, message_id)
        parsed_email = normalize_outlook_message_for_offer(message)
        cliente_resolution = resolve_cliente_for_sender_email(parsed_email.get("sender_email"))
        import_data = build_imported_email_response_data(parsed_email, cliente_resolution)
        with db_connection(autocommit=True) as conn:
            thread_match = get_offer_thread_match(conn.cursor(), parsed_email)
        if thread_match is not None:
            import_data["thread_match"] = thread_match
        return jsonify(
            {
                "success": True,
                "message": "Correo de Outlook cargado correctamente.",
                "mailbox": ((message.get("account") or {}).get("username")),
                "outlook_message": message,
                "import_data": import_data,
            }
        )
    except OutlookGraphError as exc:
        return jsonify({"success": False, "message": str(exc)}), 409
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo cargar el correo de Outlook: {str(exc)}"}), 500


@app.route("/api/outlook/messages/<path:message_id>/import", methods=["POST"])
def api_outlook_import_message(message_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para importar correos de Outlook"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    try:
        message = OutlookGraphService.get_message(session, message_id)
        parsed_email = normalize_outlook_message_for_offer(message)
        cliente_resolution = resolve_cliente_for_sender_email(parsed_email.get("sender_email"))
        response_data = build_imported_email_response_data(parsed_email, cliente_resolution)

        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            thread_match = get_offer_thread_match(cursor, parsed_email)
            if thread_match is None:
                conn.commit()
                return jsonify(
                    build_import_result_payload(
                        response_data=response_data,
                        message="Correo de Outlook preparado para crear una oferta nueva.",
                    )
                )

            conversation_id = normalize_optional_text(parsed_email.get("conversation_id"), 255)
            parsed_messages = [parsed_email]
            if conversation_id:
                conversation_messages = OutlookGraphService.list_messages_by_conversation(session, conversation_id)
                parsed_messages = [normalize_outlook_message_for_offer(item) for item in conversation_messages]

            sync_result = sync_imported_emails_into_offer(cursor, thread_match["id_oferta"], parsed_messages)
            conn.commit()

        offer_reference = sync_result["numero_oferta"] or thread_match["id_oferta"]
        if sync_result["imported_count"] > 0:
            message_text = f"Se han importado {sync_result['imported_count']} correos nuevos en la oferta {offer_reference}."
            message_key = "offer.import_existing_offer_success"
        else:
            message_text = f"No había correos nuevos que añadir en la oferta {offer_reference}."
            message_key = "offer.import_existing_offer_duplicate"

        return jsonify(
            build_import_result_payload(
                sync_result=sync_result,
                message=message_text,
                message_key=message_key,
                message_params={"offer": str(offer_reference)},
            )
        )
    except OutlookGraphError as exc:
        return jsonify({"success": False, "message": str(exc)}), 409
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo importar el correo de Outlook: {str(exc)}"}), 500


@app.route("/api/outlook/send", methods=["POST"])
def api_outlook_send_mail():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para enviar correos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        result = OutlookGraphService.send_mail(
            session,
            subject=data.get("subject"),
            body=data.get("body"),
            to_recipients=data.get("to_recipients"),
            cc_recipients=data.get("cc_recipients"),
            is_html=bool(data.get("is_html", True)),
            save_to_sent_items=bool(data.get("save_to_sent_items", True)),
        )
        return jsonify({"success": True, "message": "Correo enviado correctamente con Outlook.", **result})
    except OutlookGraphError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo enviar el correo: {str(exc)}"}), 500


@app.route("/api/ofertas/siguiente-numero", methods=["GET"])
def get_siguiente_numero_oferta():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar el siguiente número de oferta"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            siguiente_numero = get_next_numero_oferta(cursor)

        return jsonify({"success": True, "numero_oferta": siguiente_numero})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo consultar el siguiente número de oferta: {str(exc)}"}), 500


@app.route("/api/ofertas", methods=["GET"])
def list_ofertas():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar las ofertas"}), 401

    estado_id = request.args.get("estado_id", type=int)

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            ensure_offer_bom_links_schema(cursor)
            query = """
                SELECT
                    vw.id_oferta,
                    lo.id_estado,
                    e.id_departamento,
                    vw.numero_oferta,
                    vw.fecha_email,
                    vw.fecha_alta_oferta,
                    latest_interaction.fecha_limite,
                    oetc.fecha_envio_oferta,
                    vw.ref_cliente_asunto_email,
                    lo.id_cliente,
                    lo.id_bom,
                    vw.cliente,
                    b.material,
                    vw.emisor,
                    vw.observaciones_oferta,
                    COALESCE(NULLIF(LTRIM(RTRIM(gu.nombre)), ''), CAST(oetc.num_operario_responsable AS NVARCHAR(50))) AS nombre_responsable,
                    dd.nombre_departamento AS nombre_departamento_destino,
                    oetc.codigo_externo_oferta,
                    oetc.codigo_interno_oferta,
                    oetc.referencia_cliente,
                    oetc.numero_comision,
                    oetc.proyecto,
                    oetc.nombre_solicitante,
                    oetc.email_solicitante,
                    oetc.incoterm,
                    oetc.prioridad,
                    oetc.total_material_eur,
                    oetc.total_fee_eur,
                    oetc.observaciones_cliente,
                    oetc.pedido_b2b,
                    oetc.po_original,
                    oetc.sales_orders,
                    oetc.request_delivery_date,
                    STRING_AGG(CAST(ISNULL(vw.tipo_interaccion, '') AS NVARCHAR(MAX)), ' | ') AS tipos_interaccion,
                    STRING_AGG(CAST(CONVERT(VARCHAR(19), vw.fecha_interaccion, 120) AS NVARCHAR(MAX)), ' | ') AS fechas_interaccion,
                    STRING_AGG(CAST(ISNULL(vw.observaciones_interaccion, '') AS NVARCHAR(MAX)), ' | ') AS observaciones_interaccion,
                    vw.estado
                FROM ofertas.vw_listado_ofertas_interacciones vw
                INNER JOIN ofertas.listado_ofertas lo
                    ON lo.id_oferta = vw.id_oferta
                LEFT JOIN ofertas.estados e
                    ON e.id_estado = lo.id_estado
                LEFT JOIN ofertas.materiales_precio b
                    ON b.id_material_precio = lo.id_bom
                LEFT JOIN ofertas.oferta_etc oetc
                    ON oetc.id_oferta_etc = lo.id_oferta
                LEFT JOIN general.usuarios gu
                    ON gu.num_operario = oetc.num_operario_responsable
                LEFT JOIN ofertas.departamentos dd
                    ON dd.id_departamento = oetc.id_departamento_destino
                LEFT JOIN (
                    SELECT
                        oi.id_oferta,
                        oi.fecha_limite,
                        ROW_NUMBER() OVER (
                            PARTITION BY oi.id_oferta
                            ORDER BY
                                CASE WHEN oi.fecha_interaccion IS NULL THEN 1 ELSE 0 END,
                                oi.fecha_interaccion DESC,
                                oi.id_interaccion DESC
                        ) AS rn
                    FROM ofertas.oferta_interacciones oi
                ) latest_interaction
                    ON latest_interaction.id_oferta = lo.id_oferta
                   AND latest_interaction.rn = 1
            """
            params = []

            if estado_id is not None:
                query += " WHERE lo.id_estado = ?"
                params.append(estado_id)

            query += """
                GROUP BY
                    vw.id_oferta,
                    lo.id_estado,
                    e.id_departamento,
                    vw.numero_oferta,
                    vw.fecha_email,
                    vw.fecha_alta_oferta,
                    latest_interaction.fecha_limite,
                    oetc.fecha_envio_oferta,
                    vw.ref_cliente_asunto_email,
                    lo.id_cliente,
                    lo.id_bom,
                    vw.cliente,
                    b.material,
                    vw.emisor,
                    vw.observaciones_oferta,
                    COALESCE(NULLIF(LTRIM(RTRIM(gu.nombre)), ''), CAST(oetc.num_operario_responsable AS NVARCHAR(50))),
                    dd.nombre_departamento,
                    oetc.codigo_externo_oferta,
                    oetc.codigo_interno_oferta,
                    oetc.referencia_cliente,
                    oetc.numero_comision,
                    oetc.proyecto,
                    oetc.nombre_solicitante,
                    oetc.email_solicitante,
                    oetc.incoterm,
                    oetc.prioridad,
                    oetc.total_material_eur,
                    oetc.total_fee_eur,
                    oetc.observaciones_cliente,
                    oetc.pedido_b2b,
                    oetc.po_original,
                    oetc.sales_orders,
                    oetc.request_delivery_date,
                    vw.estado
                ORDER BY vw.id_oferta DESC
            """
            cursor.execute(query, tuple(params))

            ofertas = [
                {
                    "id_oferta": row[0],
                    "id_estado": row[1],
                    "id_departamento_estado": row[2],
                    "numero_oferta": row[3],
                    "fecha_email": serialize_date(row[4]),
                    "fecha_alta_oferta": serialize_date(row[5]),
                    "fecha_limite": serialize_date(row[6]),
                    "fecha_envio_oferta": serialize_date(row[7]),
                    "ref_cliente_asunto_email": row[8],
                    "id_cliente": row[9],
                    "id_bom": row[10],
                    "cliente": row[11],
                    "nombre_bom": row[12],
                    "emisor": row[13],
                    "observaciones": row[14],
                    "observaciones_oferta": row[14],
                    "nombre_responsable": row[15],
                    "nombre_departamento_destino": row[16],
                    "codigo_externo_oferta": row[17],
                    "codigo_interno_oferta": row[18],
                    "referencia_cliente": row[19],
                    "numero_comision": row[20],
                    "proyecto": row[21],
                    "nombre_solicitante": row[22],
                    "email_solicitante": row[23],
                    "incoterm": row[24],
                    "prioridad": row[25],
                    "total_material_eur": serialize_decimal(row[26]),
                    "total_fee_eur": serialize_decimal(row[27]),
                    "observaciones_cliente": row[28],
                    "pedido_b2b": row[29],
                    "po_original": row[30],
                    "sales_orders": row[31],
                    "request_delivery_date": serialize_date(row[32]),
                    "interaction_types": row[33],
                    "interaction_dates": serialize_interaction_date_list(row[34]),
                    "interaction_observaciones": row[35],
                    "estado": row[36],
                    "id_oferta": row[0],
                    "numero_oferta": row[3],
                    "fecha_email": serialize_date(row[4]),
                    "fecha_alta_oferta": serialize_date(row[5]),
                    "fecha_limite": serialize_date(row[6]),
                    "fecha_envio_oferta": serialize_date(row[7]),
                    "ref_cliente_asunto_email": row[8],
                    "id_bom": row[10],
                    "cliente": row[11],
                    "nombre_bom": row[12],
                    "emisor": row[13],
                    "observaciones_oferta": row[14],
                    "nombre_responsable": row[15],
                    "nombre_departamento_destino": row[16],
                    "codigo_externo_oferta": row[17],
                    "codigo_interno_oferta": row[18],
                    "referencia_cliente": row[19],
                    "numero_comision": row[20],
                    "proyecto": row[21],
                    "nombre_solicitante": row[22],
                    "email_solicitante": row[23],
                    "incoterm": row[24],
                    "prioridad": row[25],
                    "total_material_eur": serialize_decimal(row[26]),
                    "total_fee_eur": serialize_decimal(row[27]),
                    "observaciones_cliente": row[28],
                    "pedido_b2b": row[29],
                    "po_original": row[30],
                    "sales_orders": row[31],
                    "request_delivery_date": serialize_date(row[32]),
                    "tipo_interaccion": row[33],
                    "fecha_interaccion": serialize_interaction_date_list(row[34]),
                    "observaciones_interaccion": row[35],
                    "estado": row[36],
                }
                for row in cursor.fetchall()
            ]
            attach_offer_bom_materials(cursor, ofertas)

        for oferta in ofertas:
            chat_messages = load_offer_chat_messages(oferta["id_oferta"])
            oferta.update(build_offer_chat_summary(oferta["id_oferta"], user_data, messages=chat_messages))

        return jsonify({"success": True, "ofertas": ofertas})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar las ofertas: {str(exc)}"}), 500


@app.route("/api/ofertas-etc", methods=["GET"])
def list_ofertas_etc():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar las ofertas ETC"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()

            if not oferta_etc_table_exists(cursor):
                return jsonify({"success": False, "message": "La tabla ofertas.oferta_etc no existe todavía"}), 404

            cursor.execute(
                """
                SELECT
                    o.id_oferta_etc,
                    o.fecha_recepcion,
                    o.fecha_envio_oferta,
                    o.fecha_limite_respuesta,
                    o.id_estado,
                    e.descripcion_estado,
                    o.id_cliente,
                    c.descripcion_cliente,
                    o.num_operario_responsable,
                    gu.nombre,
                    o.id_departamento_destino,
                    d.nombre_departamento,
                    o.codigo_externo_oferta,
                    o.codigo_interno_oferta,
                    o.referencia_cliente,
                    o.numero_comision,
                    o.po_original,
                    o.pedido_b2b,
                    o.sales_orders,
                    o.request_delivery_date,
                    o.proyecto,
                    o.nombre_solicitante,
                    o.email_solicitante,
                    o.empresa_solicitante,
                    o.incoterm,
                    o.moneda,
                    o.prioridad,
                    o.es_urgente,
                    o.resumen_material_solicitado,
                    o.resumen_material_ofertado,
                    o.total_material_eur,
                    o.total_fee_eur,
                    o.total_oferta_eur,
                    o.observaciones_cliente,
                    o.observaciones_tecnicas,
                    o.observaciones_internas,
                    o.origen_registro,
                    o.activo,
                    o.fecha_creacion,
                    o.fecha_actualizacion
                FROM ofertas.oferta_etc o
                LEFT JOIN ofertas.estados e
                    ON e.id_estado = o.id_estado
                LEFT JOIN ofertas.clientes c
                    ON c.id_cliente = o.id_cliente
                LEFT JOIN ofertas.departamentos d
                    ON d.id_departamento = o.id_departamento_destino
                LEFT JOIN general.usuarios gu
                    ON gu.num_operario = o.num_operario_responsable
                ORDER BY o.fecha_creacion DESC, o.id_oferta_etc DESC
                """
            )

            ofertas = [
                {
                    "id_oferta_etc": row[0],
                    "fecha_recepcion": serialize_date(row[1]),
                    "fecha_envio_oferta": serialize_date(row[2]),
                    "fecha_limite_respuesta": serialize_date(row[3]),
                    "id_estado": row[4],
                    "estado": row[5],
                    "id_cliente": row[6],
                    "cliente": row[7],
                    "num_operario_responsable": row[8],
                    "nombre_responsable": row[9],
                    "id_departamento_destino": row[10],
                    "nombre_departamento": row[11],
                    "codigo_externo_oferta": row[12],
                    "codigo_interno_oferta": row[13],
                    "referencia_cliente": row[14],
                    "numero_comision": row[15],
                    "po_original": row[16],
                    "pedido_b2b": row[17],
                    "sales_orders": row[18],
                    "request_delivery_date": serialize_date(row[19]),
                    "proyecto": row[20],
                    "nombre_solicitante": row[21],
                    "email_solicitante": row[22],
                    "empresa_solicitante": row[23],
                    "incoterm": row[24],
                    "moneda": row[25],
                    "prioridad": row[26],
                    "es_urgente": bool(row[27]) if row[27] is not None else False,
                    "resumen_material_solicitado": row[28],
                    "resumen_material_ofertado": row[29],
                    "total_material_eur": serialize_decimal(row[30]),
                    "total_fee_eur": serialize_decimal(row[31]),
                    "total_oferta_eur": serialize_decimal(row[32]),
                    "observaciones_cliente": row[33],
                    "observaciones_tecnicas": row[34],
                    "observaciones_internas": row[35],
                    "origen_registro": row[36],
                    "activo": bool(row[37]) if row[37] is not None else True,
                    "fecha_creacion": serialize_date(row[36]),
                    "fecha_actualizacion": serialize_date(row[37]),
                }
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "ofertas_etc": ofertas})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar las ofertas ETC: {str(exc)}"}), 500


@app.route("/api/ofertas-etc/<int:oferta_etc_id>", methods=["GET"])
def get_oferta_etc(oferta_etc_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar la oferta ETC"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()

            if not oferta_etc_table_exists(cursor):
                return jsonify({"success": False, "message": "La tabla ofertas.oferta_etc no existe todavía"}), 404

            cursor.execute(
                """
                SELECT
                    o.id_oferta_etc,
                    o.fecha_recepcion,
                    o.fecha_envio_oferta,
                    o.fecha_limite_respuesta,
                    o.id_estado,
                    e.descripcion_estado,
                    o.id_cliente,
                    c.descripcion_cliente,
                    o.num_operario_responsable,
                    gu.nombre,
                    o.id_departamento_destino,
                    d.nombre_departamento,
                    o.codigo_externo_oferta,
                    o.codigo_interno_oferta,
                    o.referencia_cliente,
                    o.numero_comision,
                    o.po_original,
                    o.pedido_b2b,
                    o.sales_orders,
                    o.request_delivery_date,
                    o.proyecto,
                    o.nombre_solicitante,
                    o.email_solicitante,
                    o.empresa_solicitante,
                    o.incoterm,
                    o.moneda,
                    o.prioridad,
                    o.es_urgente,
                    o.resumen_material_solicitado,
                    o.resumen_material_ofertado,
                    o.total_material_eur,
                    o.total_fee_eur,
                    o.total_oferta_eur,
                    o.observaciones_cliente,
                    o.observaciones_tecnicas,
                    o.observaciones_internas,
                    o.origen_registro,
                    o.activo,
                    o.fecha_creacion,
                    o.fecha_actualizacion
                FROM ofertas.oferta_etc o
                LEFT JOIN ofertas.estados e
                    ON e.id_estado = o.id_estado
                LEFT JOIN ofertas.clientes c
                    ON c.id_cliente = o.id_cliente
                LEFT JOIN ofertas.departamentos d
                    ON d.id_departamento = o.id_departamento_destino
                LEFT JOIN general.usuarios gu
                    ON gu.num_operario = o.num_operario_responsable
                WHERE o.id_oferta_etc = ?
                """,
                (oferta_etc_id,),
            )
            row = cursor.fetchone()

            if not row:
                return jsonify({"success": False, "message": "Oferta ETC no encontrada"}), 404

            oferta = {
                "id_oferta_etc": row[0],
                "fecha_recepcion": serialize_date(row[1]),
                "fecha_envio_oferta": serialize_date(row[2]),
                "fecha_limite_respuesta": serialize_date(row[3]),
                "id_estado": row[4],
                "estado": row[5],
                "id_cliente": row[6],
                "cliente": row[7],
                "num_operario_responsable": row[8],
                "nombre_responsable": row[9],
                "id_departamento_destino": row[10],
                "nombre_departamento": row[11],
                "codigo_externo_oferta": row[12],
                "codigo_interno_oferta": row[13],
                "referencia_cliente": row[14],
                "numero_comision": row[15],
                "po_original": row[16],
                "pedido_b2b": row[17],
                "sales_orders": row[18],
                "request_delivery_date": serialize_date(row[19]),
                "proyecto": row[20],
                "nombre_solicitante": row[21],
                "email_solicitante": row[22],
                "empresa_solicitante": row[23],
                "incoterm": row[24],
                "moneda": row[25],
                "prioridad": row[26],
                "es_urgente": bool(row[27]) if row[27] is not None else False,
                "resumen_material_solicitado": row[28],
                "resumen_material_ofertado": row[29],
                "total_material_eur": serialize_decimal(row[30]),
                "total_fee_eur": serialize_decimal(row[31]),
                "total_oferta_eur": serialize_decimal(row[32]),
                "observaciones_cliente": row[33],
                "observaciones_tecnicas": row[34],
                "observaciones_internas": row[35],
                "origen_registro": row[36],
                "activo": bool(row[37]) if row[37] is not None else True,
                "fecha_creacion": serialize_date(row[38]),
                "fecha_actualizacion": serialize_date(row[39]),
            }

        return jsonify({"success": True, "oferta_etc": oferta})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo consultar la oferta ETC: {str(exc)}"}), 500


@app.route("/api/ofertas-etc/<int:oferta_etc_id>", methods=["PUT"])
def update_oferta_etc(oferta_etc_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para actualizar la oferta ETC"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_oferta_etc_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            actor_label = get_history_actor_label(user_data)
            payload["id_estado"] = resolve_active_estado_for_department(
                cursor,
                payload.get("id_departamento_destino"),
                fallback_estado_id=payload.get("id_estado"),
            )

            if not oferta_etc_table_exists(cursor):
                return jsonify({"success": False, "message": "La tabla ofertas.oferta_etc no existe todavía"}), 404

            cursor.execute("SELECT 1 FROM ofertas.oferta_etc WHERE id_oferta_etc = ?", (oferta_etc_id,))
            record_exists = cursor.fetchone() is not None
            current_snapshot = get_offer_etc_audit_snapshot(cursor, oferta_etc_id) if record_exists else None

            if not record_exists:
                cursor.execute("SELECT 1 FROM ofertas.listado_ofertas WHERE id_oferta = ?", (oferta_etc_id,))
                if cursor.fetchone() is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "Oferta ETC no encontrada"}), 404

            estado_metadata = get_estado_metadata(cursor, payload["id_estado"])
            if estado_metadata is None:
                conn.rollback()
                return jsonify({"success": False, "message": "Estado no encontrado"}), 404
            if not estado_metadata["activo"]:
                conn.rollback()
                return jsonify({"success": False, "message": "El estado seleccionado está inactivo y no se puede usar en procesos"}), 409

            if payload["id_cliente"] is not None:
                cursor.execute("SELECT descripcion_cliente FROM ofertas.clientes WHERE id_cliente = ?", (payload["id_cliente"],))
                client_row = cursor.fetchone()
                if client_row is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "Cliente no encontrado"}), 404
                client_label = client_row[0]
            else:
                client_label = None

            cursor.execute(
                """
                SELECT TOP 1 COALESCE(NULLIF(LTRIM(RTRIM(gu.nombre)), ''), CAST(uc.num_operario AS NVARCHAR(50)))
                FROM ofertas.usuarios_config uc
                LEFT JOIN general.usuarios gu
                    ON gu.num_operario = uc.num_operario
                WHERE uc.num_operario = ?
                """,
                (payload["num_operario_responsable"],),
            )
            responsable_row = cursor.fetchone()
            if responsable_row is None:
                conn.rollback()
                return jsonify({"success": False, "message": "Responsable no encontrado en ofertas.usuarios_config"}), 404
            responsable_label = responsable_row[0]

            cursor.execute(
                "SELECT nombre_departamento FROM ofertas.departamentos WHERE id_departamento = ?",
                (payload["id_departamento_destino"],),
            )
            department_row = cursor.fetchone()
            if department_row is None:
                conn.rollback()
                return jsonify({"success": False, "message": "Departamento no encontrado"}), 404
            department_label = department_row[0]

            updated_snapshot = {
                "fecha_recepcion": payload["fecha_recepcion"],
                "fecha_envio_oferta": payload["fecha_envio_oferta"],
                "fecha_limite_respuesta": payload["fecha_limite_respuesta"],
                "estado": estado_metadata["descripcion_estado"],
                "cliente": client_label,
                "responsable": responsable_label,
                "departamento_destino": department_label,
                "codigo_externo_oferta": payload["codigo_externo_oferta"],
                "codigo_interno_oferta": payload["codigo_interno_oferta"],
                "referencia_cliente": payload["referencia_cliente"],
                "numero_comision": payload["numero_comision"],
                "po_original": payload["po_original"],
                "pedido_b2b": payload["pedido_b2b"],
                "proyecto": payload["proyecto"],
                "sales_orders": payload["sales_orders"],
                "request_delivery_date": payload["request_delivery_date"],
                "nombre_solicitante": payload["nombre_solicitante"],
                "email_solicitante": payload["email_solicitante"],
                "empresa_solicitante": payload["empresa_solicitante"],
                "incoterm": payload["incoterm"],
                "moneda": payload["moneda"],
                "prioridad": payload["prioridad"],
                "es_urgente": bool(payload["es_urgente"]),
                "resumen_material_solicitado": payload["resumen_material_solicitado"],
                "resumen_material_ofertado": payload["resumen_material_ofertado"],
                "total_material_eur": payload["total_material_eur"],
                "total_fee_eur": payload["total_fee_eur"],
                "observaciones_cliente": payload["observaciones_cliente"],
                "observaciones_tecnicas": payload["observaciones_tecnicas"],
                "observaciones_internas": payload["observaciones_internas"],
                "origen_registro": payload["origen_registro"],
                "activo": bool(payload["activo"]),
            }

            if payload["proyecto"] is not None:
                cursor.execute(
                    "SELECT 1 FROM ofertas.proyectos WHERE LTRIM(RTRIM(descripcion_proyecto)) = LTRIM(RTRIM(?))",
                    (payload["proyecto"],),
                )
                if cursor.fetchone() is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "proyecto no encontrado en la configuración"}), 404

            if not record_exists:
                insert_oferta_etc_record(cursor, payload, explicit_id=oferta_etc_id)
            else:
                cursor.execute(
                    """
                    UPDATE ofertas.oferta_etc
                    SET fecha_recepcion = ?,
                        fecha_envio_oferta = ?,
                        fecha_limite_respuesta = ?,
                        id_estado = ?,
                        id_cliente = ?,
                        num_operario_responsable = ?,
                        id_departamento_destino = ?,
                        codigo_externo_oferta = ?,
                        codigo_interno_oferta = ?,
                        referencia_cliente = ?,
                        numero_comision = ?,
                        po_original = ?,
                        pedido_b2b = ?,
                        proyecto = ?,
                        sales_orders = ?,
                        request_delivery_date = ?,
                        nombre_solicitante = ?,
                        email_solicitante = ?,
                        empresa_solicitante = ?,
                        incoterm = ?,
                        moneda = ?,
                        prioridad = ?,
                        es_urgente = ?,
                        resumen_material_solicitado = ?,
                        resumen_material_ofertado = ?,
                        total_material_eur = ?,
                        total_fee_eur = ?,
                        observaciones_cliente = ?,
                        observaciones_tecnicas = ?,
                        observaciones_internas = ?,
                        origen_registro = ?,
                        activo = ?
                    WHERE id_oferta_etc = ?
                    """,
                    (
                        payload["fecha_recepcion"],
                        payload["fecha_envio_oferta"],
                        payload["fecha_limite_respuesta"],
                        payload["id_estado"],
                        payload["id_cliente"],
                        payload["num_operario_responsable"],
                        payload["id_departamento_destino"],
                        payload["codigo_externo_oferta"],
                        payload["codigo_interno_oferta"],
                        payload["referencia_cliente"],
                        payload["numero_comision"],
                        payload["po_original"],
                        payload["pedido_b2b"],
                        payload["proyecto"],
                        payload["sales_orders"],
                        payload["request_delivery_date"],
                        payload["nombre_solicitante"],
                        payload["email_solicitante"],
                        payload["empresa_solicitante"],
                        payload["incoterm"],
                        payload["moneda"],
                        payload["prioridad"],
                        1 if payload["es_urgente"] else 0,
                        payload["resumen_material_solicitado"],
                        payload["resumen_material_ofertado"],
                        payload["total_material_eur"],
                        payload["total_fee_eur"],
                        payload["observaciones_cliente"],
                        payload["observaciones_tecnicas"],
                        payload["observaciones_internas"],
                        payload["origen_registro"],
                        1 if payload["activo"] else 0,
                        oferta_etc_id,
                    ),
                )

                if cursor.rowcount == 0:
                    conn.rollback()
                    return jsonify({"success": False, "message": "No se pudo actualizar la oferta ETC"}), 409

                sync_offer_estado_with_etc(cursor, oferta_etc_id, payload["id_estado"])

            interaction_date = datetime.now()
            etc_change_entries = build_history_change_entries(
                current_snapshot,
                updated_snapshot,
                [
                    ("fecha_recepcion", "Fecha recepcion"),
                    ("fecha_envio_oferta", "Fecha envio oferta"),
                    ("fecha_limite_respuesta", "Fecha limite respuesta"),
                    ("estado", "Estado"),
                    ("cliente", "Cliente"),
                    ("responsable", "Responsable"),
                    ("departamento_destino", "Departamento destino"),
                    ("codigo_externo_oferta", "Codigo externo"),
                    ("codigo_interno_oferta", "Codigo interno"),
                    ("referencia_cliente", "Referencia cliente"),
                    ("numero_comision", "Numero comision"),
                    ("po_original", "PO original"),
                    ("pedido_b2b", "Pedido B2B"),
                    ("proyecto", "Proyecto"),
                    ("nombre_solicitante", "Nombre solicitante"),
                    ("email_solicitante", "Email solicitante"),
                    ("empresa_solicitante", "Empresa solicitante"),
                    ("incoterm", "Incoterm"),
                    ("moneda", "Moneda"),
                    ("prioridad", "Prioridad"),
                    ("es_urgente", "Urgente"),
                    ("resumen_material_solicitado", "Resumen material solicitado"),
                    ("resumen_material_ofertado", "Resumen material ofertado"),
                    ("total_material_eur", "Total material EUR"),
                    ("total_fee_eur", "Total fee EUR"),
                    ("observaciones_cliente", "Observaciones cliente"),
                    ("observaciones_tecnicas", "Observaciones tecnicas"),
                    ("observaciones_internas", "Observaciones internas"),
                    ("origen_registro", "Origen registro"),
                    ("activo", "Activo"),
                ],
                actor_label=actor_label,
            )
            for observation in etc_change_entries:
                insert_offer_interaction_entry(cursor, oferta_etc_id, "Edicion ETC", interaction_date, observaciones=observation)

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Oferta ETC actualizada correctamente",
                "id_oferta_etc": oferta_etc_id,
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except pyodbc.IntegrityError as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar la oferta ETC por una restricción de datos: {str(exc)}"}), 409
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar la oferta ETC: {str(exc)}"}), 500


@app.route("/api/ofertas/columnas-disponibles", methods=["GET"])
def list_available_offer_columns():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar las columnas disponibles"}), 401

    return jsonify({"success": True, "columnas": get_available_offer_columns()})


@app.route("/api/ofertas/importar-correo", methods=["POST"])
def import_oferta_from_email():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para importar correos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    uploaded_file = request.files.get("correo")
    if uploaded_file is None:
        return jsonify({"success": False, "message": "Debes adjuntar un correo .eml o .msg"}), 400

    try:
        parsed_email = parse_uploaded_email(uploaded_file)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo interpretar el correo: {str(exc)}"}), 500

    try:
        cliente_resolution = resolve_cliente_for_sender_email(parsed_email.get("sender_email"))
        matched_cliente = (cliente_resolution or {}).get("cliente")
        staged_imported_attachments = None

        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            thread_match = get_offer_thread_match(cursor, parsed_email)
            staging_bucket_name = None
            if thread_match is not None:
                staging_bucket_name = thread_match.get("numero_oferta") or thread_match.get("id_oferta")
            staged_imported_attachments = stage_imported_email_attachments(parsed_email, bucket_name=staging_bucket_name)
            if thread_match is not None:
                sync_result = sync_imported_emails_into_offer(
                    cursor,
                    thread_match["id_oferta"],
                    [parsed_email],
                    imported_email_attachment_token=staged_imported_attachments["token"] if staged_imported_attachments else None,
                )
                conn.commit()

                offer_reference = sync_result["numero_oferta"] or thread_match["id_oferta"]
                if sync_result["imported_count"] > 0:
                    message_text = f"Correo importado dentro de la oferta {offer_reference}."
                    message_key = "offer.import_existing_offer_success"
                else:
                    message_text = f"Ese correo ya estaba importado en la oferta {offer_reference}."
                    message_key = "offer.import_existing_offer_duplicate"

                return jsonify(
                    build_import_result_payload(
                        sync_result=sync_result,
                        message=message_text,
                        message_key=message_key,
                        message_params={"offer": str(offer_reference)},
                    )
                )

            conn.commit()

        message_parts = ["Correo importado correctamente."]
        if matched_cliente:
            message_parts.append(f"Cliente detectado: {matched_cliente['descripcion_cliente']}.")
        elif cliente_resolution and cliente_resolution.get("domain"):
            message_parts.append(f"No se encontró cliente para el dominio {cliente_resolution['domain']}.")
        if staged_imported_attachments:
            message_parts.append(
                f"Se han detectado {len(staged_imported_attachments['attachments'])} adjunto(s) y se añadirán al guardar la oferta."
            )

        response_data = build_imported_email_response_data(parsed_email, cliente_resolution)
        response_data["imported_email_attachment_token"] = staged_imported_attachments["token"] if staged_imported_attachments else None
        response_data["adjuntos_importados"] = staged_imported_attachments["attachments"] if staged_imported_attachments else []

        return jsonify(
            {
                "success": True,
                "message": " ".join(message_parts),
                "data": response_data,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo completar la importación del correo: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/importar-correo", methods=["POST"])
def import_email_into_offer(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para importar correos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    uploaded_file = request.files.get("correo")
    if uploaded_file is None:
        return jsonify({"success": False, "message": "Debes adjuntar un correo .eml o .msg"}), 400

    try:
        parsed_email = parse_uploaded_email(uploaded_file)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo interpretar el correo: {str(exc)}"}), 500

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_offer_exists(cursor, oferta_id)
            cursor.execute(
                "SELECT numero_oferta FROM ofertas.listado_ofertas WHERE id_oferta = ?",
                (oferta_id,),
            )
            offer_row = cursor.fetchone()
            if offer_row is None:
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            staged_imported_attachments = stage_imported_email_attachments(
                parsed_email,
                bucket_name=offer_row[0] or oferta_id,
            )
            sync_result = sync_imported_emails_into_offer(
                cursor,
                oferta_id,
                [parsed_email],
                imported_email_attachment_token=staged_imported_attachments["token"] if staged_imported_attachments else None,
            )
            conn.commit()

        offer_reference = sync_result["numero_oferta"] or oferta_id
        if sync_result["imported_count"] > 0:
            message_text = f"Correo importado dentro de la oferta {offer_reference}."
            message_key = "offer.import_existing_offer_success"
        else:
            message_text = f"Ese correo ya estaba importado en la oferta {offer_reference}."
            message_key = "offer.import_existing_offer_duplicate"

        return jsonify(
            build_import_result_payload(
                sync_result=sync_result,
                message=message_text,
                message_key=message_key,
                message_params={"offer": str(offer_reference)},
            )
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 404
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo importar el correo en la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>", methods=["GET"])
def get_oferta(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar la oferta"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            ensure_offer_bom_links_schema(cursor)
            cursor.execute(
                """
                SELECT
                    lo.id_oferta,
                    lo.numero_oferta,
                    lo.id_estado,
                    e.descripcion_estado,
                    lo.fecha_email,
                    lo.fecha_alta_oferta,
                    lo.ref_cliente_asunto_email,
                    lo.id_cliente,
                    lo.id_bom,
                    b.material,
                    lo.observaciones,
                    lo.nombre_emisor,
                    lo.email_emisor
                FROM ofertas.listado_ofertas lo
                LEFT JOIN ofertas.estados e
                    ON e.id_estado = lo.id_estado
                LEFT JOIN ofertas.materiales_precio b
                    ON b.id_material_precio = lo.id_bom
                WHERE lo.id_oferta = ?
                """,
                (oferta_id,),
            )
            row = cursor.fetchone()

            if not row:
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            oferta = {
                "id_oferta": row[0],
                "numero_oferta": row[1],
                "id_estado": row[2],
                "estado": row[3],
                "fecha_email": serialize_date(row[4]),
                "fecha_alta_oferta": serialize_date(row[5]),
                "ref_cliente_asunto_email": row[6],
                "id_cliente": row[7],
                "id_bom": row[8],
                "nombre_bom": row[9],
                "observaciones": row[10],
                "nombre_emisor": row[11],
                "email_emisor": row[12],
                "emisor": format_sender_display(row[11], row[12]),
                "interacciones": [],
                "adjuntos": list_offer_attachments(row[0], numero_oferta=row[1]),
            }
            attach_offer_bom_materials(cursor, [oferta])
            oferta.update(build_offer_chat_summary(row[0], user_data))

            cursor.execute(
                """
                SELECT
                    tipo_interaccion,
                    fecha_interaccion,
                    fecha_limite,
                    observaciones
                FROM ofertas.oferta_interacciones
                WHERE id_oferta = ?
                ORDER BY fecha_interaccion DESC, id_interaccion DESC
                """,
                (oferta_id,),
            )

            oferta["interacciones"] = [
                {
                    "tipo_interaccion": interaction_row[0],
                    "fecha_interaccion": interaction_row[1].isoformat() if interaction_row[1] else None,
                    "fecha_limite": interaction_row[2].isoformat() if interaction_row[2] else None,
                    "observaciones": interaction_row[3],
                }
                for interaction_row in cursor.fetchall()
            ]

        return jsonify({"success": True, "oferta": oferta})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo consultar la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/chat", methods=["GET"])
def get_offer_chat(oferta_id):
    if not INTERNAL_CHAT_ENABLED:
        return jsonify({"success": False, "message": "El chat interno está desactivado"}), 404

    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar el chat"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            if not ensure_offer_exists(cursor, oferta_id):
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

        messages = load_offer_chat_messages(oferta_id)
        return jsonify({"success": True, "messages": messages, **build_offer_chat_summary(oferta_id, user_data, messages=messages)})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo consultar el chat: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/chat", methods=["POST"])
def post_offer_chat_message(oferta_id):
    if not INTERNAL_CHAT_ENABLED:
        return jsonify({"success": False, "message": "El chat interno está desactivado"}), 404

    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para usar el chat"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            if not ensure_offer_exists(cursor, oferta_id):
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

        message_entry = append_offer_chat_message(oferta_id, user_data, data.get("message"))
        messages = load_offer_chat_messages(oferta_id)
        return jsonify(
            {
                "success": True,
                "message": "Mensaje enviado correctamente",
                "chat_message": message_entry,
                "messages": messages,
                **build_offer_chat_summary(oferta_id, user_data, messages=messages),
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo enviar el mensaje: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/chat/read", methods=["POST"])
def mark_offer_chat_read(oferta_id):
    if not INTERNAL_CHAT_ENABLED:
        return jsonify({"success": False, "message": "El chat interno está desactivado"}), 404

    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para actualizar el chat"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            if not ensure_offer_exists(cursor, oferta_id):
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

        messages = load_offer_chat_messages(oferta_id)
        return jsonify({"success": True, **mark_offer_chat_as_read(oferta_id, user_data, messages=messages)})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el estado del chat: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/adjuntos", methods=["POST"])
def upload_offer_attachments(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para adjuntar archivos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    uploaded_files = [file for file in request.files.getlist("archivos") if file and getattr(file, "filename", "")]
    if not uploaded_files:
        return jsonify({"success": False, "message": "Debes seleccionar al menos un archivo"}), 400

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            if not ensure_offer_exists(cursor, oferta_id):
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

        saved_attachments = [save_offer_attachment(oferta_id, uploaded_file) for uploaded_file in uploaded_files]
        return jsonify(
            {
                "success": True,
                "message": "Archivo subido correctamente." if len(saved_attachments) == 1 else "Archivos subidos correctamente.",
                "adjuntos": saved_attachments,
                "all_adjuntos": list_offer_attachments(oferta_id),
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        app.logger.exception("No se pudieron subir adjuntos para la oferta %s", oferta_id)
        return jsonify({"success": False, "message": f"No se pudieron subir los archivos: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/adjuntos/<path:filename>", methods=["GET"])
def download_offer_attachment(oferta_id, filename):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para descargar adjuntos"}), 401

    safe_filename = os.path.basename(str(filename or ""))
    if not safe_filename or safe_filename != filename:
        return jsonify({"success": False, "message": "Archivo no válido"}), 400

    attachments_dir = get_offer_attachments_dir(oferta_id)
    file_path = os.path.join(attachments_dir, safe_filename)
    if not os.path.isfile(file_path):
        return jsonify({"success": False, "message": "Adjunto no encontrado"}), 404

    metadata = load_offer_attachment_metadata(file_path)
    download_name = metadata.get("original_name") or safe_filename
    return send_from_directory(attachments_dir, safe_filename, as_attachment=True, download_name=download_name)


@app.route("/api/ofertas/<int:oferta_id>/adjuntos/<path:filename>/preview", methods=["GET"])
def preview_offer_attachment(oferta_id, filename):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para visualizar adjuntos"}), 401

    safe_filename = os.path.basename(str(filename or ""))
    if not safe_filename or safe_filename != filename:
        return jsonify({"success": False, "message": "Archivo no válido"}), 400

    attachments_dir = get_offer_attachments_dir(oferta_id)
    file_path = os.path.join(attachments_dir, safe_filename)
    if not os.path.isfile(file_path):
        return jsonify({"success": False, "message": "Adjunto no encontrado"}), 404

    metadata = load_offer_attachment_metadata(file_path)
    resolved_name = metadata.get("original_name") or safe_filename
    return send_from_directory(
        attachments_dir,
        safe_filename,
        as_attachment=False,
        download_name=resolved_name,
        mimetype=get_offer_attachment_mimetype(resolved_name),
    )


@app.route("/api/ofertas/<int:oferta_id>", methods=["PUT"])
def update_oferta(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para actualizar la oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_oferta_update_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_offer_bom_links_schema(cursor)
            actor_label = get_history_actor_label(user_data)
            current_snapshot = get_offer_audit_snapshot(cursor, oferta_id)
            if current_snapshot is None:
                conn.rollback()
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            cursor.execute("SELECT descripcion_estado FROM ofertas.estados WHERE id_estado = ?", (payload["id_estado"],))
            state_row = cursor.fetchone()
            if state_row is None:
                conn.rollback()
                return jsonify({"success": False, "message": "Estado no encontrado"}), 404
            state_label = state_row[0]

            if payload["id_cliente"] is not None:
                cursor.execute("SELECT descripcion_cliente FROM ofertas.clientes WHERE id_cliente = ?", (payload["id_cliente"],))
                client_row = cursor.fetchone()
                if client_row is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "Cliente no encontrado"}), 404
                client_label = client_row[0]
            else:
                client_label = None

            if payload["id_bom"] is not None:
                cursor.execute(
                    "SELECT material FROM ofertas.materiales_precio WHERE id_material_precio = ?",
                    (payload["id_bom"],),
                )
                bom_row = cursor.fetchone()
                if bom_row is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "BOM no encontrado"}), 404
                bom_label = bom_row[0]
            else:
                bom_label = None

            updated_snapshot = {
                "estado": state_label,
                "fecha_email": payload["fecha_email"],
                "fecha_alta_oferta": payload["fecha_alta_oferta"],
                "ref_cliente_asunto_email": payload["ref_cliente_asunto_email"],
                "cliente": client_label,
                "bom": bom_label,
                "observaciones": payload["observaciones"],
                "emisor": payload["emisor"],
            }

            if oferta_email_subject_exists(
                cursor,
                payload["fecha_email"],
                payload["ref_cliente_asunto_email"],
                excluded_id=oferta_id,
            ):
                conn.rollback()
                return jsonify({"success": False, "message": "Ya existe una oferta con la misma fecha de e-mail y el mismo asunto."}), 409

            cursor.execute(
                """
                UPDATE ofertas.listado_ofertas
                SET id_estado = ?,
                    fecha_email = ?,
                    fecha_alta_oferta = ?,
                    ref_cliente_asunto_email = ?,
                    id_cliente = ?,
                    id_bom = ?,
                    observaciones = ?,
                    nombre_emisor = ?,
                    email_emisor = ?
                WHERE id_oferta = ?
                """,
                (
                    payload["id_estado"],
                    payload["fecha_email"],
                    payload["fecha_alta_oferta"],
                    payload["ref_cliente_asunto_email"],
                    payload["id_cliente"],
                    payload["id_bom"],
                    payload["observaciones"],
                    payload["nombre_emisor"],
                    payload["email_emisor"],
                    oferta_id,
                ),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            if payload["id_bom"] is not None:
                add_offer_bom_material_link(cursor, oferta_id, payload["id_bom"])
            else:
                sync_offer_primary_bom(cursor, oferta_id)

            interaction_date = datetime.now()
            offer_change_entries = build_history_change_entries(
                current_snapshot,
                updated_snapshot,
                [
                    ("estado", "Estado"),
                    ("fecha_email", "Fecha email"),
                    ("fecha_alta_oferta", "Fecha alta oferta"),
                    ("ref_cliente_asunto_email", "Asunto"),
                    ("cliente", "Cliente"),
                    ("bom", "BOM"),
                    ("observaciones", "Observaciones"),
                    ("emisor", "Emisor"),
                ],
                actor_label=actor_label,
            )
            for observation in offer_change_entries:
                insert_offer_interaction_entry(cursor, oferta_id, "Edicion oferta", interaction_date, observaciones=observation)

            conn.commit()

        return jsonify({"success": True, "message": "Oferta actualizada correctamente"})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>", methods=["DELETE"])
def delete_oferta(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para eliminar la oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response("Solo los usuarios con rol Manager pueden eliminar ofertas.")

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_offer_bom_price_override_schema(cursor)

            cursor.execute(
                "SELECT numero_oferta FROM ofertas.listado_ofertas WHERE id_oferta = ?",
                (oferta_id,),
            )
            oferta_row = cursor.fetchone()
            if not oferta_row:
                conn.rollback()
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            numero_oferta = oferta_row[0] or oferta_id

            cursor.execute(
                "DELETE FROM ofertas.oferta_interacciones WHERE id_oferta = ?",
                (oferta_id,),
            )
            cursor.execute(
                "DELETE FROM ofertas.oferta_correos_importados WHERE id_oferta = ?",
                (oferta_id,),
            )
            cursor.execute(
                "DELETE FROM ofertas.oferta_bom_precio_override WHERE id_oferta = ?",
                (oferta_id,),
            )
            cursor.execute(
                "DELETE FROM ofertas.oferta_bom_materiales WHERE id_oferta = ?",
                (oferta_id,),
            )
            cursor.execute(
                "DELETE FROM ofertas.listado_ofertas WHERE id_oferta = ?",
                (oferta_id,),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            conn.commit()

        delete_offer_attachment_storage(oferta_id, numero_oferta=numero_oferta)

        return jsonify(
            {
                "success": True,
                "message": f"Oferta {numero_oferta} eliminada correctamente",
                "id_oferta": oferta_id,
                "numero_oferta": numero_oferta,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo eliminar la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/verificar-duplicado-correo", methods=["POST"])
def check_oferta_duplicate_email():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para comprobar el correo"}), 401

    data = request.get_json(silent=True) or {}

    try:
        fecha_email = normalize_required_date(data.get("fecha_email"), "fecha_email")
        ref_cliente_asunto_email = normalize_optional_text(data.get("ref_cliente_asunto_email"), 500)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection() as conn:
            cursor = conn.cursor()
            exists = oferta_email_subject_exists(cursor, fecha_email, ref_cliente_asunto_email)

        return jsonify(
            {
                "success": True,
                "exists": exists,
                "message": "Ya existe una oferta con la misma fecha de e-mail y el mismo asunto." if exists else "",
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo comprobar el correo: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/estado", methods=["POST"])
def update_oferta_estado(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para actualizar el estado de la oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_oferta_estado_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    notification_result = None
    sender_department_name = None
    sender_departments = user_data.get("departamentos") if isinstance(user_data, dict) else None
    if isinstance(sender_departments, list) and sender_departments:
        sender_department_name = normalize_optional_text(sender_departments[0].get("nombre_departamento"), 255)

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            cursor.execute(
                """
                SELECT
                    lo.id_oferta,
                    lo.id_estado,
                    e.descripcion_estado
                FROM ofertas.listado_ofertas lo
                LEFT JOIN ofertas.estados e
                    ON e.id_estado = lo.id_estado
                WHERE lo.id_oferta = ?
                """,
                (oferta_id,),
            )
            oferta_row = cursor.fetchone()

            if not oferta_row:
                conn.rollback()
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            previous_estado_id = oferta_row[1]
            previous_estado = (oferta_row[2] or "Sin estado").strip()
            previous_estado_metadata = get_estado_metadata(cursor, previous_estado_id) if previous_estado_id is not None else None
            previous_department_name = normalize_optional_text(
                (previous_estado_metadata or {}).get("nombre_departamento"),
                255,
            )

            next_estado_metadata = get_estado_metadata(cursor, payload["id_estado"])
            if next_estado_metadata is None:
                conn.rollback()
                return jsonify({"success": False, "message": "Estado siguiente no encontrado"}), 404
            if not next_estado_metadata["activo"]:
                conn.rollback()
                return jsonify({"success": False, "message": "El estado siguiente está inactivo y no se puede seleccionar"}), 409

            if previous_estado_id == payload["id_estado"]:
                conn.rollback()
                return jsonify({"success": False, "message": "Debes seleccionar un estado distinto al actual"}), 400

            next_estado = (next_estado_metadata["descripcion_estado"] or "Sin estado").strip()
            next_department_name = normalize_optional_text(next_estado_metadata.get("nombre_departamento"), 255)
            interaction_date = datetime.now()
            interaction_type = f"{previous_estado} -> {next_estado}"
            actor_label = get_history_actor_label(user_data)
            interaction_comment = normalize_optional_text(payload["comentario"])
            if actor_label:
                interaction_comment = f"{interaction_comment}\nActualizado por: {actor_label}" if interaction_comment else f"Actualizado por: {actor_label}"

            cursor.execute(
                """
                UPDATE ofertas.listado_ofertas
                SET id_estado = ?
                WHERE id_oferta = ?
                """,
                (payload["id_estado"], oferta_id),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            cursor.execute(
                """
                INSERT INTO ofertas.oferta_interacciones (
                    id_oferta,
                    tipo_interaccion,
                    fecha_interaccion,
                    fecha_limite,
                    observaciones
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (oferta_id, interaction_type, interaction_date, payload["fecha_limite"], interaction_comment),
            )

            notification_payload = build_estado_manager_notification(
                cursor,
                oferta_id,
                payload["id_estado"],
                sender_department_name=sender_department_name,
                source_department_name=previous_department_name,
                target_department_name=next_department_name,
            )

            conn.commit()

            notification_result = send_estado_manager_notification(notification_payload)

        return jsonify(
            {
                "success": True,
                "message": "Estado actualizado correctamente",
                "fecha_interaccion": interaction_date.isoformat(),
                "fecha_limite": payload["fecha_limite"].isoformat() if payload["fecha_limite"] else None,
                "tipo_interaccion": interaction_type,
                "estado_anterior": previous_estado,
                "estado_siguiente": next_estado,
                "notification": notification_result,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el estado de la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/reasignar", methods=["POST"])
def reassign_oferta_responsable(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para reasignar ofertas"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response("Solo los managers pueden reasignar tareas.")

    data = request.get_json(silent=True) or {}
    try:
        responsable_num_operario = int(data.get("num_operario_responsable"))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Debes seleccionar un usuario valido"}), 400

    manager_department_ids = get_user_department_ids(user_data)
    if not manager_department_ids:
        return jsonify({"success": False, "message": "Tu usuario no tiene ningun departamento asignado"}), 409

    assigning_user_name = normalize_optional_text(
        user_data.get("nombre") or user_data.get("display_name") or user_data.get("email"),
        255,
    )

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if not oferta_etc_table_exists(cursor):
                conn.rollback()
                return jsonify({"success": False, "message": "La tabla de ofertas ETC no está disponible"}), 409

            cursor.execute(
                """
                SELECT
                    lo.id_oferta,
                    lo.numero_oferta,
                    lo.id_estado,
                    lo.id_cliente,
                    e.id_departamento,
                    d.nombre_departamento,
                    oetc.id_oferta_etc,
                    oetc.num_operario_responsable
                FROM ofertas.listado_ofertas lo
                LEFT JOIN ofertas.estados e
                    ON e.id_estado = lo.id_estado
                LEFT JOIN ofertas.departamentos d
                    ON d.id_departamento = e.id_departamento
                LEFT JOIN ofertas.oferta_etc oetc
                    ON oetc.id_oferta_etc = lo.id_oferta
                WHERE lo.id_oferta = ?
                """,
                (oferta_id,),
            )
            offer_row = cursor.fetchone()

            if not offer_row:
                conn.rollback()
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            offer_context = {
                "id_oferta": offer_row[0],
                "numero_oferta": offer_row[1],
                "id_estado": offer_row[2],
                "id_cliente": offer_row[3],
            }
            estado_department_id = offer_row[4]
            department_name = normalize_optional_text(offer_row[5], 255)
            etc_id = offer_row[6]
            previous_responsable = offer_row[7]

            if estado_department_id is None:
                conn.rollback()
                return jsonify({"success": False, "message": "La oferta no tiene un departamento asociado por estado"}), 409

            if int(estado_department_id) not in manager_department_ids:
                conn.rollback()
                return jsonify({"success": False, "message": "Solo puedes reasignar ofertas de tu departamento"}), 403

            if etc_id is None:
                etc_id = ensure_oferta_etc_record(
                    cursor,
                    offer_context,
                    estado_department_id,
                    responsable_num_operario=previous_responsable,
                )

            cursor.execute(
                """
                SELECT
                    uc.num_operario,
                    uc.id_departamento,
                    uc.email,
                    gu.nombre
                FROM ofertas.usuarios_config uc
                LEFT JOIN general.usuarios gu
                    ON gu.num_operario = uc.num_operario
                WHERE uc.num_operario = ?
                """,
                (responsable_num_operario,),
            )
            assignee_row = cursor.fetchone()

            if not assignee_row:
                conn.rollback()
                return jsonify({"success": False, "message": "El usuario seleccionado no existe"}), 404

            assignee_department_id = assignee_row[1]
            if assignee_department_id is None or int(assignee_department_id) != int(estado_department_id):
                conn.rollback()
                return jsonify({"success": False, "message": "Solo puedes reasignar a usuarios del mismo departamento"}), 409

            if previous_responsable is not None and int(previous_responsable) == responsable_num_operario:
                conn.rollback()
                return jsonify({"success": False, "message": "La oferta ya está asignada a ese usuario"}), 409

            cursor.execute(
                """
                UPDATE ofertas.oferta_etc
                SET num_operario_responsable = ?,
                    id_departamento_destino = ?
                WHERE id_oferta_etc = ?
                """,
                (responsable_num_operario, estado_department_id, etc_id),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "No se pudo actualizar la reasignacion"}), 409

            interaction_date = datetime.now()
            assignee_name = normalize_optional_text(assignee_row[3], 255) or str(responsable_num_operario)
            interaction_comment = f"Responsable reasignado a {assignee_name}"
            if assigning_user_name:
                interaction_comment += f" por {assigning_user_name}"

            cursor.execute(
                """
                INSERT INTO ofertas.oferta_interacciones (
                    id_oferta,
                    tipo_interaccion,
                    fecha_interaccion,
                    observaciones
                )
                VALUES (?, ?, ?, ?)
                """,
                (oferta_id, "Reasignacion", interaction_date, interaction_comment),
            )

            notification_payload = build_reassignment_notification(
                cursor,
                oferta_id,
                assigned_by_name=assigning_user_name,
            )

            conn.commit()

        notification_result = send_reassignment_notification(notification_payload)
        offer_label = normalize_optional_text(offer_row[1], 100) or f"ID {oferta_id}"
        response_message = f"Oferta {offer_label} reasignada correctamente"
        if department_name:
            response_message += f" en {department_name}"
        if notification_result and notification_result.get("skipped") and notification_result.get("message"):
            response_message += f". {notification_result['message']}"

        return jsonify(
            {
                "success": True,
                "message": response_message,
                "oferta_id": oferta_id,
                "num_operario_responsable": responsable_num_operario,
                "responsable_nombre": assignee_name,
                "id_departamento_estado": estado_department_id,
                "fecha_interaccion": interaction_date.isoformat(),
                "notification": notification_result,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo reasignar la oferta: {str(exc)}"}), 500


@app.route("/api/clientes", methods=["GET"])
def list_clientes():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los clientes"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT id_cliente, descripcion_cliente, dominio
                FROM ofertas.clientes
                ORDER BY descripcion_cliente ASC
                """
            )
            clientes = [
                {"id_cliente": row[0], "descripcion_cliente": row[1], "dominio": row[2]}
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "clientes": clientes})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los clientes: {str(exc)}"}), 500


@app.route("/api/clientes", methods=["POST"])
def create_cliente():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para crear clientes"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_cliente_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if cliente_exists(cursor, payload["descripcion_cliente"]):
                return jsonify({"success": False, "message": "Ya existe un cliente con esa descripción"}), 409
            if cliente_domain_exists(cursor, payload["dominio"]):
                return jsonify({"success": False, "message": "Ya existe un cliente con ese dominio"}), 409

            cursor.execute(
                """
                INSERT INTO ofertas.clientes (descripcion_cliente, dominio)
                OUTPUT INSERTED.id_cliente
                VALUES (?, ?)
                """,
                (payload["descripcion_cliente"], payload["dominio"]),
            )
            inserted = cursor.fetchone()
            inserted_id = inserted[0] if inserted else None
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Cliente creado correctamente",
                "cliente": {
                    "id_cliente": inserted_id,
                    "descripcion_cliente": payload["descripcion_cliente"],
                    "dominio": payload["dominio"],
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo crear el cliente: {str(exc)}"}), 500


@app.route("/api/clientes/<int:cliente_id>", methods=["PUT"])
def update_cliente(cliente_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para editar clientes"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_cliente_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if cliente_exists(cursor, payload["descripcion_cliente"], excluded_id=cliente_id):
                return jsonify({"success": False, "message": "Ya existe un cliente con esa descripción"}), 409
            if cliente_domain_exists(cursor, payload["dominio"], excluded_id=cliente_id):
                return jsonify({"success": False, "message": "Ya existe un cliente con ese dominio"}), 409

            cursor.execute(
                """
                UPDATE ofertas.clientes
                SET descripcion_cliente = ?, dominio = ?
                WHERE id_cliente = ?
                """,
                (payload["descripcion_cliente"], payload["dominio"], cliente_id),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Cliente no encontrado"}), 404

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Cliente actualizado correctamente",
                "cliente": {
                    "id_cliente": cliente_id,
                    "descripcion_cliente": payload["descripcion_cliente"],
                    "dominio": payload["dominio"],
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el cliente: {str(exc)}"}), 500


@app.route("/api/proyectos", methods=["GET"])
def list_proyectos():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los proyectos"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT id_proyecto, descripcion_proyecto
                FROM ofertas.proyectos
                ORDER BY descripcion_proyecto ASC
                """
            )
            proyectos = [
                {"id_proyecto": row[0], "descripcion_proyecto": row[1]}
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "proyectos": proyectos})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los proyectos: {str(exc)}"}), 500


@app.route("/api/proyectos", methods=["POST"])
def create_proyecto():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para crear proyectos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_proyecto_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if proyecto_exists(cursor, payload["descripcion_proyecto"]):
                return jsonify({"success": False, "message": "Ya existe un proyecto con esa descripción"}), 409

            cursor.execute(
                """
                INSERT INTO ofertas.proyectos (descripcion_proyecto)
                OUTPUT INSERTED.id_proyecto
                VALUES (?)
                """,
                (payload["descripcion_proyecto"],),
            )
            inserted = cursor.fetchone()
            inserted_id = inserted[0] if inserted else None
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "proyecto creado correctamente",
                "proyecto": {
                    "id_proyecto": inserted_id,
                    "descripcion_proyecto": payload["descripcion_proyecto"],
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo crear el proyecto: {str(exc)}"}), 500


@app.route("/api/proyectos/<int:proyecto_id>", methods=["PUT"])
def update_proyecto(proyecto_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para editar proyectos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_proyecto_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if proyecto_exists(cursor, payload["descripcion_proyecto"], excluded_id=proyecto_id):
                return jsonify({"success": False, "message": "Ya existe un proyecto con esa descripción"}), 409

            cursor.execute(
                """
                UPDATE ofertas.proyectos
                SET descripcion_proyecto = ?
                WHERE id_proyecto = ?
                """,
                (payload["descripcion_proyecto"], proyecto_id),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "proyecto no encontrado"}), 404

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "proyecto actualizado correctamente",
                "proyecto": {
                    "id_proyecto": proyecto_id,
                    "descripcion_proyecto": payload["descripcion_proyecto"],
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el proyecto: {str(exc)}"}), 500


@app.route("/api/materiales-precio", methods=["GET"])
def list_materiales_precio():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los materiales BOM"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            ensure_bom_catalog_schema(cursor)
            materiales = list_materiales_precio_snapshot(cursor)

        return jsonify({"success": True, "materiales": materiales})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los materiales BOM: {str(exc)}"}), 500


@app.route("/api/materiales-precio", methods=["POST"])
def create_material_precio():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para guardar precios BOM"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_material_precio_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_bom_catalog_schema(cursor)
            inserted_bom_id = upsert_bom_catalog_row(cursor, payload, bom_id=None)
            cursor.execute(
                """
                SELECT
                    mp.id_material_precio,
                    mp.material,
                    mp.precio,
                    mp.fecha_creacion,
                    mp.part_nr,
                    mp.mat_description,
                    mp.new_sales_price,
                    mp.notas,
                    mp.nuevos_cp357,
                    mp.nuevos_cp361,
                    mp.nuevos_cp365,
                    mp.anulados_cp365,
                    mp.nuevos_cp369,
                    mp.anulados_cp369,
                    mp.nuevos_cp373,
                    mp.anulados_cp373
                FROM ofertas.materiales_precio mp
                WHERE mp.id_material_precio = ?
                """,
                (inserted_bom_id,),
            )
            inserted = cursor.fetchone()
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Precio BOM guardado correctamente. Se ha creado un nuevo registro para conservar el histórico.",
                "material": build_bom_catalog_record(inserted) if inserted else None,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo guardar el precio BOM: {str(exc)}"}), 500


@app.route("/api/boms", methods=["GET"])
def list_boms():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los BOM"}), 401

    only_active = request.args.get("only_active", default=0, type=int) == 1

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            ensure_bom_catalog_schema(cursor)
            boms = list_boms_catalog(cursor, only_active=only_active)

        return jsonify({"success": True, "boms": boms})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los BOM: {str(exc)}"}), 500


@app.route("/api/boms", methods=["POST"])
def create_bom():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para crear BOM"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_bom_mutation_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_bom_catalog_schema(cursor)

            if find_conflicting_bom(cursor, payload) is not None:
                conn.rollback()
                return jsonify({"success": False, "message": "Ya existe un BOM con ese Part Nr o material"}), 409

            inserted_bom_id = upsert_bom_catalog_row(cursor, payload, bom_id=None)
            cursor.execute(
                """
                SELECT
                    mp.id_material_precio,
                    mp.material,
                    mp.precio,
                    mp.fecha_creacion,
                    mp.part_nr,
                    mp.mat_description,
                    mp.new_sales_price,
                    mp.notas,
                    mp.nuevos_cp357,
                    mp.nuevos_cp361,
                    mp.nuevos_cp365,
                    mp.anulados_cp365,
                    mp.nuevos_cp369,
                    mp.anulados_cp369,
                    mp.nuevos_cp373,
                    mp.anulados_cp373
                FROM ofertas.materiales_precio mp
                WHERE mp.id_material_precio = ?
                """,
                (inserted_bom_id,),
            )
            inserted = cursor.fetchone()
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "BOM creado correctamente",
                "bom": build_bom_catalog_record(inserted) if inserted else None,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo crear el BOM: {str(exc)}"}), 500


@app.route("/api/boms/<int:bom_id>", methods=["PUT"])
def update_bom(bom_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para editar BOM"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_bom_mutation_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_bom_catalog_schema(cursor)

            if find_conflicting_bom(cursor, payload, exclude_bom_id=bom_id) is not None:
                conn.rollback()
                return jsonify({"success": False, "message": "Ya existe otro BOM con ese Part Nr o material"}), 409

            upsert_bom_catalog_row(cursor, payload, bom_id=bom_id)
            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "BOM no encontrado"}), 404

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "BOM actualizado correctamente",
                "bom": {
                    "id_bom": bom_id,
                    "part_nr": payload["part_nr"],
                    "mat_description": payload["mat_description"],
                    "new_sales_price": serialize_decimal(payload["new_sales_price"]),
                    "notas": payload.get("notas"),
                    "material": payload["material"],
                    "precio": serialize_decimal(payload["precio"]),
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el BOM: {str(exc)}"}), 500


@app.route("/api/boms/<int:bom_id>", methods=["DELETE"])
def delete_bom(bom_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para eliminar BOM"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            result = delete_bom_catalog_material(cursor, bom_id)
            conn.commit()

        affected_offer_count = result.get("affected_offer_count", 0)
        deleted_override_count = result.get("deleted_override_count", 0)
        detail_parts = []
        if affected_offer_count:
            detail_parts.append(f"{affected_offer_count} ofertas actualizadas")
        if deleted_override_count:
            detail_parts.append(f"{deleted_override_count} excepciones eliminadas")

        detail_suffix = f" ({', '.join(detail_parts)})" if detail_parts else ""

        return jsonify(
            {
                "success": True,
                "message": f"BOM eliminado correctamente{detail_suffix}.",
                "deleted_bom": {
                    "id_bom": bom_id,
                    "material": result.get("material"),
                },
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 404
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo eliminar el BOM: {str(exc)}"}), 500


@app.route("/api/boms/export", methods=["GET"])
def export_boms_csv():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para exportar los BOM"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            ensure_bom_catalog_schema(cursor)
            boms = list_boms_catalog(cursor)

        csv_content = build_bom_catalog_csv_content(boms)
        filename = f"bom_catalog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return Response(
            csv_content.encode("utf-8-sig"),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron exportar los BOM: {str(exc)}"}), 500


@app.route("/api/boms/import", methods=["POST"])
def import_boms_csv():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para importar BOM"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    uploaded_file = request.files.get("archivo")
    if uploaded_file is None or not getattr(uploaded_file, "filename", ""):
        return jsonify({"success": False, "message": "Debes seleccionar un archivo CSV o Excel"}), 400

    _, extension = os.path.splitext(uploaded_file.filename or "")
    normalized_extension = extension.lower()
    if normalized_extension not in {".csv", ".xlsx", ".xlsm"}:
        return jsonify({"success": False, "message": "Solo se admite importación CSV o Excel (.xlsx/.xlsm) en este menú BOM"}), 400

    try:
        if normalized_extension == ".csv":
            imported_rows = parse_bom_catalog_csv(uploaded_file)
        else:
            imported_rows = parse_bom_catalog_excel(uploaded_file)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            result = replace_bom_catalog_from_rows(cursor, imported_rows)
            conn.commit()

        summary = (
            f"Importación BOM completada: {result['updated_count']} actualizados, "
            f"{result['inserted_count']} insertados y {result['deleted_count']} eliminados."
        )

        # Persistir última importación en archivo JSON (visible para todos los usuarios)
        try:
            last_import_path = os.path.join(RUNTIME_DATA_DIR, "last_bom_import.json")
            os.makedirs(RUNTIME_DATA_DIR, exist_ok=True)
            with open(last_import_path, "w", encoding="utf-8") as f:
                json.dump({
                    "file_name": uploaded_file.filename,
                    "imported_at": datetime.now().isoformat(),
                    "updated_count": result["updated_count"],
                    "inserted_count": result["inserted_count"],
                    "deleted_count": result["deleted_count"],
                }, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        return jsonify(
            {
                "success": True,
                "message": "Importación BOM completada correctamente",
                "summary": summary,
                "result": result,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": "No se pudo importar el catálogo BOM"}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": "No se pudo importar el catálogo BOM"}), 500


@app.route("/api/boms/last-import", methods=["GET"])
def get_last_bom_import():
    """Devuelve la información de la última importación del catálogo BOM."""
    last_import_path = os.path.join(RUNTIME_DATA_DIR, "last_bom_import.json")
    try:
        if os.path.isfile(last_import_path):
            with open(last_import_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return jsonify({"success": True, **data})
        return jsonify({"success": True, "file_name": None, "imported_at": None})
    except Exception:
        return jsonify({"success": True, "file_name": None, "imported_at": None})


@app.route("/api/estados", methods=["GET"])
def list_estados():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los estados"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    e.id_estado,
                    e.descripcion_estado,
                    e.orden,
                    ISNULL(o.total_ofertas, 0) AS total_ofertas,
                    e.id_departamento,
                    ISNULL(d.nombre_departamento, '') AS nombre_departamento,
                    ISNULL(e.emoji_sidebar, '') AS emoji_sidebar,
                    ISNULL(e.activo, 1) AS activo
                FROM ofertas.estados e
                LEFT JOIN (
                    SELECT id_estado, COUNT(*) AS total_ofertas
                    FROM ofertas.listado_ofertas
                    GROUP BY id_estado
                ) o
                    ON o.id_estado = e.id_estado
                LEFT JOIN ofertas.departamentos d
                    ON d.id_departamento = e.id_departamento
                WHERE e.id_estado > 0
                ORDER BY
                    CASE WHEN e.orden IS NULL THEN 1 ELSE 0 END,
                    e.orden ASC,
                    e.descripcion_estado ASC
                """
            )
            estados = [
                {
                    "id_estado": row[0],
                    "descripcion_estado": row[1],
                    "orden": row[2],
                    "total_ofertas": row[3],
                    "id_departamento": row[4],
                    "nombre_departamento": row[5],
                    "emoji_sidebar": row[6] or "",
                    "activo": bool(row[7]),
                }
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "estados": estados})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los estados: {str(exc)}"}), 500


@app.route("/api/estados", methods=["POST"])
def create_estado():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para crear estados"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_estado_payload(data)
        id_departamento = normalize_optional_int(data.get("id_departamento"), "id_departamento")
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if estado_exists(cursor, payload["descripcion_estado"]):
                return jsonify({"success": False, "message": "Ya existe un estado con esa descripción"}), 409

            if id_departamento is not None:
                cursor.execute("SELECT 1 FROM ofertas.departamentos WHERE id_departamento = ?", (id_departamento,))
                if cursor.fetchone() is None:
                    return jsonify({"success": False, "message": "Departamento no encontrado"}), 404

            if payload["orden"] is None:
                cursor.execute("SELECT ISNULL(MAX(orden), 0) + 1 FROM ofertas.estados")
                next_order_row = cursor.fetchone()
                payload["orden"] = next_order_row[0] if next_order_row and next_order_row[0] is not None else 1

            cursor.execute(
                """
                INSERT INTO ofertas.estados (descripcion_estado, orden, id_departamento, emoji_sidebar, activo)
                OUTPUT INSERTED.id_estado
                VALUES (?, ?, ?, ?, ?)
                """,
                (payload["descripcion_estado"], payload["orden"], id_departamento, payload["emoji_sidebar"], 1 if payload["activo"] else 0),
            )
            inserted = cursor.fetchone()
            inserted_id = inserted[0] if inserted else None
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Estado creado correctamente",
                "estado": {
                    "id_estado": inserted_id,
                    "descripcion_estado": payload["descripcion_estado"],
                    "orden": payload["orden"],
                    "id_departamento": id_departamento,
                    "emoji_sidebar": payload["emoji_sidebar"],
                    "activo": payload["activo"],
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo crear el estado: {str(exc)}"}), 500


@app.route("/api/estados/<int:estado_id>", methods=["PUT"])
def update_estado(estado_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para editar estados"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_estado_payload(data)
        id_departamento = normalize_optional_int(data.get("id_departamento"), "id_departamento")
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if estado_exists(cursor, payload["descripcion_estado"], excluded_id=estado_id):
                return jsonify({"success": False, "message": "Ya existe un estado con esa descripción"}), 409

            if id_departamento is not None:
                cursor.execute("SELECT 1 FROM ofertas.departamentos WHERE id_departamento = ?", (id_departamento,))
                if cursor.fetchone() is None:
                    return jsonify({"success": False, "message": "Departamento no encontrado"}), 404

            cursor.execute(
                """
                UPDATE ofertas.estados
                SET descripcion_estado = ?, orden = ?, id_departamento = ?, emoji_sidebar = ?, activo = ?
                WHERE id_estado = ?
                """,
                (payload["descripcion_estado"], payload["orden"], id_departamento, payload["emoji_sidebar"], 1 if payload["activo"] else 0, estado_id),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Estado no encontrado"}), 404

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Estado actualizado correctamente",
                "estado": {
                    "id_estado": estado_id,
                    "descripcion_estado": payload["descripcion_estado"],
                    "orden": payload["orden"],
                    "emoji_sidebar": payload["emoji_sidebar"],
                    "activo": payload["activo"],
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el estado: {str(exc)}"}), 500


@app.route("/api/estados/reorder", methods=["POST"])
def reorder_estados():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para reordenar estados"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    items = (request.get_json(silent=True) or {}).get("orden", [])
    if not isinstance(items, list) or not items:
        return jsonify({"success": False, "message": "Payload inválido: se espera {orden: [{id_estado, orden}]}"}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            for item in items:
                cursor.execute(
                    "UPDATE ofertas.estados SET orden = ? WHERE id_estado = ?",
                    (int(item["orden"]), int(item["id_estado"])),
                )
            conn.commit()
        return jsonify({"success": True, "message": "orden actualizado"})
    except (KeyError, TypeError, ValueError) as exc:
        return jsonify({"success": False, "message": f"Datos inválidos: {str(exc)}"}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo reordenar los estados: {str(exc)}"}), 500


@app.route("/api/estados/<estado_key>/columnas", methods=["GET"])
def list_configuracion_columnas(estado_key):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar la configuración de columnas"}), 401

    try:
        estado_id = parse_config_scope_id(estado_key)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            if not ensure_config_scope_exists(cursor, estado_id):
                return jsonify({"success": False, "message": "Estado no encontrado"}), 404
            if is_global_config_scope(estado_id):
                ensure_global_config_default_columns(cursor)

            cursor.execute(
                """
                SELECT id_config, id_estado, columna, descripcion_columna, orden_columna
                FROM ofertas.configuracioncolumnas
                WHERE id_estado = ?
                ORDER BY ISNULL(orden_columna, 2147483647), id_config ASC
                """,
                (estado_id,),
            )
            configuraciones = [
                {
                    "id_config": row[0],
                    "id_estado": row[1],
                    "columna": normalize_offer_column_name(row[2]),
                    "descripcion_columna": row[3],
                    "orden_columna": row[4],
                }
                for row in cursor.fetchall()
                if normalize_offer_column_name(row[2]) in get_available_offer_column_map()
            ]

        return jsonify({"success": True, "configuraciones": configuraciones})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo consultar la configuración de columnas: {str(exc)}"}), 500


@app.route("/api/estados/<estado_key>/columnas", methods=["POST"])
def create_configuracion_columna(estado_key):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para crear columnas de configuración"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    try:
        estado_id = parse_config_scope_id(estado_key)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    data = request.get_json(silent=True) or {}
    available_column_map = get_available_offer_column_map()

    try:
        columnas = data.get("columnas") if isinstance(data.get("columnas"), list) else None
        if columnas:
            columnas = [normalize_offer_column_name(normalize_required_text(column, "columna", 100)) for column in columnas]
            payload = {
                "columna": columnas[0],
                "descripcion_columna": normalize_optional_text(data.get("descripcion_columna"), 255),
                "orden_columna": normalize_optional_int(data.get("orden_columna"), "orden columna"),
            }
        else:
            payload = build_configuracion_columna_payload(data)
            payload["columna"] = normalize_offer_column_name(payload["columna"])
            columnas = [payload["columna"]]
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    if any(column is None for column in columnas):
        return jsonify({"success": False, "message": "Una o más columnas ya no existen en la estructura actual"}), 400

    invalid_columns = [column for column in columnas if column not in available_column_map]
    if invalid_columns:
        return jsonify({"success": False, "message": f"columnas no válidas: {', '.join(invalid_columns)}"}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            if not ensure_config_scope_exists(cursor, estado_id):
                conn.rollback()
                return jsonify({"success": False, "message": "Estado no encontrado"}), 404

            cursor.execute(
                "SELECT ISNULL(MAX(orden_columna), 0) FROM ofertas.configuracioncolumnas WHERE id_estado = ?",
                (estado_id,),
            )
            current_max_order = cursor.fetchone()[0] or 0

            inserted_configurations = []
            for index, column in enumerate(columnas):
                if configuracion_columna_exists(cursor, estado_id, column):
                    conn.rollback()
                    return jsonify({"success": False, "message": f"Ya existe una columna con ese nombre para el estado seleccionado: {column}"}), 409

                order_value = payload["orden_columna"] + index if payload["orden_columna"] is not None else current_max_order + index + 1
                description_value = payload["descripcion_columna"] if len(columnas) == 1 and payload["descripcion_columna"] else available_column_map[column]["label"]

                cursor.execute(
                    """
                    INSERT INTO ofertas.configuracioncolumnas (id_estado, columna, descripcion_columna, orden_columna)
                    OUTPUT INSERTED.id_config
                    VALUES (?, ?, ?, ?)
                    """,
                    (
                        estado_id,
                        column,
                        description_value,
                        order_value,
                    ),
                )
                inserted = cursor.fetchone()
                inserted_configurations.append(
                    {
                        "id_config": inserted[0] if inserted else None,
                        "id_estado": estado_id,
                        "columna": column,
                        "descripcion_columna": description_value,
                        "orden_columna": order_value,
                    }
                )

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "columna de configuración creada correctamente",
                "configuracion": inserted_configurations[0] if len(inserted_configurations) == 1 else None,
                "configuraciones": inserted_configurations,
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo crear la columna de configuración: {str(exc)}"}), 500


@app.route("/api/configuracion-columnas/<int:config_id>", methods=["PUT"])
def update_configuracion_columna(config_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para editar columnas de configuración"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}
    available_column_map = get_available_offer_column_map()

    try:
        payload = build_configuracion_columna_payload(data)
        payload = build_configuracion_columna_payload(data)
        payload["columna"] = normalize_offer_column_name(payload["columna"])
        id_estado = parse_config_scope_id(data.get("id_estado"))
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    if payload["columna"] is None:
        return jsonify({"success": False, "message": "La columna indicada ya no existe en la estructura actual"}), 400

    if payload["columna"] not in available_column_map:
        return jsonify({"success": False, "message": f"columna no válida: {payload['columna']}"}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            if not ensure_config_scope_exists(cursor, id_estado):
                conn.rollback()
                return jsonify({"success": False, "message": "Estado no encontrado"}), 404

            if configuracion_columna_exists(cursor, id_estado, payload["columna"], excluded_id=config_id):
                conn.rollback()
                return jsonify({"success": False, "message": "Ya existe una columna con ese nombre para el estado seleccionado"}), 409

            cursor.execute(
                """
                UPDATE ofertas.configuracioncolumnas
                SET id_estado = ?,
                    columna = ?,
                    descripcion_columna = ?,
                    orden_columna = ?
                WHERE id_config = ?
                """,
                (
                    id_estado,
                    payload["columna"],
                    payload["descripcion_columna"],
                    payload["orden_columna"],
                    config_id,
                ),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Configuración de columna no encontrada"}), 404

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "columna de configuración actualizada correctamente",
                "configuracion": {
                    "id_config": config_id,
                    "id_estado": id_estado,
                    "columna": payload["columna"],
                    "descripcion_columna": payload["descripcion_columna"],
                    "orden_columna": payload["orden_columna"],
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar la columna de configuración: {str(exc)}"}), 500


@app.route("/api/configuracion-columnas/<int:config_id>", methods=["DELETE"])
def delete_configuracion_columna(config_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para eliminar columnas de configuración"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            cursor.execute(
                "DELETE FROM ofertas.configuracioncolumnas WHERE id_config = ?",
                (config_id,),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Configuración de columna no encontrada"}), 404

            conn.commit()

        return jsonify({"success": True, "message": "columna de configuración eliminada correctamente"})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo eliminar la columna de configuración: {str(exc)}"}), 500


@app.route("/api/configuracion-columnas/reorder", methods=["POST"])
def reorder_configuracion_columnas():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para reordenar columnas de configuración"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    items = (request.get_json(silent=True) or {}).get("orden", [])
    if not isinstance(items, list) or not items:
        return jsonify({"success": False, "message": "Payload inválido: se espera {orden: [{id_config, orden_columna}]}"}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            for item in items:
                cursor.execute(
                    "UPDATE ofertas.configuracioncolumnas SET orden_columna = ? WHERE id_config = ?",
                    (int(item["orden_columna"]), int(item["id_config"])),
                )
            conn.commit()

        return jsonify({"success": True, "message": "orden de columnas actualizado"})
    except (KeyError, TypeError, ValueError) as exc:
        return jsonify({"success": False, "message": f"Datos inválidos: {str(exc)}"}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo reordenar la configuración de columnas: {str(exc)}"}), 500


# =====================================================
# ENDPOINTS PARA DEPARTAMENTOS
# =====================================================

@app.route("/api/departamentos", methods=["GET"])
def list_departamentos():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los departamentos"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    id_departamento,
                    nombre_departamento,
                    descripcion,
                    estado_activo,
                    fecha_creacion
                FROM ofertas.departamentos
                ORDER BY nombre_departamento ASC
                """
            )
            departamentos = [
                {
                    "id_departamento": row[0],
                    "nombre_departamento": row[1],
                    "descripcion": row[2],
                    "estado_activo": bool(row[3]),
                    "fecha_creacion": row[4].isoformat() if row[4] else None,
                }
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "departamentos": departamentos})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los departamentos: {str(exc)}"}), 500


@app.route("/api/departamentos", methods=["POST"])
def create_departamento():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para crear departamentos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        nombre = normalize_required_text(data.get("nombre_departamento"), "Nombre departamento", 255)
        descripcion = normalize_optional_text(data.get("descripcion"), 500)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            # Verificar si ya existe un departamento con ese nombre
            cursor.execute(
                "SELECT 1 FROM ofertas.departamentos WHERE LOWER(LTRIM(RTRIM(nombre_departamento))) = LOWER(LTRIM(RTRIM(?)))",
                (nombre,),
            )
            if cursor.fetchone() is not None:
                return jsonify({"success": False, "message": "Ya existe un departamento con ese nombre"}), 409

            cursor.execute(
                """
                INSERT INTO ofertas.departamentos (nombre_departamento, descripcion, estado_activo)
                OUTPUT INSERTED.id_departamento, INSERTED.fecha_creacion
                VALUES (?, ?, 1)
                """,
                (nombre, descripcion),
            )
            inserted = cursor.fetchone()
            inserted_id = inserted[0] if inserted else None
            fecha_creacion = inserted[1].isoformat() if inserted and inserted[1] else None
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Departamento creado correctamente",
                "departamento": {
                    "id_departamento": inserted_id,
                    "nombre_departamento": nombre,
                    "descripcion": descripcion,
                    "estado_activo": True,
                    "fecha_creacion": fecha_creacion,
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo crear el departamento: {str(exc)}"}), 500


@app.route("/api/departamentos/<int:departamento_id>", methods=["PUT"])
def update_departamento(departamento_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para editar departamentos"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        nombre = normalize_required_text(data.get("nombre_departamento"), "Nombre departamento", 255)
        descripcion = normalize_optional_text(data.get("descripcion"), 500)
        estado_activo = bool(data.get("estado_activo", True))
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            # Verificar si existe otro departamento con ese nombre
            cursor.execute(
                "SELECT 1 FROM ofertas.departamentos WHERE LOWER(LTRIM(RTRIM(nombre_departamento))) = LOWER(LTRIM(RTRIM(?))) AND id_departamento <> ?",
                (nombre, departamento_id),
            )
            if cursor.fetchone() is not None:
                return jsonify({"success": False, "message": "Ya existe otro departamento con ese nombre"}), 409

            cursor.execute(
                """
                UPDATE ofertas.departamentos
                SET nombre_departamento = ?, descripcion = ?, estado_activo = ?
                WHERE id_departamento = ?
                """,
                (nombre, descripcion, int(estado_activo), departamento_id),
            )

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "Departamento no encontrado"}), 404

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Departamento actualizado correctamente",
                "departamento": {
                    "id_departamento": departamento_id,
                    "nombre_departamento": nombre,
                    "descripcion": descripcion,
                    "estado_activo": estado_activo,
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el departamento: {str(exc)}"}), 500


# =====================================================
# ENDPOINTS PARA CONFIGURACION DE USUARIOS
# =====================================================

@app.route("/api/usuarios-config/<int:num_operario>", methods=["GET"])
def get_usuario_config(num_operario):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar la configuración del usuario"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            config = get_user_config(cursor, num_operario)

        return jsonify(
            {
                "success": True,
                "num_operario": num_operario,
                "email": config["email"],
                "id_rol": config["id_rol"],
                "rol": config["nombre_rol"],
                "departamento": config["departamento"],
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo consultar la configuración del usuario: {str(exc)}"}), 500


@app.route("/api/usuarios-config", methods=["POST"])
def upsert_usuario_config():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para guardar la configuración de usuario"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        num_operario = normalize_required_int(data.get("num_operario"), "Num_operario")
        email = normalize_optional_email(data.get("email"), "Email")
        id_departamento = normalize_optional_int(data.get("id_departamento"), "id_departamento")
        id_rol = normalize_required_int(data.get("id_rol"), "id_rol")
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            if id_departamento is not None:
                cursor.execute("SELECT 1 FROM ofertas.departamentos WHERE id_departamento = ?", (id_departamento,))
                if cursor.fetchone() is None:
                    return jsonify({"success": False, "message": "Departamento no encontrado"}), 404

            cursor.execute("SELECT 1 FROM ofertas.roles WHERE id_rol = ?", (id_rol,))
            if cursor.fetchone() is None:
                return jsonify({"success": False, "message": "Rol no encontrado"}), 404

            cursor.execute(
                """
                MERGE ofertas.usuarios_config AS target
                USING (SELECT ? AS Num_operario, ? AS Email, ? AS id_departamento, ? AS id_rol) AS source
                    ON target.Num_operario = source.Num_operario
                WHEN MATCHED THEN
                    UPDATE SET Email = source.Email, id_departamento = source.id_departamento, id_rol = source.id_rol
                WHEN NOT MATCHED THEN
                    INSERT (Num_operario, Email, id_departamento, id_rol)
                    VALUES (source.Num_operario, source.Email, source.id_departamento, source.id_rol);
                """,
                (num_operario, email, id_departamento, id_rol),
            )
            conn.commit()

        return jsonify({"success": True, "message": "Configuración de usuario guardada correctamente"})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo guardar la configuración del usuario: {str(exc)}"}), 500


@app.route("/api/roles", methods=["GET"])
def list_roles():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los roles"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT id_rol, nombre_rol, descripcion
                FROM ofertas.roles
                ORDER BY id_rol ASC
                """
            )
            roles = [
                {"id_rol": row[0], "nombre_rol": row[1], "descripcion": row[2]}
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "roles": roles})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los roles: {str(exc)}"}), 500


@app.route("/api/usuarios-general", methods=["GET"])
def list_usuarios_general():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los usuarios generales"}), 401

    search = normalize_optional_text(request.args.get("search"), 255)

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            query = """
                SELECT num_operario, nombre
                FROM general.usuarios
            """
            params = []

            if search:
                query += " WHERE CAST(num_operario AS VARCHAR(50)) LIKE ? OR nombre LIKE ?"
                params.extend([f"%{search}%", f"%{search}%"])

            query += " ORDER BY nombre ASC, num_operario ASC"
            cursor.execute(query, tuple(params))

            usuarios = [
                {
                    "num_operario": row[0],
                    "nombre": row[1],
                }
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "usuarios": usuarios})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los usuarios generales: {str(exc)}"}), 500


@app.route("/api/usuarios", methods=["GET"])
def list_usuarios():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los usuarios"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    ISNULL(u.id_usuario, 0) AS id_usuario,
                    uc.num_operario,
                    u.nombre,
                    uc.email,
                    uc.id_rol,
                    uc.id_departamento,
                    ISNULL(u.nivel_permisos, 0) AS nivel_permisos,
                    COALESCE(r.nombre_rol, 'Estandar') AS nombre_rol,
                    ISNULL(d.nombre_departamento, '') AS departamento
                FROM ofertas.usuarios_config uc
                INNER JOIN general.usuarios u
                    ON u.num_operario = uc.num_operario
                LEFT JOIN ofertas.roles r
                    ON r.id_rol = uc.id_rol
                LEFT JOIN ofertas.departamentos d
                    ON d.id_departamento = uc.id_departamento
                ORDER BY uc.num_operario ASC
                """
            )
            usuarios = [
                {
                    "id_usuario": row[0],
                    "num_operario": row[1],
                    "nombre": row[2],
                    "email": row[3],
                    "id_rol": row[4],
                    "id_departamento": row[5],
                    "nivel_permisos": row[6],
                    "rol": row[7],
                    "departamentos": row[8] or "",
                    "nombre_departamento": row[8] or "",
                }
                for row in cursor.fetchall()
            ]

        return jsonify({"success": True, "usuarios": usuarios})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los usuarios: {str(exc)}"}), 500


@app.route("/api/usuarios", methods=["POST"])
def create_usuario():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para crear usuarios"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_usuario_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            general_user = get_general_user(cursor, payload["num_operario"])
            if general_user is None:
                return jsonify({"success": False, "message": "El número de operario no existe en general.usuarios"}), 404

            if normalize_optional_text(general_user["nombre"], 255) != payload["nombre"]:
                return jsonify({"success": False, "message": "El nombre no coincide con el usuario de general.usuarios"}), 409

            role_id = payload["id_rol"]
            if role_id is None:
                role_id = get_role_id_by_name(cursor, payload["rol"])

            if role_id is None:
                return jsonify({"success": False, "message": "Rol no encontrado"}), 404

            cursor.execute("SELECT nombre_rol FROM ofertas.roles WHERE id_rol = ?", (role_id,))
            role_row = cursor.fetchone()
            if role_row is None:
                return jsonify({"success": False, "message": "Rol no encontrado"}), 404

            role_name = role_row[0]
            nivel_permisos = get_role_permission_level(role_name)

            if payload["id_departamento"] is not None:
                cursor.execute("SELECT 1 FROM ofertas.departamentos WHERE id_departamento = ?", (payload["id_departamento"],))
                if cursor.fetchone() is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "Departamento no encontrado"}), 404

            cursor.execute(
                """
                MERGE ofertas.usuarios_config AS target
                USING (SELECT ? AS num_operario, ? AS email, ? AS id_departamento, ? AS id_rol) AS source
                    ON target.num_operario = source.num_operario
                WHEN MATCHED THEN
                    UPDATE SET email = source.email, id_departamento = source.id_departamento, id_rol = source.id_rol
                WHEN NOT MATCHED THEN
                    INSERT (num_operario, email, id_departamento, id_rol)
                    VALUES (source.num_operario, source.email, source.id_departamento, source.id_rol);
                """,
                (payload["num_operario"], payload["email"], payload["id_departamento"], role_id),
            )

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Configuración de usuario guardada correctamente",
                "usuario": {
                    "id_usuario": general_user["id_usuario"],
                    "num_operario": payload["num_operario"],
                    "nombre": payload["nombre"],
                    "email": payload["email"],
                    "rol": role_name,
                    "nivel_permisos": nivel_permisos,
                },
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo crear el usuario: {str(exc)}"}), 500


@app.route("/api/usuarios/<int:num_operario>", methods=["PUT"])
def update_usuario(num_operario):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para actualizar usuarios"}), 401
    if is_read_only_user(user_data):
        return read_only_response()
    if not is_manager_user(user_data):
        return manager_only_response()

    data = request.get_json(silent=True) or {}
    data["num_operario"] = num_operario

    try:
        payload = build_usuario_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            general_user = get_general_user(cursor, payload["num_operario"])
            if general_user is None:
                return jsonify({"success": False, "message": "El número de operario no existe en general.usuarios"}), 404

            if normalize_optional_text(general_user["nombre"], 255) != payload["nombre"]:
                return jsonify({"success": False, "message": "El nombre no coincide con el usuario de general.usuarios"}), 409

            cursor.execute("SELECT 1 FROM ofertas.usuarios_config WHERE num_operario = ?", (payload["num_operario"],))
            if cursor.fetchone() is None:
                conn.rollback()
                return jsonify({"success": False, "message": "Usuario no encontrado"}), 404

            role_id = payload["id_rol"]
            if role_id is None:
                role_id = get_role_id_by_name(cursor, payload["rol"])

            if role_id is None:
                return jsonify({"success": False, "message": "Rol no encontrado"}), 404

            cursor.execute("SELECT nombre_rol FROM ofertas.roles WHERE id_rol = ?", (role_id,))
            role_row = cursor.fetchone()
            if role_row is None:
                return jsonify({"success": False, "message": "Rol no encontrado"}), 404

            if payload["id_departamento"] is not None:
                cursor.execute("SELECT 1 FROM ofertas.departamentos WHERE id_departamento = ?", (payload["id_departamento"],))
                if cursor.fetchone() is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "Departamento no encontrado"}), 404

            cursor.execute(
                """
                UPDATE ofertas.usuarios_config
                SET email = ?,
                    id_departamento = ?,
                    id_rol = ?
                WHERE num_operario = ?
                """,
                (payload["email"], payload["id_departamento"], role_id, payload["num_operario"]),
            )

            conn.commit()

        return jsonify({"success": True, "message": "Usuario actualizado correctamente"})
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar el usuario: {str(exc)}"}), 500


@app.route("/api/ofertas", methods=["POST"])
def create_oferta():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para guardar una oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_oferta_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_offer_bom_links_schema(cursor)

            if payload["id_bom"] is not None:
                cursor.execute(
                    "SELECT 1 FROM ofertas.materiales_precio WHERE id_material_precio = ?",
                    (payload["id_bom"],),
                )
                if cursor.fetchone() is None:
                    conn.rollback()
                    return jsonify({"success": False, "message": "BOM no encontrado"}), 404

            inserted_id, numero_oferta = insert_oferta_record(cursor, payload)

            if inserted_id is not None and payload["id_bom"] is not None:
                add_offer_bom_material_link(cursor, inserted_id, payload["id_bom"])

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Oferta guardada correctamente",
                "id_oferta": inserted_id,
                "numero_oferta": numero_oferta,
            }
        )
    except DuplicateOfertaError as exc:
        return jsonify({"success": False, "message": str(exc)}), 409
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo guardar la oferta: {str(exc)}"}), 500


def insert_oferta_record(cursor, payload):
    if oferta_email_subject_exists(
        cursor,
        payload["fecha_email"],
        payload["ref_cliente_asunto_email"],
    ):
        raise DuplicateOfertaError("Ya existe una oferta con la misma fecha de e-mail y el mismo asunto.")

    cursor.execute("CREATE TABLE #inserted_oferta (id_oferta INT)")
    cursor.execute(
        """
        INSERT INTO ofertas.listado_ofertas (
            fecha_email,
            fecha_alta_oferta,
            ref_cliente_asunto_email,
            id_cliente,
            id_bom,
            observaciones,
            nombre_emisor,
            email_emisor
        )
        OUTPUT INSERTED.id_oferta INTO #inserted_oferta
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            payload["fecha_email"],
            payload["fecha_alta_oferta"],
            payload["ref_cliente_asunto_email"],
            payload["id_cliente"],
            payload["id_bom"],
            payload["observaciones"],
            payload["nombre_emisor"],
            payload["email_emisor"],
        ),
    )
    cursor.execute("SELECT id_oferta FROM #inserted_oferta")
    inserted = cursor.fetchone()
    inserted_id = inserted[0] if inserted else None

    numero_oferta = None
    if inserted_id is not None:
        cursor.execute(
            "SELECT numero_oferta FROM ofertas.listado_ofertas WHERE id_oferta = ?",
            (inserted_id,),
        )
        numero_row = cursor.fetchone()
        numero_oferta = numero_row[0] if numero_row else None

    return inserted_id, numero_oferta


@app.route("/api/ofertas/<int:oferta_id>/boms", methods=["GET"])
def get_offer_boms(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para consultar los BOM de la oferta"}), 401

    try:
        with db_connection(autocommit=True) as conn:
            cursor = conn.cursor()
            ensure_offer_bom_links_schema(cursor)
            if not ensure_offer_exists(cursor, oferta_id):
                return jsonify({"success": False, "message": "Oferta no encontrada"}), 404

            boms = list_offer_bom_materials(cursor, oferta_id)

        return jsonify(
            {
                "success": True,
                "boms": boms,
                "nombre_bom": ", ".join(item["material"] for item in boms if item.get("material")),
            }
        )
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron consultar los BOM de la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/boms", methods=["POST"])
def add_offer_bom(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para modificar los BOM de la oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        material_id = normalize_required_int(data.get("id_material_precio") or data.get("id_bom"), "BOM")
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            result = add_offer_bom_material_link(cursor, oferta_id, material_id)
            selected_boms = list_offer_bom_materials(cursor, oferta_id)
            if result["added"]:
                actor_name = normalize_optional_text(
                    user_data.get("nombre") or user_data.get("display_name") or user_data.get("email"),
                    255,
                )
                insert_offer_interaction_entry(
                    cursor,
                    oferta_id,
                    "Edicion oferta",
                    datetime.now(),
                    observaciones=f"BOM añadido: {result['material']} (por {actor_name})",
                )
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "BOM añadido a la oferta" if result["added"] else "El BOM ya estaba asociado a la oferta",
                "boms": selected_boms,
                "nombre_bom": ", ".join(item["material"] for item in selected_boms if item.get("material")),
                "id_bom": selected_boms[0]["id_material_precio"] if selected_boms else None,
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 404
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo añadir el BOM a la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/boms/<int:material_id>/cantidad", methods=["PUT"])
def update_offer_bom_quantity(oferta_id, material_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para modificar la cantidad BOM"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}
    try:
        cantidad = int(data.get("cantidad", 1))
        if cantidad < 1:
            return jsonify({"success": False, "message": "La cantidad debe ser al menos 1"}), 400
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Cantidad inválida"}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            ensure_offer_bom_links_schema(cursor)

            cursor.execute(
                "UPDATE ofertas.oferta_bom_materiales SET cantidad = ? WHERE id_oferta = ? AND id_material_precio = ?",
                (cantidad, oferta_id, material_id),
            )
            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({"success": False, "message": "El BOM no está asociado a la oferta"}), 404
            conn.commit()

        return jsonify({"success": True, "message": "Cantidad actualizada", "cantidad": cantidad})
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo actualizar la cantidad: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/boms/<int:material_id>/precio-override", methods=["PUT"])
def set_offer_bom_price_override(oferta_id, material_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para modificar el precio BOM de la oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        precio_oferta = normalize_optional_decimal(data.get("precio"), "precio")
        if precio_oferta is None:
            raise ValueError("El nuevo precio es obligatorio")
        precio_oferta = precio_oferta.quantize(Decimal("0.01"))
        if precio_oferta < 0:
            raise ValueError("El precio BOM de la oferta no puede ser negativo")
        comentario = normalize_optional_text(data.get("comentario"), 500)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            result = upsert_offer_bom_material_price_override(
                cursor,
                oferta_id,
                material_id,
                precio_oferta,
                user_data=user_data,
                comentario=comentario,
            )
            selected_boms = list_offer_bom_materials(cursor, oferta_id)
            insert_offer_interaction_entry(
                cursor,
                oferta_id,
                "Edicion oferta",
                datetime.now(),
                observaciones=(
                    f"Precio BOM especifico {'establecido' if result['created'] else 'actualizado'}: "
                    f"{result['material']} ({serialize_decimal(result['precio_catalogo'])} -> {serialize_decimal(result['precio_oferta'])})"
                ),
            )
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Precio BOM especifico de la oferta guardado correctamente.",
                "boms": selected_boms,
                "nombre_bom": ", ".join(item["material"] for item in selected_boms if item.get("material")),
                "id_bom": selected_boms[0]["id_material_precio"] if selected_boms else None,
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 404
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo guardar el precio BOM de la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/boms/<int:material_id>/precio-override", methods=["DELETE"])
def clear_offer_bom_price_override(oferta_id, material_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para modificar el precio BOM de la oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            result = remove_offer_bom_material_price_override(cursor, oferta_id, material_id)
            selected_boms = list_offer_bom_materials(cursor, oferta_id)
            if result["removed"]:
                insert_offer_interaction_entry(
                    cursor,
                    oferta_id,
                    "Edicion oferta",
                    datetime.now(),
                    observaciones=(
                        f"Precio BOM especifico eliminado: {result['material']} "
                        f"(vuelve a catalogo {serialize_decimal(result['precio_catalogo'])})"
                    ),
                )
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "La oferta vuelve a usar el precio del catalogo." if result["removed"] else "El material ya usaba el precio del catalogo.",
                "boms": selected_boms,
                "nombre_bom": ", ".join(item["material"] for item in selected_boms if item.get("material")),
                "id_bom": selected_boms[0]["id_material_precio"] if selected_boms else None,
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 404
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo limpiar el precio BOM de la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/boms/<int:material_id>", methods=["DELETE"])
def remove_offer_bom(oferta_id, material_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para modificar los BOM de la oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            result = remove_offer_bom_material_link(cursor, oferta_id, material_id)
            selected_boms = list_offer_bom_materials(cursor, oferta_id)
            if result["removed"]:
                actor_name = normalize_optional_text(
                    user_data.get("nombre") or user_data.get("display_name") or user_data.get("email"),
                    255,
                )
                insert_offer_interaction_entry(
                    cursor,
                    oferta_id,
                    "Edicion oferta",
                    datetime.now(),
                    observaciones=f"BOM eliminado: {result['material']} (por {actor_name})",
                )
            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "BOM eliminado de la oferta" if result["removed"] else "El BOM no estaba asociado a la oferta",
                "boms": selected_boms,
                "nombre_bom": ", ".join(item["material"] for item in selected_boms if item.get("material")),
                "id_bom": selected_boms[0]["id_material_precio"] if selected_boms else None,
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 404
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo eliminar el BOM de la oferta: {str(exc)}"}), 500


@app.route("/api/ofertas/<int:oferta_id>/bom-to-etc", methods=["POST"])
def sync_offer_bom_to_etc(oferta_id):
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para guardar BOM en ETC"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()

            # Obtener materiales BOM de la oferta respetando overrides de precio por oferta
            cursor.execute(
                """
                SELECT mp.material,
                       mp.part_nr,
                       COALESCE(ao.precio_oferta, mp.precio) AS precio,
                       obm.cantidad
                FROM ofertas.oferta_bom_materiales obm
                INNER JOIN ofertas.materiales_precio mp
                    ON mp.id_material_precio = obm.id_material_precio
                LEFT JOIN ofertas.oferta_bom_precio_override ao
                    ON ao.id_oferta = obm.id_oferta
                   AND ao.id_material_precio = obm.id_material_precio
                   AND ao.activo = 1
                WHERE obm.id_oferta = ?
                ORDER BY obm.fecha_asignacion DESC, mp.material ASC
                """,
                (oferta_id,),
            )
            bom_rows = cursor.fetchall()

            now = datetime.now()

            if bom_rows:
                # Construir resumen y total
                material_summary = "\n".join(
                    f"• x{row[3]} [{row[1]}] {row[0]} — {row[2]:.2f}" if row[2] is not None else f"• x{row[3]} [{row[1]}] {row[0]}"
                    for row in bom_rows
                )
                total_material = sum((row[2] or 0) * (row[3] or 1) for row in bom_rows)
            else:
                material_summary = ""
                total_material = 0

            # Comprobar si existe registro ETC
            cursor.execute(
                "SELECT 1 FROM ofertas.oferta_etc WHERE id_oferta_etc = ?",
                (oferta_id,),
            )
            etc_exists = cursor.fetchone() is not None

            if etc_exists:
                cursor.execute(
                    """
                    UPDATE ofertas.oferta_etc
                    SET resumen_material_solicitado = ?,
                        total_material_eur = ?,
                        fecha_actualizacion = ?
                    WHERE id_oferta_etc = ?
                    """,
                    (material_summary or None, total_material or None, now, oferta_id),
                )
            else:
                cursor.execute("SET IDENTITY_INSERT ofertas.oferta_etc ON")
                try:
                    cursor.execute(
                        """
                        INSERT INTO ofertas.oferta_etc (
                            id_oferta_etc, id_estado, moneda, prioridad, es_urgente,
                            origen_registro, activo, fecha_creacion, fecha_actualizacion,
                            resumen_material_solicitado, total_material_eur
                        ) VALUES (?, 1, 'EUR', 'NORMAL', 0, 'MANUAL', 1, ?, ?, ?, ?)
                        """,
                        (oferta_id, now, now, material_summary or None, total_material or None),
                    )
                finally:
                    cursor.execute("SET IDENTITY_INSERT ofertas.oferta_etc OFF")

            conn.commit()

        return jsonify({
            "success": True,
            "message": "Materiales BOM volcados a ETC correctamente",
            "total_material_eur": total_material,
        })
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudieron guardar los materiales en ETC: {str(exc)}"}), 500


def insert_oferta_etc_record(cursor, payload, explicit_id=None):
    if not oferta_etc_table_exists(cursor):
        raise ValueError("La tabla ofertas.oferta_etc no existe todavía")

    payload = {**payload}
    payload["id_estado"] = resolve_active_estado_for_department(
        cursor,
        payload.get("id_departamento_destino"),
        fallback_estado_id=payload.get("id_estado"),
    )

    estado_metadata = get_estado_metadata(cursor, payload["id_estado"])
    if estado_metadata is None:
        raise ValueError("Estado no encontrado")
    if not estado_metadata["activo"]:
        raise ValueError("El estado seleccionado está inactivo y no se puede usar en procesos")

    if payload["id_cliente"] is not None:
        cursor.execute("SELECT 1 FROM ofertas.clientes WHERE id_cliente = ?", (payload["id_cliente"],))
        if cursor.fetchone() is None:
            raise ValueError("Cliente no encontrado")

    if payload["num_operario_responsable"] is not None:
        cursor.execute(
            "SELECT 1 FROM ofertas.usuarios_config WHERE num_operario = ?",
            (payload["num_operario_responsable"],),
        )
        if cursor.fetchone() is None:
            raise ValueError("Responsable no encontrado en ofertas.usuarios_config")

    if payload["id_departamento_destino"] is not None:
        cursor.execute(
            "SELECT 1 FROM ofertas.departamentos WHERE id_departamento = ?",
            (payload["id_departamento_destino"],),
        )
        if cursor.fetchone() is None:
            raise ValueError("Departamento no encontrado")

    if payload["proyecto"] is not None:
        cursor.execute(
            "SELECT 1 FROM ofertas.proyectos WHERE LTRIM(RTRIM(descripcion_proyecto)) = LTRIM(RTRIM(?))",
            (payload["proyecto"],),
        )
        if cursor.fetchone() is None:
            raise ValueError("proyecto no encontrado en la configuración")

    cursor.execute("CREATE TABLE #inserted_oferta_etc (id_oferta_etc INT)")
    insert_params = (
        payload["fecha_recepcion"],
        payload["fecha_envio_oferta"],
        payload["fecha_limite_respuesta"],
        payload["id_estado"],
        payload["id_cliente"],
        payload["num_operario_responsable"],
        payload["id_departamento_destino"],
        payload["codigo_externo_oferta"],
        payload["codigo_interno_oferta"],
        payload["referencia_cliente"],
        payload["numero_comision"],
        payload["po_original"],
        payload["pedido_b2b"],
        payload["proyecto"],
        payload["sales_orders"],
        payload["request_delivery_date"],
        payload["nombre_solicitante"],
        payload["email_solicitante"],
        payload["empresa_solicitante"],
        payload["incoterm"],
        payload["moneda"],
        payload["prioridad"],
        1 if payload["es_urgente"] else 0,
        payload["resumen_material_solicitado"],
        payload["resumen_material_ofertado"],
        payload["total_material_eur"],
        payload["total_fee_eur"],
        payload["observaciones_cliente"],
        payload["observaciones_tecnicas"],
        payload["observaciones_internas"],
        payload["origen_registro"],
        1 if payload["activo"] else 0,
    )

    if explicit_id is not None:
        cursor.execute("SET IDENTITY_INSERT ofertas.oferta_etc ON")
        try:
            cursor.execute(
                """
                INSERT INTO ofertas.oferta_etc (
                    id_oferta_etc,
                    fecha_recepcion,
                    fecha_envio_oferta,
                    fecha_limite_respuesta,
                    id_estado,
                    id_cliente,
                    num_operario_responsable,
                    id_departamento_destino,
                    codigo_externo_oferta,
                    codigo_interno_oferta,
                    referencia_cliente,
                    numero_comision,
                    po_original,
                    pedido_b2b,
                    proyecto,
                    sales_orders,
                    request_delivery_date,
                    nombre_solicitante,
                    email_solicitante,
                    empresa_solicitante,
                    incoterm,
                    moneda,
                    prioridad,
                    es_urgente,
                    resumen_material_solicitado,
                    resumen_material_ofertado,
                    total_material_eur,
                    total_fee_eur,
                    observaciones_cliente,
                    observaciones_tecnicas,
                    observaciones_internas,
                    origen_registro,
                    activo
                )
                OUTPUT INSERTED.id_oferta_etc INTO #inserted_oferta_etc
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (explicit_id, *insert_params),
            )
        finally:
            cursor.execute("SET IDENTITY_INSERT ofertas.oferta_etc OFF")
    else:
        cursor.execute(
            """
            INSERT INTO ofertas.oferta_etc (
                fecha_recepcion,
                fecha_envio_oferta,
                fecha_limite_respuesta,
                id_estado,
                id_cliente,
                num_operario_responsable,
                id_departamento_destino,
                codigo_externo_oferta,
                codigo_interno_oferta,
                referencia_cliente,
                numero_comision,
                po_original,
                pedido_b2b,
                proyecto,
                sales_orders,
                request_delivery_date,
                nombre_solicitante,
                email_solicitante,
                empresa_solicitante,
                incoterm,
                moneda,
                prioridad,
                es_urgente,
                resumen_material_solicitado,
                resumen_material_ofertado,
                total_material_eur,
                total_fee_eur,
                observaciones_cliente,
                observaciones_tecnicas,
                observaciones_internas,
                origen_registro,
                activo
            )
            OUTPUT INSERTED.id_oferta_etc INTO #inserted_oferta_etc
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            insert_params,
        )
    cursor.execute("SELECT id_oferta_etc FROM #inserted_oferta_etc")
    inserted = cursor.fetchone()
    inserted_id = inserted[0] if inserted else None
    if explicit_id is not None:
        sync_offer_estado_with_etc(cursor, explicit_id, payload["id_estado"])
    return inserted_id


def build_minimal_oferta_etc_payload(oferta_row, estado_department_id, responsable_num_operario=None):
    return {
        "fecha_recepcion": None,
        "fecha_envio_oferta": None,
        "fecha_limite_respuesta": None,
        "id_estado": oferta_row.get("id_estado") or 1,
        "id_cliente": oferta_row.get("id_cliente"),
        "num_operario_responsable": responsable_num_operario,
        "id_departamento_destino": estado_department_id,
        "codigo_externo_oferta": None,
        "codigo_interno_oferta": None,
        "referencia_cliente": oferta_row.get("numero_oferta"),
        "numero_comision": None,
        "po_original": None,
        "pedido_b2b": None,
        "proyecto": None,
        "sales_orders": oferta_row.get("sales_orders"),
        "request_delivery_date": oferta_row.get("request_delivery_date"),
        "nombre_solicitante": None,
        "email_solicitante": None,
        "empresa_solicitante": None,
        "incoterm": None,
        "moneda": "EUR",
        "prioridad": "NORMAL",
        "es_urgente": False,
        "resumen_material_solicitado": None,
        "resumen_material_ofertado": None,
        "total_material_eur": None,
        "total_fee_eur": None,
        "observaciones_cliente": None,
        "observaciones_tecnicas": None,
        "observaciones_internas": None,
        "origen_registro": "MANUAL",
        "activo": True,
    }


def ensure_oferta_etc_record(cursor, oferta_row, estado_department_id, responsable_num_operario=None):
    oferta_id = oferta_row.get("id_oferta")
    if oferta_id is None:
        raise ValueError("Oferta no encontrada para crear el registro ETC")

    cursor.execute("SELECT id_oferta_etc FROM ofertas.oferta_etc WHERE id_oferta_etc = ?", (oferta_id,))
    existing_row = cursor.fetchone()
    if existing_row is not None:
        return existing_row[0]

    payload = build_minimal_oferta_etc_payload(
        oferta_row,
        estado_department_id,
        responsable_num_operario=responsable_num_operario,
    )
    return insert_oferta_etc_record(cursor, payload, explicit_id=oferta_id)


@app.route("/api/ofertas-completa", methods=["POST"])
def create_oferta_completa():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para guardar una oferta"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}
    oferta_data = data.get("oferta") or {}
    oferta_etc_data = data.get("oferta_etc") or {}
    imported_email_attachment_token = data.get("imported_email_attachment_token") or oferta_data.get("imported_email_attachment_token")
    imported_email_metadata = data.get("imported_email_metadata") or oferta_data.get("imported_email_metadata")

    try:
        oferta_payload = build_oferta_payload(oferta_data)
        oferta_etc_payload = build_oferta_etc_payload(oferta_etc_data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    inserted_id = None
    staged_attachment_payloads = []
    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            inserted_id, numero_oferta = insert_oferta_record(cursor, oferta_payload)
            inserted_etc_id = insert_oferta_etc_record(cursor, oferta_etc_payload, explicit_id=inserted_id)
            if imported_email_attachment_token:
                staged_attachment_payloads = move_staged_imported_email_attachments_to_offer(
                    inserted_id,
                    imported_email_attachment_token,
                    numero_oferta=numero_oferta,
                )
            conn.commit()

            # --- Notificar a los managers del departamento del estado de la oferta ---
            notification_result = None
            try:
                with db_connection(autocommit=True) as notify_conn:
                    notify_cursor = notify_conn.cursor()
                    notify_cursor.execute(
                        "SELECT id_estado FROM ofertas.listado_ofertas WHERE id_oferta = ?",
                        (inserted_id,),
                    )
                    estado_row = notify_cursor.fetchone()
                    if estado_row and estado_row[0] is not None:
                        notification_payload = build_new_offer_notification(
                            notify_cursor,
                            inserted_id,
                            estado_row[0],
                        )
                        notification_result = send_estado_manager_notification(notification_payload)
            except Exception:
                app.logger.exception(
                    "No se pudo enviar la notificacion de nueva oferta para id_oferta=%s",
                    inserted_id,
                )

            return jsonify(
                {
                    "success": True,
                    "message": "Oferta y ETC guardados correctamente",
                    "id_oferta": inserted_id,
                    "numero_oferta": numero_oferta,
                    "id_oferta_etc": inserted_etc_id,
                    "adjuntos": staged_attachment_payloads,
                    "notification": notification_result,
                }
            )
    except DuplicateOfertaError as exc:
        return jsonify({"success": False, "message": str(exc)}), 409
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except pyodbc.IntegrityError as exc:
        return jsonify({"success": False, "message": f"No se pudo guardar la oferta completa por una restricción de datos: {str(exc)}"}), 409
    except Exception as exc:
        if inserted_id and staged_attachment_payloads:
            cleanup_offer_attachment_entries(inserted_id, [item.get("stored_name") for item in staged_attachment_payloads if item.get("stored_name")])
        return jsonify({"success": False, "message": f"No se pudo guardar la oferta completa: {str(exc)}"}), 500
@app.route("/api/ofertas-etc", methods=["POST"])
def create_oferta_etc():
    user_data = get_logged_user_data()
    if not user_data:
        return jsonify({"success": False, "message": "Debes iniciar sesión para guardar una oferta ETC"}), 401
    if is_read_only_user(user_data):
        return read_only_response()

    data = request.get_json(silent=True) or {}

    try:
        payload = build_oferta_etc_payload(data)
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    try:
        with db_connection(autocommit=False) as conn:
            cursor = conn.cursor()
            inserted_id = insert_oferta_etc_record(cursor, payload)

            conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "Oferta ETC guardada correctamente",
                "id_oferta_etc": inserted_id,
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "message": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
    except pyodbc.IntegrityError as exc:
        return jsonify({"success": False, "message": f"No se pudo guardar la oferta ETC por una restricción de datos: {str(exc)}"}), 409
    except Exception as exc:
        return jsonify({"success": False, "message": f"No se pudo guardar la oferta ETC: {str(exc)}"}), 500


def create_app():
    return app


def resolve_ssl_file(file_name):
    if not file_name:
        return None

    normalized_name = str(file_name).strip()
    if not normalized_name:
        return None

    if os.path.isabs(normalized_name):
        return normalized_name if os.path.exists(normalized_name) else None

    candidate_dirs = []
    candidate_dirs.append(PROJECT_DIR)
    if getattr(sys, "frozen", False):
        exe_dir = os.path.dirname(sys.executable)
        candidate_dirs.append(exe_dir)
        parent_dir = os.path.dirname(exe_dir)
        if parent_dir and parent_dir != exe_dir:
            candidate_dirs.append(parent_dir)

    seen = set()
    for base_dir in candidate_dirs:
        if not base_dir:
            continue
        candidate_path = os.path.normpath(os.path.join(base_dir, normalized_name))
        if candidate_path in seen:
            continue
        seen.add(candidate_path)
        if os.path.exists(candidate_path):
            return candidate_path
    return None


def get_runtime_ssl_context(use_https):
    if not use_https:
        return None

    cert_path = resolve_ssl_file(app.config.get("SSL_CERT_FILE"))
    key_path = resolve_ssl_file(app.config.get("SSL_KEY_FILE"))
    if cert_path and os.path.basename(cert_path).lower() == "cert.pem":
        cert_chain_candidate = os.path.join(os.path.dirname(cert_path), "cert_chain.pem")
        if os.path.exists(cert_chain_candidate):
            cert_path = cert_chain_candidate
    if cert_path and key_path:
        return (cert_path, key_path)

    if app.config.get("AUTO_GENERATE_SSL_CERT", True):
        return "adhoc"

    missing_files = []
    if not cert_path:
        missing_files.append(str(app.config.get("SSL_CERT_FILE") or "cert.pem"))
    if not key_path:
        missing_files.append(str(app.config.get("SSL_KEY_FILE") or "key.pem"))
    raise RuntimeError(
        "Falta configurar el certificado HTTPS del servidor. "
        f"No se encontraron: {', '.join(missing_files)}. "
        "Copia cert.pem y key.pem junto al EXE o define SSL_CERT_FILE/SSL_KEY_FILE."
    )


if __name__ == "__main__":
    redirect_uri = get_env_value("OAUTH_REDIRECT_URI", "AZURE_REDIRECT_URI", "MICROSOFT_REDIRECT_URI", "OUTLOOK_REDIRECT_URI", default="")
    parsed_redirect = urlparse(redirect_uri) if redirect_uri else None
    use_https = parsed_redirect is not None and parsed_redirect.scheme.lower() == "https"
    ssl_context = get_runtime_ssl_context(use_https)
    app.run(
        host="0.0.0.0",
        port=app.config["APP_PORT"],
        debug=app.config["DEBUG"],
        ssl_context=ssl_context,
    )
